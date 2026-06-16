"""
数据导入导出模块
支持Excel/CSV格式的数据导入导出
"""

from typing import List, Dict, Any, Optional
from datetime import datetime
import csv
import os


class DataExportService:
    """数据导出服务"""
    
    def __init__(self, task_model, contact_model):
        self.task_model = task_model
        self.contact_model = contact_model
        
    def export_tasks_to_csv(self, tasks: List[Dict[str, Any]], filepath: str) -> bool:
        """导出任务到CSV"""
        if not tasks:
            return False
            
        try:
            # CSV表头
            headers = [
                '任务ID', '任务名称', '状态', '重要程度', '责任人',
                '关键模块', '任务内容', '咨询者姓名', '咨询者部门',
                '预期答复时间', '创建时间', '产品型号', '行业', '备注'
            ]
            
            with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                
                for task in tasks:
                    row = [
                        task.get('task_id', ''),
                        task.get('task_name', ''),
                        task.get('status', ''),
                        task.get('importance', ''),
                        task.get('responder_name', ''),
                        task.get('key_module', ''),
                        task.get('task_content', ''),
                        task.get('consultant_name', ''),
                        task.get('consultant_dept', ''),
                        str(task.get('expected_time', ''))[:19] if task.get('expected_time') else '',
                        str(task.get('task_time', ''))[:19] if task.get('task_time') else '',
                        task.get('product_model', ''),
                        task.get('industry', ''),
                        task.get('remarks', '')
                    ]
                    writer.writerow(row)
                    
            return True
        except Exception:
            return False
            
    def export_tasks_to_excel(self, tasks: List[Dict[str, Any]], filepath: str) -> bool:
        """导出任务到Excel（需要openpyxl）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "任务列表"
            
            # 表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入表头
            headers = [
                '序号', '任务ID', '任务名称', '状态', '重要程度', '责任人',
                '关键模块', '咨询者姓名', '咨询者部门', '预期答复时间', '创建时间'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
            # 写入数据
            for row, task in enumerate(tasks, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=task.get('task_id', ''))
                ws.cell(row=row, column=3, value=task.get('task_name', ''))
                ws.cell(row=row, column=4, value=task.get('status', ''))
                ws.cell(row=row, column=5, value=task.get('importance', ''))
                ws.cell(row=row, column=6, value=task.get('responder_name', ''))
                ws.cell(row=row, column=7, value=task.get('key_module', ''))
                ws.cell(row=row, column=8, value=task.get('consultant_name', ''))
                ws.cell(row=row, column=9, value=task.get('consultant_dept', ''))
                ws.cell(row=row, column=10, value=str(task.get('expected_time', ''))[:19] if task.get('expected_time') else '')
                ws.cell(row=row, column=11, value=str(task.get('task_time', ''))[:19] if task.get('task_time') else '')
                
            # 调整列宽
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
                
            wb.save(filepath)
            return True
        except ImportError:
            # 如果没有openpyxl，降级为CSV
            return self.export_tasks_to_csv(tasks, filepath.replace('.xlsx', '.csv'))
        except Exception:
            return False
            
    def generate_export_filename(self, export_type: str) -> str:
        """生成导出文件名"""
        now = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"任务_{export_type}_{now}"


class DataImportService:
    """数据导入服务"""
    
    def __init__(self, contact_model, recommendation_model):
        self.contact_model = contact_model
        self.recommendation_model = recommendation_model
        
    def import_contacts_from_csv(self, filepath: str) -> tuple:
        """
        从CSV导入通讯录
        返回: (成功数量, 总数量, 错误列表)
        """
        imported = 0
        errors = []
        
        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                # 尝试多种分隔符
                try:
                    reader = csv.DictReader(f)
                except:
                    f.seek(0)
                    reader = csv.DictReader(f, delimiter=';')
                    
                for row_num, row in enumerate(reader, 2):
                    try:
                        contact = {
                            'name': row.get('姓名', row.get('name', '')).strip(),
                            'employee_id': row.get('工号', row.get('employee_id', '')).strip(),
                            'phone': row.get('手机号', row.get('phone', '')).strip(),
                            'email': row.get('邮箱', row.get('email', '')).strip(),
                            'department': row.get('部门', row.get('department', '')).strip(),
                            'position': row.get('职位', row.get('position', '')).strip()
                        }
                        
                        if contact['name']:
                            self.contact_model.add_contact(contact)
                            imported += 1
                        else:
                            errors.append(f"行{row_num}: 姓名为空")
                    except Exception as e:
                        errors.append(f"行{row_num}: {str(e)}")
                        
        except Exception as e:
            return 0, 0, [f"读取文件失败: {str(e)}"]
            
        return imported, imported + len(errors), errors
        
    def import_contacts_from_text(self, content: str) -> tuple:
        """
        从文本导入通讯录（支持多种格式）
        格式1: 姓名/工号
        格式2: 姓名+手机号
        格式3: 姓名+邮箱
        返回: (成功数量, 总数量, 错误列表)
        """
        imported = 0
        errors = []
        lines = content.strip().split('\n')
        
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if not line:
                continue
                
            try:
                contact = {'name': '', 'employee_id': '', 'phone': '', 'email': '', 'department': '', 'position': ''}
                
                # 格式: 姓名/工号
                if '/' in line:
                    parts = line.split('/')
                    contact['name'] = parts[0].strip()
                    if len(parts) > 1:
                        contact['employee_id'] = parts[1].strip()
                        
                # 格式: 姓名+手机号（手机号11位数字）
                elif any(len(p) == 11 and p.isdigit() for p in line.split()):
                    for part in line.split():
                        if len(part) == 11 and part.isdigit():
                            contact['phone'] = part
                        elif '@' in part:
                            contact['email'] = part
                        elif not contact['name']:
                            contact['name'] = part
                            
                # 格式: 姓名+邮箱
                elif '@' in line:
                    for part in line.split():
                        if '@' in part:
                            contact['email'] = part
                        elif not contact['name']:
                            contact['name'] = part
                            
                # 纯姓名
                else:
                    contact['name'] = line
                    
                if contact['name']:
                    self.contact_model.add_contact(contact)
                    imported += 1
                else:
                    errors.append(f"行{line_num}: 无法解析 - {line}")
                    
            except Exception as e:
                errors.append(f"行{line_num}: {str(e)}")
                
        return imported, len(lines), errors
        
    def import_recommendations_from_csv(self, filepath: str) -> tuple:
        """
        从CSV导入推荐库
        格式: 姓名，关键模块[, 其他字段...]
        返回: (成功数量, 总数量, 错误列表)
        """
        imported = 0
        errors = []
        
        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                
                for row_num, row in enumerate(reader, 2):
                    try:
                        if len(row) < 2:
                            errors.append(f"行{row_num}: 数据不完整")
                            continue
                            
                        recommendation = {
                            'name': row[0].strip(),
                            'key_module': row[1].strip(),
                            'employee_id': row[2].strip() if len(row) > 2 else '',
                            'phone': row[3].strip() if len(row) > 3 else '',
                            'email': row[4].strip() if len(row) > 4 else '',
                            'department': row[5].strip() if len(row) > 5 else '',
                            'position': row[6].strip() if len(row) > 6 else '',
                            'expertise': row[7].strip() if len(row) > 7 else ''
                        }
                        
                        if recommendation['name'] and recommendation['key_module']:
                            self.recommendation_model.add_recommendation(recommendation)
                            imported += 1
                        else:
                            errors.append(f"行{row_num}: 姓名或关键模块为空")
                    except Exception as e:
                        errors.append(f"行{row_num}: {str(e)}")
                        
        except Exception as e:
            return 0, 0, [f"读取文件失败: {str(e)}"]
            
        return imported, imported + len(errors), errors


class DataValidationService:
    """数据校验服务"""
    
    @staticmethod
    def validate_task_data(data: Dict[str, Any]) -> tuple:
        """
        校验任务数据
        返回: (是否有效, 错误列表)
        """
        errors = []
        
        # 必填字段检查
        required_fields = ['task_name', 'consultant_name']
        for field in required_fields:
            if not data.get(field):
                errors.append(f"必填字段 {field} 不能为空")
                
        # 任务名称长度检查
        if data.get('task_name') and len(data['task_name']) > 100:
            errors.append("任务名称不能超过100个字符")
            
        # 联系方式格式检查
        contact = data.get('consultant_contact', '')
        if contact:
            if '@' in contact and '.' not in contact.split('@')[-1]:
                errors.append("邮箱格式不正确")
            elif contact.isdigit() and len(contact) != 11:
                errors.append("手机号应为11位数字")
                
        return len(errors) == 0, errors
        
    @staticmethod
    def validate_contact_data(data: Dict[str, Any]) -> tuple:
        """
        校验通讯录数据
        返回: (是否有效, 错误列表)
        """
        errors = []
        
        # 姓名检查
        if not data.get('name'):
            errors.append("姓名不能为空")
            
        # 邮箱格式检查
        email = data.get('email', '')
        if email and '@' not in email:
            errors.append("邮箱格式不正确")
            
        # 手机号格式检查
        phone = data.get('phone', '')
        if phone and (not phone.isdigit() or len(phone) not in [7, 11]):
            errors.append("手机号格式不正确")
            
        return len(errors) == 0, errors
