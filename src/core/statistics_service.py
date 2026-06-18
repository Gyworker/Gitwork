# -*- coding: utf-8 -*-
"""
统计分析服务模块
提供完整的数据统计分析功能

版本: V1.0
功能:
- 任务统计（状态分布、重要程度分布）
- 责任人统计（任务数量、进行中数量）
- 关键模块统计（模块分布、热度排行）
- 时间趋势分析（按日/周/月统计）
- 导出统计报告
"""

import os
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from collections import defaultdict

from src.core.logger import get_logger
from src.database.connection import get_db_connection

logger = get_logger(__name__)


@dataclass
class TaskSummary:
    """任务统计摘要"""
    total: int = 0
    by_status: Dict[str, int] = None
    by_importance: Dict[str, int] = None
    today_created: int = 0
    today_completed: int = 0
    overdue: int = 0
    
    def __post_init__(self):
        if self.by_status is None:
            self.by_status = {}
        if self.by_importance is None:
            self.by_importance = {}
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'total': self.total,
            'by_status': self.by_status,
            'by_importance': self.by_importance,
            'today_created': self.today_created,
            'today_completed': self.today_completed,
            'overdue': self.overdue
        }


@dataclass
class ResponderStat:
    """责任人统计"""
    name: str = ""
    total_tasks: int = 0
    in_progress: int = 0
    completed: int = 0
    avg_duration: float = 0.0
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'responder': self.name,
            'total_tasks': self.total_tasks,
            'in_progress': self.in_progress,
            'completed': self.completed,
            'avg_duration': round(self.avg_duration, 1)
        }


@dataclass
class ModuleStat:
    """关键模块统计"""
    module: str = ""
    count: int = 0
    percentage: float = 0.0
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'module': self.module,
            'count': self.count,
            'percentage': round(self.percentage, 1)
        }


@dataclass
class TrendData:
    """趋势数据"""
    date: str = ""
    created: int = 0
    completed: int = 0
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'date': self.date,
            'created': self.created,
            'completed': self.completed
        }


class StatisticsService:
    """
    统计分析服务
    
    功能:
    1. 任务统计 - 总数、状态分布、重要程度分布
    2. 责任人统计 - 任务数量排行、处理效率
    3. 关键模块统计 - 模块分布、热度排行
    4. 时间趋势 - 按日/周/月统计任务变化
    5. 效率分析 - 平均处理时长、完成率
    6. 导出报告 - 生成文本/JSON报告
    """
    
    def __init__(self, db_path: str = None):
        """
        初始化统计服务
        
        Args:
            db_path: 数据库路径
        """
        self.db_path = db_path
        self._db = None
    
    @property
    def db(self):
        """获取数据库连接"""
        if self._db is None:
            self._db = get_db_connection()
        return self._db
    
    def get_task_summary(self) -> Dict[str, Any]:
        """
        获取任务统计摘要
        
        Returns:
            Dict: 统计结果
        """
        try:
            summary = TaskSummary()
            
            # 总任务数
            result = self.db.execute_query("SELECT COUNT(*) as cnt FROM tasks")
            summary.total = result[0]['cnt'] if result else 0
            
            # 按状态统计
            result = self.db.execute_query(
                "SELECT status, COUNT(*) as cnt FROM tasks GROUP BY status"
            )
            summary.by_status = {r['status']: r['cnt'] for r in result}
            
            # 按重要程度统计
            result = self.db.execute_query(
                "SELECT importance, COUNT(*) as cnt FROM tasks GROUP BY importance"
            )
            summary.by_importance = {r['importance']: r['cnt'] for r in result}
            
            # 今日新增
            today = datetime.now().strftime('%Y-%m-%d')
            result = self.db.execute_query(
                "SELECT COUNT(*) as cnt FROM tasks WHERE DATE(created_at) = ?",
                (today,)
            )
            summary.today_created = result[0]['cnt'] if result else 0
            
            # 今日完成
            result = self.db.execute_query(
                "SELECT COUNT(*) as cnt FROM tasks WHERE DATE(updated_at) = ? AND status = '完成'",
                (today,)
            )
            summary.today_completed = result[0]['cnt'] if result else 0
            
            # 逾期任务
            result = self.db.execute_query(
                """
                SELECT COUNT(*) as cnt FROM tasks 
                WHERE status NOT IN ('完成', '已答复') 
                AND expected_time < datetime('now')
                """
            )
            summary.overdue = result[0]['cnt'] if result else 0
            
            return summary.to_dict()
            
        except Exception as e:
            logger.error(f"获取任务摘要失败: {e}")
            return TaskSummary().to_dict()
    
    def get_status_distribution(self) -> List[Dict[str, Any]]:
        """
        获取任务状态分布
        
        Returns:
            List: 状态分布列表
        """
        try:
            total = self.db.execute_query("SELECT COUNT(*) as cnt FROM tasks")
            total_count = total[0]['cnt'] if total else 1
            
            result = self.db.execute_query(
                "SELECT status, COUNT(*) as cnt FROM tasks GROUP BY status ORDER BY cnt DESC"
            )
            
            return [
                {
                    'status': r['status'],
                    'count': r['cnt'],
                    'percentage': round(r['cnt'] / total_count * 100, 1)
                }
                for r in result
            ]
            
        except Exception as e:
            logger.error(f"获取状态分布失败: {e}")
            return []
    
    def get_importance_distribution(self) -> List[Dict[str, Any]]:
        """
        获取重要程度分布
        
        Returns:
            List: 重要程度分布列表
        """
        try:
            total = self.db.execute_query("SELECT COUNT(*) as cnt FROM tasks")
            total_count = total[0]['cnt'] if total else 1
            
            result = self.db.execute_query(
                "SELECT importance, COUNT(*) as cnt FROM tasks GROUP BY importance ORDER BY cnt DESC"
            )
            
            return [
                {
                    'importance': r['importance'],
                    'count': r['cnt'],
                    'percentage': round(r['cnt'] / total_count * 100, 1)
                }
                for r in result
            ]
            
        except Exception as e:
            logger.error(f"获取重要程度分布失败: {e}")
            return []
    
    def get_responder_stats(self, limit: int = 10) -> List[Dict[str, Any]]:
        """
        获取责任人统计
        
        Args:
            limit: 返回数量限制
            
        Returns:
            List: 责任人统计列表
        """
        try:
            result = self.db.execute_query(
                """
                SELECT 
                    responder as name,
                    COUNT(*) as total_tasks,
                    SUM(CASE WHEN status = '进行中' THEN 1 ELSE 0 END) as in_progress,
                    SUM(CASE WHEN status = '完成' THEN 1 ELSE 0 END) as completed
                FROM tasks
                WHERE responder IS NOT NULL AND responder != ''
                GROUP BY responder
                ORDER BY total_tasks DESC
                LIMIT ?
                """,
                (limit,)
            )
            
            stats = []
            for r in result:
                stat = ResponderStat(
                    name=r['name'],
                    total_tasks=r['total_tasks'],
                    in_progress=r['in_progress'],
                    completed=r['completed']
                )
                
                # 计算平均处理时长
                if stat.completed > 0:
                    duration_result = self.db.execute_query(
                        """
                        SELECT AVG(
                            julianday(updated_at) - julianday(created_at)
                        ) as avg_days
                        FROM tasks
                        WHERE responder = ? AND status = '完成'
                        """,
                        (r['name'],)
                    )
                    if duration_result:
                        stat.avg_duration = (duration_result[0]['avg_days'] or 0) * 24  # 转换为小时
                
                stats.append(stat.to_dict())
            
            return stats
            
        except Exception as e:
            logger.error(f"获取责任人统计失败: {e}")
            return []
    
    def get_module_stats(self, limit: int = 10) -> List[Dict[str, Any]]:
        """
        获取关键模块统计
        
        Args:
            limit: 返回数量限制
            
        Returns:
            List: 模块统计列表
        """
        try:
            # 统计每个模块的任务数量
            result = self.db.execute_query(
                """
                SELECT key_module, COUNT(*) as cnt
                FROM tasks
                WHERE key_module IS NOT NULL AND key_module != ''
                GROUP BY key_module
                ORDER BY cnt DESC
                LIMIT ?
                """,
                (limit,)
            )
            
            total = sum(r['cnt'] for r in result)
            
            return [
                {
                    'module': r['key_module'],
                    'count': r['cnt'],
                    'percentage': round(r['cnt'] / total * 100, 1) if total > 0 else 0
                }
                for r in result
            ]
            
        except Exception as e:
            logger.error(f"获取模块统计失败: {e}")
            return []
    
    def get_trend_data(
        self,
        days: int = 30,
        granularity: str = 'day'
    ) -> List[Dict[str, Any]]:
        """
        获取趋势数据
        
        Args:
            days: 统计天数
            granularity: 粒度 ('day', 'week', 'month')
            
        Returns:
            List: 趋势数据列表
        """
        try:
            if granularity == 'day':
                return self._get_daily_trend(days)
            elif granularity == 'week':
                return self._get_weekly_trend(days)
            elif granularity == 'month':
                return self._get_monthly_trend(days)
            else:
                return []
                
        except Exception as e:
            logger.error(f"获取趋势数据失败: {e}")
            return []
    
    def _get_daily_trend(self, days: int) -> List[Dict[str, Any]]:
        """获取每日趋势"""
        result = self.db.execute_query(
            """
            SELECT 
                DATE(created_at) as date,
                COUNT(*) as created
            FROM tasks
            WHERE created_at >= datetime('now', ?)
            GROUP BY DATE(created_at)
            ORDER BY date
            """,
            (f'-{days} days',)
        )
        
        # 获取每日完成数
        completed = self.db.execute_query(
            """
            SELECT 
                DATE(updated_at) as date,
                COUNT(*) as completed
            FROM tasks
            WHERE updated_at >= datetime('now', ?) AND status = '完成'
            GROUP BY DATE(updated_at)
            ORDER BY date
            """,
            (f'-{days} days',)
        )
        
        completed_map = {r['date']: r['completed'] for r in completed}
        
        return [
            {
                'date': r['date'],
                'created': r['created'],
                'completed': completed_map.get(r['date'], 0)
            }
            for r in result
        ]
    
    def _get_weekly_trend(self, weeks: int) -> List[Dict[str, Any]]:
        """获取每周趋势"""
        result = self.db.execute_query(
            """
            SELECT 
                strftime('%Y-W%W', created_at) as week,
                COUNT(*) as created
            FROM tasks
            WHERE created_at >= datetime('now', ?)
            GROUP BY week
            ORDER BY week
            """,
            (f'-{weeks * 7} days',)
        )
        
        return [
            {
                'date': r['week'],
                'created': r['created'],
                'completed': 0
            }
            for r in result
        ]
    
    def _get_monthly_trend(self, months: int) -> List[Dict[str, Any]]:
        """获取每月趋势"""
        result = self.db.execute_query(
            """
            SELECT 
                strftime('%Y-%m', created_at) as month,
                COUNT(*) as created
            FROM tasks
            WHERE created_at >= datetime('now', ?)
            GROUP BY month
            ORDER BY month
            """,
            (f'-{months} months',)
        )
        
        return [
            {
                'date': r['month'],
                'created': r['created'],
                'completed': 0
            }
            for r in result
        ]
    
    def get_efficiency_analysis(self) -> Dict[str, Any]:
        """
        获取效率分析
        
        Returns:
            Dict: 效率分析结果
        """
        try:
            # 平均处理时长
            avg_duration = self.db.execute_query(
                """
                SELECT AVG(
                    julianday(CASE 
                        WHEN status = '完成' THEN updated_at 
                        ELSE datetime('now') 
                    END) - julianday(created_at)
                ) as avg_days
                FROM tasks
                """
            )
            
            avg_days = avg_duration[0]['avg_days'] if avg_duration and avg_duration[0]['avg_days'] else 0
            
            # 完成率
            total = self.db.execute_query("SELECT COUNT(*) as cnt FROM tasks")
            completed = self.db.execute_query(
                "SELECT COUNT(*) as cnt FROM tasks WHERE status = '完成'"
            )
            
            total_count = total[0]['cnt'] if total else 1
            completed_count = completed[0]['cnt'] if completed else 0
            completion_rate = completed_count / total_count * 100 if total_count > 0 else 0
            
            # 平均响应时长
            avg_response = self.db.execute_query(
                """
                SELECT AVG(
                    julianday(CASE 
                        WHEN status IN ('已答复', '完成') THEN updated_at 
                        ELSE datetime('now') 
                    END) - julianday(created_at)
                ) as avg_days
                FROM tasks
                """
            )
            
            return {
                'avg_duration_hours': round(avg_days * 24, 1),
                'avg_response_hours': round(
                    (avg_response[0]['avg_days'] or 0) * 24, 1
                ),
                'completion_rate': round(completion_rate, 1),
                'total_tasks': total_count,
                'completed_tasks': completed_count,
                'in_progress_tasks': total_count - completed_count
            }
            
        except Exception as e:
            logger.error(f"获取效率分析失败: {e}")
            return {}
    
    def get_department_stats(self) -> List[Dict[str, Any]]:
        """
        获取部门统计
        
        Returns:
            List: 部门统计列表
        """
        try:
            result = self.db.execute_query(
                """
                SELECT 
                    department,
                    COUNT(*) as total,
                    SUM(CASE WHEN status = '完成' THEN 1 ELSE 0 END) as completed
                FROM tasks
                WHERE department IS NOT NULL AND department != ''
                GROUP BY department
                ORDER BY total DESC
                """
            )
            
            total_tasks = sum(r['total'] for r in result)
            
            return [
                {
                    'department': r['department'],
                    'total': r['total'],
                    'completed': r['completed'],
                    'in_progress': r['total'] - r['completed'],
                    'percentage': round(r['total'] / total_tasks * 100, 1) if total_tasks > 0 else 0
                }
                for r in result
            ]
            
        except Exception as e:
            logger.error(f"获取部门统计失败: {e}")
            return []
    
    def get_consultant_stats(self) -> List[Dict[str, Any]]:
        """
        获取咨询者统计
        
        Returns:
            List: 咨询者统计列表
        """
        try:
            result = self.db.execute_query(
                """
                SELECT 
                    consultant_name,
                    consultant_contact,
                    COUNT(*) as task_count
                FROM tasks
                WHERE consultant_name IS NOT NULL AND consultant_name != ''
                GROUP BY consultant_name
                ORDER BY task_count DESC
                LIMIT 20
                """
            )
            
            return [
                {
                    'name': r['consultant_name'],
                    'contact': r['consultant_contact'] or '',
                    'task_count': r['task_count']
                }
                for r in result
            ]
            
        except Exception as e:
            logger.error(f"获取咨询者统计失败: {e}")
            return []
    
    def generate_summary_report(self) -> str:
        """
        生成统计摘要报告
        
        Returns:
            str: 报告文本
        """
        summary = self.get_task_summary()
        efficiency = self.get_efficiency_analysis()
        status_dist = self.get_status_distribution()
        importance_dist = self.get_importance_distribution()
        responder_stats = self.get_responder_stats(5)
        module_stats = self.get_module_stats(5)
        
        report = []
        report.append("=" * 60)
        report.append("市场咨询任务跟踪工具 - 统计报告")
        report.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("=" * 60)
        report.append("")
        
        # 概览
        report.append("【任务概览】")
        report.append(f"  总任务数: {summary.get('total', 0)}")
        report.append(f"  今日新增: {summary.get('today_created', 0)}")
        report.append(f"  今日完成: {summary.get('today_completed', 0)}")
        report.append(f"  逾期任务: {summary.get('overdue', 0)}")
        report.append("")
        
        # 状态分布
        report.append("【状态分布】")
        for s in status_dist:
            report.append(f"  {s['status']}: {s['count']} ({s['percentage']}%)")
        report.append("")
        
        # 重要程度分布
        report.append("【重要程度分布】")
        for i in importance_dist:
            report.append(f"  {i['importance']}: {i['count']} ({i['percentage']}%)")
        report.append("")
        
        # 效率分析
        report.append("【效率分析】")
        report.append(f"  平均处理时长: {efficiency.get('avg_duration_hours', 0)} 小时")
        report.append(f"  平均响应时长: {efficiency.get('avg_response_hours', 0)} 小时")
        report.append(f"  完成率: {efficiency.get('completion_rate', 0)}%")
        report.append("")
        
        # 责任人排行
        report.append("【责任人排行 Top 5】")
        for i, r in enumerate(responder_stats, 1):
            report.append(f"  {i}. {r['responder']}: {r['total_tasks']}个任务")
        report.append("")
        
        # 热门模块
        report.append("【热门模块 Top 5】")
        for i, m in enumerate(module_stats, 1):
            report.append(f"  {i}. {m['module']}: {m['count']}个任务")
        report.append("")
        
        report.append("=" * 60)
        report.append("报告结束")
        
        return "\n".join(report)
    
    def export_report(self, format: str = 'txt', filepath: str = None) -> str:
        """
        导出统计报告
        
        Args:
            format: 报告格式 ('txt', 'json', 'csv', 'excel')
            filepath: 输出文件路径
            
        Returns:
            str: 报告内容或文件路径
        """
        if format == 'txt':
            return self.generate_summary_report()
        elif format == 'json':
            return self._export_json(filepath)
        elif format == 'csv':
            return self._export_csv(filepath)
        elif format == 'excel':
            return self._export_excel(filepath)
        else:
            raise ValueError(f"不支持的格式: {format}")
    
    def _export_json(self, filepath: str = None) -> str:
        """导出JSON格式报告"""
        import json
        
        data = {
            'generated_at': datetime.now().isoformat(),
            'summary': self.get_task_summary(),
            'status_distribution': self.get_status_distribution(),
            'importance_distribution': self.get_importance_distribution(),
            'responder_stats': self.get_responder_stats(),
            'module_stats': self.get_module_stats(),
            'efficiency': self.get_efficiency_analysis(),
            'department_stats': self.get_department_stats()
        }
        
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        
        if filepath:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(json_str)
        
        return json_str
    
    def _export_csv(self, filepath: str = None) -> str:
        """导出CSV格式报告"""
        import csv
        
        if not filepath:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = f'统计报告_{timestamp}.csv'
        
        with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            
            # 摘要
            summary = self.get_task_summary()
            writer.writerow(['任务统计摘要'])
            writer.writerow(['总任务数', summary.get('total', 0)])
            writer.writerow(['今日新增', summary.get('today_created', 0)])
            writer.writerow(['今日完成', summary.get('today_completed', 0)])
            writer.writerow(['逾期任务', summary.get('overdue', 0)])
            writer.writerow([])
            
            # 状态分布
            writer.writerow(['状态分布'])
            writer.writerow(['状态', '数量', '占比'])
            for s in self.get_status_distribution():
                writer.writerow([s['status'], s['count'], f"{s['percentage']}%"])
            writer.writerow([])
            
            # 责任人统计
            writer.writerow(['责任人统计'])
            writer.writerow(['责任人', '总任务', '进行中', '已完成'])
            for r in self.get_responder_stats():
                writer.writerow([r['responder'], r['total_tasks'], r['in_progress'], r['completed']])
        
        return filepath
    
    def _export_excel(self, filepath: str = None) -> str:
        """
        导出Excel格式报告（多Sheet）
        
        Args:
            filepath: 输出文件路径
            
        Returns:
            str: 文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("请安装openpyxl: pip install openpyxl")
            raise ImportError("导出Excel需要安装openpyxl库")
        
        if not filepath:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = f'统计报告_{timestamp}.xlsx'
        
        # 创建工作簿
        wb = Workbook()
        
        # 样式定义
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # =========================================================================
        # Sheet1: 统计概览
        # =========================================================================
        ws_summary = wb.active
        ws_summary.title = "统计概览"
        
        row = 1
        # 标题
        ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws_summary.cell(row=row, column=1, value="市场咨询任务跟踪工具 - 统计报告")
        ws_summary.cell(row=row, column=1).font = title_font
        ws_summary.cell(row=row, column=1).alignment = center_align
        
        row += 1
        ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws_summary.cell(row=row, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws_summary.cell(row=row, column=1).alignment = center_align
        
        # 获取数据
        summary = self.get_task_summary()
        efficiency = self.get_efficiency_analysis()
        
        row += 2
        # 任务概览标题
        ws_summary.cell(row=row, column=1, value="【任务概览】").font = subtitle_font
        row += 1
        
        # 概览数据
        overview_data = [
            ("总任务数", summary.get('total', 0)),
            ("今日新增", summary.get('today_created', 0)),
            ("今日完成", summary.get('today_completed', 0)),
            ("逾期任务", summary.get('overdue', 0)),
        ]
        
        for label, value in overview_data:
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)
            ws_summary.cell(row=row, column=1).border = thin_border
            ws_summary.cell(row=row, column=2).border = thin_border
            row += 1
        
        row += 1
        # 效率分析标题
        ws_summary.cell(row=row, column=1, value="【效率分析】").font = subtitle_font
        row += 1
        
        # 效率数据
        efficiency_data = [
            ("平均处理时长", f"{efficiency.get('avg_duration_hours', 0)} 小时"),
            ("平均响应时长", f"{efficiency.get('avg_response_hours', 0)} 小时"),
            ("完成率", f"{efficiency.get('completion_rate', 0)}%"),
            ("已完成任务", efficiency.get('completed_tasks', 0)),
            ("进行中任务", efficiency.get('in_progress_tasks', 0)),
        ]
        
        for label, value in efficiency_data:
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)
            ws_summary.cell(row=row, column=1).border = thin_border
            ws_summary.cell(row=row, column=2).border = thin_border
            row += 1
        
        # 设置列宽
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 20
        
        # =========================================================================
        # Sheet2: 状态分布
        # =========================================================================
        ws_status = wb.create_sheet("状态分布")
        
        row = 1
        ws_status.cell(row=row, column=1, value="任务状态分布统计").font = title_font
        ws_status.cell(row=row, column=1).alignment = center_align
        ws_status.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        
        row += 2
        # 表头
        headers = ["状态", "数量", "占比"]
        for col, header in enumerate(headers, 1):
            cell = ws_status.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        row += 1
        # 数据
        status_data = self.get_status_distribution()
        for item in status_data:
            ws_status.cell(row=row, column=1, value=item['status']).border = thin_border
            ws_status.cell(row=row, column=2, value=item['count']).border = thin_border
            ws_status.cell(row=row, column=3, value=f"{item['percentage']}%").border = thin_border
            ws_status.cell(row=row, column=2).alignment = center_align
            ws_status.cell(row=row, column=3).alignment = center_align
            row += 1
        
        # 设置列宽
        ws_status.column_dimensions['A'].width = 15
        ws_status.column_dimensions['B'].width = 12
        ws_status.column_dimensions['C'].width = 12
        
        # =========================================================================
        # Sheet3: 责任人统计
        # =========================================================================
        ws_responder = wb.create_sheet("责任人统计")
        
        row = 1
        ws_responder.cell(row=row, column=1, value="责任人任务统计").font = title_font
        ws_responder.cell(row=row, column=1).alignment = center_align
        ws_responder.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        
        row += 2
        # 表头
        headers = ["责任人", "总任务", "进行中", "已完成", "平均时长(小时)"]
        for col, header in enumerate(headers, 1):
            cell = ws_responder.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        row += 1
        # 数据
        responder_data = self.get_responder_stats(20)
        for item in responder_data:
            ws_responder.cell(row=row, column=1, value=item['responder']).border = thin_border
            ws_responder.cell(row=row, column=2, value=item['total_tasks']).border = thin_border
            ws_responder.cell(row=row, column=3, value=item['in_progress']).border = thin_border
            ws_responder.cell(row=row, column=4, value=item['completed']).border = thin_border
            ws_responder.cell(row=row, column=5, value=item['avg_duration']).border = thin_border
            for col in range(2, 6):
                ws_responder.cell(row=row, column=col).alignment = center_align
            row += 1
        
        # 设置列宽
        ws_responder.column_dimensions['A'].width = 20
        ws_responder.column_dimensions['B'].width = 10
        ws_responder.column_dimensions['C'].width = 10
        ws_responder.column_dimensions['D'].width = 10
        ws_responder.column_dimensions['E'].width = 15
        
        # =========================================================================
        # Sheet4: 模块统计
        # =========================================================================
        ws_module = wb.create_sheet("模块统计")
        
        row = 1
        ws_module.cell(row=row, column=1, value="关键模块统计").font = title_font
        ws_module.cell(row=row, column=1).alignment = center_align
        ws_module.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        
        row += 2
        # 表头
        headers = ["模块名称", "任务数量", "占比"]
        for col, header in enumerate(headers, 1):
            cell = ws_module.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        row += 1
        # 数据
        module_data = self.get_module_stats(50)
        for item in module_data:
            ws_module.cell(row=row, column=1, value=item['module']).border = thin_border
            ws_module.cell(row=row, column=2, value=item['count']).border = thin_border
            ws_module.cell(row=row, column=3, value=f"{item['percentage']}%").border = thin_border
            ws_module.cell(row=row, column=2).alignment = center_align
            ws_module.cell(row=row, column=3).alignment = center_align
            row += 1
        
        # 设置列宽
        ws_module.column_dimensions['A'].width = 30
        ws_module.column_dimensions['B'].width = 12
        ws_module.column_dimensions['C'].width = 12
        
        # =========================================================================
        # Sheet5: 趋势分析
        # =========================================================================
        ws_trend = wb.create_sheet("趋势分析")
        
        row = 1
        ws_trend.cell(row=row, column=1, value="任务创建趋势（近30天）").font = title_font
        ws_trend.cell(row=row, column=1).alignment = center_align
        ws_trend.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        
        row += 2
        # 表头
        headers = ["日期", "新增任务", "完成任务"]
        for col, header in enumerate(headers, 1):
            cell = ws_trend.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        row += 1
        # 数据
        trend_data = self.get_trend_data(days=30, granularity='day')
        for item in trend_data:
            ws_trend.cell(row=row, column=1, value=item['date']).border = thin_border
            ws_trend.cell(row=row, column=2, value=item['created']).border = thin_border
            ws_trend.cell(row=row, column=3, value=item['completed']).border = thin_border
            for col in range(2, 4):
                ws_trend.cell(row=row, column=col).alignment = center_align
            row += 1
        
        # 设置列宽
        ws_trend.column_dimensions['A'].width = 15
        ws_trend.column_dimensions['B'].width = 12
        ws_trend.column_dimensions['C'].width = 12
        
        # =========================================================================
        # Sheet6: 部门统计
        # =========================================================================
        ws_dept = wb.create_sheet("部门统计")
        
        row = 1
        ws_dept.cell(row=row, column=1, value="部门任务统计").font = title_font
        ws_dept.cell(row=row, column=1).alignment = center_align
        ws_dept.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        
        row += 2
        # 表头
        headers = ["部门", "总任务", "已完成", "进行中", "占比"]
        for col, header in enumerate(headers, 1):
            cell = ws_dept.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        row += 1
        # 数据
        dept_data = self.get_department_stats()
        for item in dept_data:
            ws_dept.cell(row=row, column=1, value=item['department']).border = thin_border
            ws_dept.cell(row=row, column=2, value=item['total']).border = thin_border
            ws_dept.cell(row=row, column=3, value=item['completed']).border = thin_border
            ws_dept.cell(row=row, column=4, value=item['in_progress']).border = thin_border
            ws_dept.cell(row=row, column=5, value=f"{item['percentage']}%").border = thin_border
            for col in range(2, 6):
                ws_dept.cell(row=row, column=col).alignment = center_align
            row += 1
        
        # 设置列宽
        ws_dept.column_dimensions['A'].width = 20
        ws_dept.column_dimensions['B'].width = 10
        ws_dept.column_dimensions['C'].width = 10
        ws_dept.column_dimensions['D'].width = 10
        ws_dept.column_dimensions['E'].width = 10
        
        # =========================================================================
        # Sheet7: 重要程度分布
        # =========================================================================
        ws_importance = wb.create_sheet("重要程度")
        
        row = 1
        ws_importance.cell(row=row, column=1, value="任务重要程度分布").font = title_font
        ws_importance.cell(row=row, column=1).alignment = center_align
        ws_importance.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        
        row += 2
        # 表头
        headers = ["重要程度", "数量", "占比"]
        for col, header in enumerate(headers, 1):
            cell = ws_importance.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        row += 1
        # 数据
        importance_data = self.get_importance_distribution()
        for item in importance_data:
            ws_importance.cell(row=row, column=1, value=item['importance']).border = thin_border
            ws_importance.cell(row=row, column=2, value=item['count']).border = thin_border
            ws_importance.cell(row=row, column=3, value=f"{item['percentage']}%").border = thin_border
            ws_importance.cell(row=row, column=2).alignment = center_align
            ws_importance.cell(row=row, column=3).alignment = center_align
            row += 1
        
        # 设置列宽
        ws_importance.column_dimensions['A'].width = 15
        ws_importance.column_dimensions['B'].width = 12
        ws_importance.column_dimensions['C'].width = 12
        
        # 保存文件
        wb.save(filepath)
        logger.info(f"Excel报告已导出: {filepath}")
        
        return filepath


# =============================================================================
# 单例实例
# =============================================================================

_statistics_service: Optional[StatisticsService] = None


def get_statistics_service(db_path: str = None) -> StatisticsService:
    """获取统计服务单例"""
    global _statistics_service
    if _statistics_service is None:
        _statistics_service = StatisticsService(db_path)
    return _statistics_service
