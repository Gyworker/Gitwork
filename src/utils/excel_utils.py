# -*- coding: utf-8 -*-
"""
Excel工具模块
提供公共的Excel操作功能

版本：V4.1 (优化版)
减少代码重复，提取公共函数
"""

from typing import List, Dict, Optional, Any, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# 样式常量
# =============================================================================

# 默认表头样式颜色
DEFAULT_HEADER_COLOR = "4472C4"
GREEN_HEADER_COLOR = "70AD47"
ORANGE_HEADER_COLOR = "ED7D31"

# 默认列宽
DEFAULT_COLUMN_WIDTH = 15
MIN_COLUMN_WIDTH = 8
MAX_COLUMN_WIDTH = 50

# 边框样式
DEFAULT_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 字体样式
DEFAULT_HEADER_FONT = Font(bold=True, color="FFFFFF")
DEFAULT_DATA_FONT = Font(bold=False)


# =============================================================================
# 样式工厂函数
# =============================================================================

def create_header_style(color: str = DEFAULT_HEADER_COLOR) -> Dict[str, Any]:
    """创建表头样式"""
    return {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color=color, end_color=color, fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
    }


def create_data_style(horizontal: str = "left", 
                      vertical: str = "center",
                      wrap_text: bool = False) -> Dict[str, Any]:
    """创建数据单元格样式"""
    return {
        'alignment': Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text),
        'border': DEFAULT_BORDER,
    }


def apply_style_to_cell(cell, style: Dict[str, Any]):
    """应用样式到单元格"""
    if 'font' in style:
        cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    if 'alignment' in style:
        cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']


# =============================================================================
# 列宽管理
# =============================================================================

def calculate_column_width(content: str, min_width: int = MIN_COLUMN_WIDTH, 
                           max_width: int = MAX_COLUMN_WIDTH) -> int:
    """
    计算合适的列宽
    
    Args:
        content: 内容
        min_width: 最小宽度
        max_width: 最大宽度
        
    Returns:
        合适的列宽
    """
    # 中文字符按2倍计算
    width = len(content) * 1.2
    return max(min_width, min(int(width), max_width))


def set_column_width(ws, column: int, width: int):
    """设置单列宽度"""
    ws.column_dimensions[get_column_letter(column)].width = width


def auto_adjust_column_width(ws, default_width: int = DEFAULT_COLUMN_WIDTH,
                             min_width: int = MIN_COLUMN_WIDTH,
                             max_width: int = MAX_COLUMN_WIDTH) -> None:
    """自动调整列宽"""
    for col_idx in range(1, ws.max_column + 1):
        max_content_length = 0
        
        # 遍历该列所有单元格
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                content = str(cell.value)
                max_content_length = max(max_content_length, len(content))
        
        # 计算合适的宽度
        width = calculate_column_width(
            " " * max_content_length,
            min_width, 
            max_width
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# =============================================================================
# Excel操作函数
# =============================================================================

def write_cell(ws, row: int, column: int, value: Any, 
               style: Optional[Dict[str, Any]] = None):
    """
    写入单元格
    
    Args:
        ws: 工作表
        row: 行号
        column: 列号
        value: 值
        style: 样式（可选）
    """
    cell = ws.cell(row=row, column=column, value=value)
    if style:
        apply_style_to_cell(cell, style)


def write_row(ws, row: int, data: List[Any], 
              start_column: int = 1,
              style: Optional[Dict[str, Any]] = None):
    """
    写入一行数据
    
    Args:
        ws: 工作表
        row: 行号
        data: 数据列表
        start_column: 起始列号
        style: 样式（可选）
    """
    for col_idx, value in enumerate(data, start_column):
        write_cell(ws, row, col_idx, value, style)


def create_excel_header(ws, headers: List[str],
                        fill_color: str = DEFAULT_HEADER_COLOR) -> None:
    """创建标准Excel表头"""
    style = create_header_style(fill_color)
    write_row(ws, 1, headers, style=style)


def write_data_rows(ws, data: List[List[Any]], start_row: int = 2) -> None:
    """写入数据行"""
    for row_idx, row_data in enumerate(data, start_row):
        write_row(ws, row_idx, row_data)


# =============================================================================
# 工作簿创建函数
# =============================================================================

def create_workbook(title: str = "Sheet1") -> Tuple[Workbook, Any]:
    """
    创建工作簿
    
    Args:
        title: 工作表名称
        
    Returns:
        (工作簿, 工作表) 元组
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def create_standard_excel(headers: List[str],
                           data: List[List[Any]],
                           title: str = "Sheet1",
                           fill_color: str = DEFAULT_HEADER_COLOR) -> Workbook:
    """创建标准Excel工作簿"""
    wb, ws = create_workbook(title)
    create_excel_header(ws, headers, fill_color)
    write_data_rows(ws, data)
    auto_adjust_column_width(ws)
    return wb


# =============================================================================
# 数据转换函数
# =============================================================================

def dict_to_row(data: Dict, keys: List[str], 
                default: Any = '') -> List[Any]:
    """将字典转换为行数据"""
    return [data.get(key, default) for key in keys]


def dict_list_to_rows(data_list: List[Dict], keys: List[str]) -> List[List[Any]]:
    """将字典列表转换为行数据列表"""
    return [dict_to_row(data, keys) for data in data_list]


def rows_to_dict_list(rows: List[List[Any]], keys: List[str]) -> List[Dict]:
    """将行数据列表转换为字典列表"""
    return [dict(zip(keys, row)) for row in rows]


# =============================================================================
# 文件操作函数
# =============================================================================

def save_workbook(wb: Workbook, output_path: str) -> str:
    """保存工作簿"""
    wb.save(output_path)
    return output_path


def create_export_excel(output_path: str,
                       headers: List[str],
                       data: List[List[Any]],
                       sheet_name: str = "导出数据",
                       fill_color: str = DEFAULT_HEADER_COLOR) -> str:
    """创建导出Excel文件"""
    wb = create_standard_excel(headers, data, sheet_name, fill_color)
    return save_workbook(wb, output_path)


# =============================================================================
# 模板生成函数
# =============================================================================

def create_template(output_path: str,
                   headers: List[str],
                   example_data: Optional[List[List[Any]]] = None,
                   fill_color: str = GREEN_HEADER_COLOR) -> str:
    """创建导入模板"""
    wb, ws = create_workbook("模板")
    create_excel_header(ws, headers, fill_color)
    
    if example_data:
        write_data_rows(ws, example_data)
    
    auto_adjust_column_width(ws)
    return save_workbook(wb, output_path)


# =============================================================================
# 高级功能
# =============================================================================

def add_header_row_number(ws, start_row: int = 1):
    """为表头添加序号"""
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=start_row, column=col_idx).value
        if header:
            ws.cell(row=start_row, column=col_idx).value = f"{col_idx}. {header}"


def add_data_row_number(ws, number_column: int = 0, 
                        start_row: int = 2,
                        header_text: str = "序号"):
    """为数据添加行号"""
    # 如果需要添加序号列
    if number_column > 0:
        # 插入序号列
        ws.insert_cols(number_column)
        # 写入表头
        ws.cell(row=1, column=number_column).value = header_text
        apply_style_to_cell(
            ws.cell(row=1, column=number_column),
            create_header_style()
        )
        # 写入序号
        for row_idx in range(start_row, ws.max_row + 1):
            ws.cell(row=row_idx, column=number_column).value = row_idx - start_row + 1


def freeze_panes(ws, cell: str = "A2"):
    """冻结窗格"""
    ws.freeze_panes = cell


def add_auto_filter(ws, ref_range: Optional[str] = None):
    """添加自动筛选"""
    if ref_range is None:
        ref_range = f"A1:{get_column_letter(ws.max_column)}1"
    ws.auto_filter.ref = ref_range


# =============================================================================
# 批量处理工具
# =============================================================================

class ExcelBatchWriter:
    """Excel批量写入器"""
    
    def __init__(self, ws, start_row: int = 1, 
                 batch_size: int = 100):
        """
        初始化批量写入器
        
        Args:
            ws: 工作表
            start_row: 起始行
            batch_size: 批量大小
        """
        self.ws = ws
        self.current_row = start_row
        self.batch_size = batch_size
        self._row_count = 0
    
    def write_row(self, data: List[Any], 
                  style: Optional[Dict[str, Any]] = None):
        """写入一行"""
        write_row(self.ws, self.current_row, data, style=style)
        self.current_row += 1
        self._row_count += 1
    
    def write_rows(self, rows: List[List[Any]],
                   style: Optional[Dict[str, Any]] = None):
        """批量写入多行"""
        for row in rows:
            self.write_row(row, style)
    
    def write_dict(self, data: Dict, keys: List[str],
                   style: Optional[Dict[str, Any]] = None):
        """写入字典数据"""
        row = dict_to_row(data, keys)
        self.write_row(row, style)
    
    @property
    def row_count(self) -> int:
        """已写入行数"""
        return self._row_count


def export_dict_list(output_path: str,
                     data_list: List[Dict],
                     headers: List[str],
                     sheet_name: str = "导出数据",
                     include_row_numbers: bool = True) -> str:
    """
    导出字典列表为Excel
    
    Args:
        output_path: 输出路径
        data_list: 字典列表
        headers: 表头
        sheet_name: 工作表名称
        include_row_numbers: 是否包含行号
        
    Returns:
        保存的文件路径
    """
    wb, ws = create_workbook(sheet_name)
    
    # 添加表头
    if include_row_numbers:
        ws.cell(row=1, column=1).value = "序号"
        apply_style_to_cell(ws.cell(row=1, column=1), create_header_style())
        headers_full = ["序号"] + headers
        for col, header in enumerate(headers, 2):
            ws.cell(row=1, column=col).value = header
            apply_style_to_cell(ws.cell(row=1, column=col), create_header_style())
    else:
        create_excel_header(ws, headers)
    
    # 批量写入数据
    writer = ExcelBatchWriter(ws, start_row=2)
    for idx, data in enumerate(data_list, 1):
        if include_row_numbers:
            row_data = [idx] + dict_to_row(data, headers)
        else:
            row_data = dict_to_row(data, headers)
        writer.write_row(row_data)
    
    # 自动调整列宽
    auto_adjust_column_width(ws)
    
    return save_workbook(wb, output_path)
