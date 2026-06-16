# -*- coding: utf-8 -*-
"""
增强映射学习模块单元测试
"""

import os
import sys
import pytest
from unittest.mock import Mock, patch, MagicMock
import tempfile
from openpyxl import Workbook

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from src.learning.enhanced_mapping import (
    MappingRule, MappingLearner
)


class TestMappingRule:
    """映射规则数据类测试"""

    def test_create_rule(self):
        """测试创建规则"""
        rule = MappingRule(
            keyword="市场",
            responsible="张三",
            confidence=0.9,
            source="excel"
        )

        assert rule.keyword == "市场"
        assert rule.responsible == "张三"
        assert rule.confidence == 0.9
        assert rule.source == "excel"

    def test_to_dict(self):
        """测试转换为字典"""
        rule = MappingRule(keyword="技术", responsible="李四")
        data = rule.to_dict()

        assert isinstance(data, dict)
        assert data['keyword'] == "技术"
        assert data['responsible'] == "李四"

    def test_from_dict(self):
        """测试从字典创建"""
        data = {
            'id': 1,
            'keyword': '销售',
            'responsible': '王五',
            'confidence': 0.85,
            'source': 'history',
            'usage_count': 10,
            'last_used': '2024-06-16',
            'created_at': '2024-06-01',
            'updated_at': '2024-06-16'
        }
        rule = MappingRule.from_dict(data)

        assert rule.keyword == '销售'
        assert rule.responsible == '王五'
        assert rule.usage_count == 10


class TestMappingLearner:
    """映射学习器测试"""

    @pytest.fixture
    def mock_db(self):
        """创建模拟数据库"""
        db = Mock()
        db.execute_query.return_value = []
        return db

    @pytest.fixture
    def temp_mapping_excel(self):
        """创建临时映射Excel文件"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "映射规则"
            headers = ['关键词', '责任人']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            test_data = [
                ['市场推广', '张三'],
                ['技术开发', '李四'],
                ['产品设计', '王五'],
            ]
            for row_idx, row_data in enumerate(test_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            wb.save(f.name)
            wb.close()
        yield f.name
        if os.path.exists(f.name):
            os.unlink(f.name)

    def test_init(self, mock_db):
        """测试初始化"""
        learner = MappingLearner(mock_db)
        assert learner.db == mock_db
        assert learner.rules == {}

    def test_learn_from_excel(self, mock_db, temp_mapping_excel):
        """测试从Excel学习"""
        learner = MappingLearner(mock_db)
        count = learner.learn_from_excel(temp_mapping_excel)
        assert count == 3
        assert '市场推广' in learner.rules
        assert '技术开发' in learner.rules

    def test_learn_from_text(self, mock_db):
        """测试从文本学习"""
        learner = MappingLearner(mock_db)
        rule = learner.learn_from_text("市场推广活动", "张三")
        assert rule.responsible == "张三"
        assert len(learner.rules) > 0

    def test_learn_from_history(self, mock_db):
        """测试从历史数据学习"""
        tasks = [
            {'module': '市场', 'responsible': '张三'},
            {'module': '市场', 'responsible': '张三'},
            {'module': '技术', 'responsible': '李四'},
        ]
        learner = MappingLearner(mock_db)
        count = learner.learn_from_history(tasks)
        assert count == 2

    def test_recommend_exact_match(self, mock_db):
        """测试精确匹配推荐"""
        learner = MappingLearner(mock_db)
        learner._add_rule("市场推广", "张三")
        responsible, confidence = learner.recommend("市场推广活动策划")
        assert responsible == "张三"
        assert confidence > 0

    def test_recommend_no_match(self, mock_db):
        """测试无匹配"""
        learner = MappingLearner(mock_db)
        responsible, confidence = learner.recommend("完全不相关的任务")
        assert responsible is None
        assert confidence == 0.0

    def test_export_to_excel(self, mock_db):
        """测试导出到Excel"""
        learner = MappingLearner(mock_db)
        learner._add_rule("市场", "张三")
        learner._add_rule("技术", "李四")
        output_path = tempfile.mktemp(suffix='.xlsx')
        saved_path = learner.export_to_excel(output_path)
        assert os.path.exists(saved_path)
        import pandas as pd
        df = pd.read_excel(saved_path)
        assert len(df) >= 2
        os.unlink(saved_path)

    def test_delete_rule(self, mock_db):
        """测试删除规则"""
        learner = MappingLearner(mock_db)
        learner._add_rule("测试", "用户A")
        assert "测试" in learner.rules
        result = learner.delete_rule("测试")
        assert result == True
        assert "测试" not in learner.rules

    def test_extract_keywords(self, mock_db):
        """测试关键词提取"""
        learner = MappingLearner(mock_db)
        keywords = learner._extract_keywords("这是一个市场推广活动")
        assert len(keywords) > 0


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
