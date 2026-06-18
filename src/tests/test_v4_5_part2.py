# -*- coding: utf-8 -*-
"""
V4.5 功能测试 - 第二部分
测试企业微信解析、映射学习定时任务、数据统计服务

版本: V1.0
"""

import unittest
import os
import tempfile
import json
from datetime import datetime
from unittest.mock import Mock, patch, MagicMock


class TestWeChatParser(unittest.TestCase):
    """企业微信解析器测试"""
    
    def setUp(self):
        """初始化测试环境"""
        from src.content.wechat_parser import WeChatParser
        self.parser = WeChatParser()
        self.temp_dir = tempfile.mkdtemp()
    
    def test_is_available(self):
        """测试解析器可用性"""
        self.assertTrue(self.parser.is_available())
    
    def test_get_supported_formats(self):
        """测试支持的格式"""
        formats = self.parser.get_supported_formats()
        self.assertIn('.txt', formats[0])
        self.assertIn('.json', formats[1])
        self.assertIn('.csv', formats[2])
    
    def test_parse_txt_format(self):
        """测试解析文本格式"""
        # 创建测试文件
        test_file = os.path.join(self.temp_dir, 'test.txt')
        content = """[2026-06-18 10:00] 张三: 请问关于MAC认证的问题
[2026-06-18 10:05] 李四: MAC认证需要配置RADIUS服务器
[2026-06-18 10:10] 张三: 谢谢
"""
        with open(test_file, 'w', encoding='utf-8') as f:
            f.write(content)
        
        # 解析文件
        record = self.parser.parse_file(test_file)
        
        self.assertEqual(record.chat_name, 'test')
        self.assertEqual(record.message_count, 3)
        self.assertGreaterEqual(record.member_count, 1)
        
        # 清理
        os.remove(test_file)
    
    def test_parse_json_format(self):
        """测试解析JSON格式"""
        # 创建测试文件
        test_file = os.path.join(self.temp_dir, 'test.json')
        data = {
            'chatName': '测试会话',
            'messages': [
                {'sender': '张三', 'content': '802.1x认证问题', 'time': '2026-06-18 10:00'},
                {'sender': '李四', 'content': '需要检查交换机配置', 'time': '2026-06-18 10:05'},
            ]
        }
        with open(test_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
        
        # 解析文件
        record = self.parser.parse_file(test_file)
        
        self.assertEqual(record.chat_name, '测试会话')
        self.assertEqual(record.message_count, 2)
        
        # 清理
        os.remove(test_file)
    
    def test_parse_csv_format(self):
        """测试解析CSV格式"""
        # 创建测试文件
        test_file = os.path.join(self.temp_dir, 'test.csv')
        content = """时间,发送者,内容
2026-06-18 10:00,张三,Portal认证咨询
2026-06-18 10:05,李四,需要提供具体型号
"""
        with open(test_file, 'w', encoding='utf-8') as f:
            f.write(content)
        
        # 解析文件
        record = self.parser.parse_file(test_file)
        
        self.assertEqual(record.message_count, 2)
        
        # 清理
        os.remove(test_file)
    
    def test_extract_key_modules(self):
        """测试提取关键模块"""
        content = """
        客户咨询关于MAC认证和802.1x认证的问题。
        需要配置RADIUS服务器和AC设备。
        产品型号为S5500-24P-SI。
        """
        
        modules = self.parser._extract_key_modules(content)
        
        self.assertIn('MAC认证', modules)
        self.assertIn('802.1x认证', modules)
        self.assertIn('RADIUS', modules)
        self.assertIn('AC', modules)
    
    def test_extract_products(self):
        """测试提取产品型号"""
        content = """
        产品型号：
        - S5500-24P-SI
        - S5120-28P-EI
        - MSR830-5BEI
        """
        
        products = self.parser._extract_products(content)
        
        self.assertIn('S5500-24P-SI', products)
        self.assertIn('S5120-28P-EI', products)
    
    def test_convert_to_task(self):
        """测试转换为任务信息"""
        from src.content.wechat_parser import WeChatChatRecord, WeChatMessage
        
        record = WeChatChatRecord(
            chat_name='测试会话',
            start_time='2026-06-18 10:00',
            end_time='2026-06-18 10:30',
            messages=[
                WeChatMessage(sender='张三', time='10:00', content='MAC认证配置问题'),
                WeChatMessage(sender='李四', time='10:05', content='需要配置RADIUS'),
            ]
        )
        
        result = self.parser._convert_to_task(record)
        
        self.assertEqual(result.source_type, 'wechat')
        self.assertEqual(result.source, '测试会话')
        self.assertEqual(result.consultant_name, '张三')
        self.assertIn('MAC认证', result.key_module)
        self.assertIn('RADIUS', result.key_module)
    
    def test_parse_content_with_text(self):
        """测试解析文本内容（不读取文件）"""
        content = """
        [2026-06-18 10:00] 张三: OSPF路由协议问题咨询
        [2026-06-18 10:05] 李四: 需要查看设备配置
        """
        
        result = self.parser.parse_content(content)
        
        self.assertEqual(result.source_type, 'wechat')
        self.assertIn('OSPF', result.key_module)
        self.assertEqual(result.consultant_name, '张三')


class TestMappingScheduler(unittest.TestCase):
    """映射学习定时任务调度器测试"""
    
    def setUp(self):
        """初始化测试环境"""
        from src.learning.mapping_scheduler import (
            MappingScheduler,
            ScheduleType,
            TaskStatus
        )
        self.MappingScheduler = MappingScheduler
        self.ScheduleType = ScheduleType
        self.TaskStatus = TaskStatus
        self.scheduler = MappingScheduler()
    
    def test_init(self):
        """测试初始化"""
        self.assertFalse(self.scheduler.is_running())
        self.assertEqual(len(self.scheduler.tasks), 0)
    
    def test_add_daily_task(self):
        """测试添加每日任务"""
        task_id = self.scheduler.add_daily_task(
            name="测试任务",
            execution_time="01:00",
            job_func=lambda: "success"
        )
        
        self.assertIn(task_id, self.scheduler.tasks)
        task = self.scheduler.tasks[task_id]
        self.assertEqual(task.name, "测试任务")
        self.assertEqual(task.schedule_type, self.ScheduleType.DAILY)
        self.assertTrue(task.enabled)
    
    def test_add_weekly_task(self):
        """测试添加每周任务"""
        task_id = self.scheduler.add_weekly_task(
            name="每周任务",
            execution_time="02:00",
            execution_day=0  # 周一
        )
        
        self.assertIn(task_id, self.scheduler.tasks)
        task = self.scheduler.tasks[task_id]
        self.assertEqual(task.schedule_type, self.ScheduleType.WEEKLY)
        self.assertEqual(task.execution_day, 0)
    
    def test_add_monthly_task(self):
        """测试添加每月任务"""
        task_id = self.scheduler.add_monthly_task(
            name="每月任务",
            execution_time="03:00",
            execution_day=1  # 每月1日
        )
        
        self.assertIn(task_id, self.scheduler.tasks)
        task = self.scheduler.tasks[task_id]
        self.assertEqual(task.schedule_type, self.ScheduleType.MONTHLY)
        self.assertEqual(task.execution_day, 1)
    
    def test_remove_task(self):
        """测试移除任务"""
        task_id = self.scheduler.add_daily_task(name="待移除任务")
        self.assertIn(task_id, self.scheduler.tasks)
        
        result = self.scheduler.remove_task(task_id)
        self.assertTrue(result)
        self.assertNotIn(task_id, self.scheduler.tasks)
    
    def test_enable_disable_task(self):
        """测试启用/禁用任务"""
        task_id = self.scheduler.add_daily_task(name="测试任务")
        task = self.scheduler.tasks[task_id]
        
        # 禁用
        self.scheduler.disable_task(task_id)
        self.assertFalse(task.enabled)
        
        # 启用
        self.scheduler.enable_task(task_id)
        self.assertTrue(task.enabled)
    
    def test_execute_now(self):
        """测试立即执行任务"""
        executed = []
        
        def job():
            executed.append(True)
            return "任务执行成功"
        
        task_id = self.scheduler.add_daily_task(
            name="立即执行任务",
            job_func=job
        )
        
        result = self.scheduler.execute_now(task_id)
        
        self.assertTrue(result)
        self.assertEqual(len(executed), 1)
    
    def test_execution_history(self):
        """测试执行历史"""
        task_id = self.scheduler.add_daily_task(
            name="测试历史任务",
            job_func=lambda: "ok"
        )
        
        self.scheduler.execute_now(task_id)
        
        history = self.scheduler.get_execution_history()
        self.assertGreaterEqual(len(history), 1)
        
        task_history = self.scheduler.get_execution_history(task_id)
        self.assertEqual(len(task_history), 1)
        self.assertEqual(task_history[0].task_name, "测试历史任务")


class TestStatisticsService(unittest.TestCase):
    """统计数据服务测试"""
    
    def setUp(self):
        """初始化测试环境"""
        from src.core.statistics_service import StatisticsService
        
        # 创建模拟数据库
        self.mock_db = Mock()
        self.service = StatisticsService()
        self.service._db = self.mock_db
    
    def test_get_task_summary(self):
        """测试获取任务统计摘要"""
        # 模拟数据库返回值
        self.mock_db.execute_query.side_effect = [
            [{'cnt': 100}],  # 总数
            [{'status': '进行中', 'cnt': 30}, {'status': '完成', 'cnt': 70}],  # 状态
            [{'importance': '高', 'cnt': 20}, {'importance': '中', 'cnt': 80}],  # 重要程度
            [{'cnt': 5}],  # 今日新增
            [{'cnt': 3}],  # 今日完成
            [{'cnt': 2}],  # 逾期
        ]
        
        result = self.service.get_task_summary()
        
        self.assertEqual(result['total'], 100)
        self.assertEqual(result['by_status']['进行中'], 30)
        self.assertEqual(result['by_status']['完成'], 70)
        self.assertEqual(result['today_created'], 5)
        self.assertEqual(result['today_completed'], 3)
    
    def test_get_status_distribution(self):
        """测试获取状态分布"""
        self.mock_db.execute_query.side_effect = [
            [{'cnt': 100}],  # 总数
            [
                {'status': '进行中', 'cnt': 30},
                {'status': '完成', 'cnt': 50},
                {'status': '挂起', 'cnt': 20}
            ]
        ]
        
        result = self.service.get_status_distribution()
        
        self.assertEqual(len(result), 3)
        self.assertEqual(result[0]['status'], '进行中')
        self.assertEqual(result[0]['percentage'], 30.0)
    
    def test_get_importance_distribution(self):
        """测试获取重要程度分布"""
        self.mock_db.execute_query.side_effect = [
            [{'cnt': 100}],
            [
                {'importance': '高', 'cnt': 25},
                {'importance': '中', 'cnt': 50},
                {'importance': '低', 'cnt': 25}
            ]
        ]
        
        result = self.service.get_importance_distribution()
        
        self.assertEqual(len(result), 3)
        self.assertEqual(result[0]['percentage'], 25.0)
    
    def test_get_responder_stats(self):
        """测试获取责任人统计"""
        self.mock_db.execute_query.side_effect = [
            [
                {'name': '张三', 'total_tasks': 30, 'in_progress': 10, 'completed': 20},
                {'name': '李四', 'total_tasks': 25, 'in_progress': 5, 'completed': 20}
            ],
            [{'avg_days': 1.5}]  # 平均时长
        ]
        
        result = self.service.get_responder_stats(limit=5)
        
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]['responder'], '张三')
        self.assertEqual(result[0]['total_tasks'], 30)
        self.assertEqual(result[0]['avg_duration'], 36.0)  # 1.5天 = 36小时
    
    def test_get_module_stats(self):
        """测试获取模块统计"""
        self.mock_db.execute_query.side_effect = [
            [
                {'key_module': 'MAC认证', 'cnt': 20},
                {'key_module': '802.1x认证', 'cnt': 15},
                {'key_module': 'Portal认证', 'cnt': 10}
            ]
        ]
        
        result = self.service.get_module_stats(limit=5)
        
        self.assertEqual(len(result), 3)
        self.assertEqual(result[0]['module'], 'MAC认证')
        self.assertEqual(result[0]['percentage'], 44.4)  # 20/45 * 100
    
    def test_generate_summary_report(self):
        """测试生成摘要报告"""
        self.mock_db.execute_query.side_effect = [
            [{'cnt': 100}],
            [{'status': '进行中', 'cnt': 30}, {'status': '完成', 'cnt': 70}],
            [{'importance': '高', 'cnt': 20}, {'importance': '中', 'cnt': 80}],
            [{'cnt': 5}],
            [{'cnt': 3}],
            [{'cnt': 2}],
            [{'avg_days': 1.5}],
            [{'cnt': 100}],
            [{'cnt': 70}],
            [{'avg_days': 1.0}],
            [
                {'name': '张三', 'total_tasks': 30, 'in_progress': 10, 'completed': 20}
            ],
            [
                {'key_module': 'MAC认证', 'cnt': 20}
            ]
        ]
        
        report = self.service.generate_summary_report()
        
        self.assertIn('任务概览', report)
        self.assertIn('总任务数: 100', report)
        self.assertIn('今日新增: 5', report)
        self.assertIn('今日完成: 3', report)
        self.assertIn('状态分布', report)
        self.assertIn('效率分析', report)


class TestStatisticsServiceDatabase(unittest.TestCase):
    """统计数据服务数据库测试"""
    
    def setUp(self):
        """初始化测试环境"""
        from src.core.statistics_service import StatisticsService
        
        # 使用内存数据库测试
        self.service = StatisticsService()
    
    def test_get_task_summary_empty(self):
        """测试空数据库统计"""
        result = self.service.get_task_summary()
        
        self.assertEqual(result['total'], 0)
        self.assertEqual(result['today_created'], 0)
    
    def test_get_status_distribution_empty(self):
        """测试空数据库状态分布"""
        result = self.service.get_status_distribution()
        
        self.assertEqual(len(result), 0)


class TestStatisticsExcelExport(unittest.TestCase):
    """统计服务Excel导出测试"""
    
    def setUp(self):
        """初始化测试环境"""
        from src.core.statistics_service import StatisticsService
        
        self.service = StatisticsService()
        self.temp_dir = tempfile.mkdtemp()
    
    def test_export_report_excel_format(self):
        """测试Excel格式导出"""
        # 模拟数据库返回
        with patch.object(self.service, 'get_task_summary') as mock_summary, \
             patch.object(self.service, 'get_efficiency_analysis') as mock_eff, \
             patch.object(self.service, 'get_status_distribution') as mock_status, \
             patch.object(self.service, 'get_importance_distribution') as mock_importance, \
             patch.object(self.service, 'get_responder_stats') as mock_responder, \
             patch.object(self.service, 'get_module_stats') as mock_module, \
             patch.object(self.service, 'get_department_stats') as mock_dept, \
             patch.object(self.service, 'get_trend_data') as mock_trend:
            
            mock_summary.return_value = {'total': 100, 'today_created': 5, 'today_completed': 3, 'overdue': 2}
            mock_eff.return_value = {'avg_duration_hours': 24.0, 'avg_response_hours': 12.0, 'completion_rate': 70.0, 'completed_tasks': 70, 'in_progress_tasks': 30}
            mock_status.return_value = [{'status': '进行中', 'count': 30, 'percentage': 30.0}, {'status': '完成', 'count': 70, 'percentage': 70.0}]
            mock_importance.return_value = [{'importance': '高', 'count': 25, 'percentage': 25.0}]
            mock_responder.return_value = [{'responder': '张三', 'total_tasks': 30, 'in_progress': 10, 'completed': 20, 'avg_duration': 24.0}]
            mock_module.return_value = [{'module': 'MAC认证', 'count': 20, 'percentage': 20.0}]
            mock_dept.return_value = [{'department': '技术部', 'total': 50, 'completed': 35, 'in_progress': 15, 'percentage': 50.0}]
            mock_trend.return_value = [{'date': '2026-06-18', 'created': 5, 'completed': 3}]
            
            # 导出Excel
            filepath = os.path.join(self.temp_dir, 'test_report.xlsx')
            result = self.service.export_report(format='excel', filepath=filepath)
            
            # 验证文件生成
            self.assertTrue(os.path.exists(result))
            self.assertTrue(result.endswith('.xlsx'))
            
            # 验证文件内容
            try:
                from openpyxl import load_workbook
                wb = load_workbook(result)
                
                # 验证Sheet
                self.assertIn('统计概览', wb.sheetnames)
                self.assertIn('状态分布', wb.sheetnames)
                self.assertIn('责任人统计', wb.sheetnames)
                self.assertIn('模块统计', wb.sheetnames)
                self.assertIn('趋势分析', wb.sheetnames)
                self.assertIn('部门统计', wb.sheetnames)
                self.assertIn('重要程度', wb.sheetnames)
                
                # 验证统计概览Sheet内容
                ws_summary = wb['统计概览']
                self.assertEqual(ws_summary.cell(row=1, column=1).value, '市场咨询任务跟踪工具 - 统计报告')
                
                wb.close()
            except ImportError:
                self.skipTest("openpyxl未安装")
    
    def test_export_excel_default_filepath(self):
        """测试Excel导出默认文件名"""
        with patch.object(self.service, 'get_task_summary') as mock_summary, \
             patch.object(self.service, 'get_efficiency_analysis') as mock_eff, \
             patch.object(self.service, 'get_status_distribution') as mock_status, \
             patch.object(self.service, 'get_importance_distribution') as mock_importance, \
             patch.object(self.service, 'get_responder_stats') as mock_responder, \
             patch.object(self.service, 'get_module_stats') as mock_module, \
             patch.object(self.service, 'get_department_stats') as mock_dept, \
             patch.object(self.service, 'get_trend_data') as mock_trend:
            
            mock_summary.return_value = {'total': 0}
            mock_eff.return_value = {}
            mock_status.return_value = []
            mock_importance.return_value = []
            mock_responder.return_value = []
            mock_module.return_value = []
            mock_dept.return_value = []
            mock_trend.return_value = []
            
            # 不指定路径，生成默认文件名
            result = self.service.export_report(format='excel')
            
            self.assertTrue(os.path.exists(result))
            self.assertTrue('统计报告_' in result)
            self.assertTrue(result.endswith('.xlsx'))
    
    def test_export_unsupported_format(self):
        """测试不支持的格式"""
        with self.assertRaises(ValueError) as context:
            self.service.export_report(format='pdf')
        
        self.assertIn('不支持的格式', str(context.exception))
    
    def test_export_excel_with_all_data(self):
        """测试导出完整数据"""
        with patch.object(self.service, 'get_task_summary') as mock_summary, \
             patch.object(self.service, 'get_efficiency_analysis') as mock_eff, \
             patch.object(self.service, 'get_status_distribution') as mock_status, \
             patch.object(self.service, 'get_importance_distribution') as mock_importance, \
             patch.object(self.service, 'get_responder_stats') as mock_responder, \
             patch.object(self.service, 'get_module_stats') as mock_module, \
             patch.object(self.service, 'get_department_stats') as mock_dept, \
             patch.object(self.service, 'get_trend_data') as mock_trend:
            
            mock_summary.return_value = {'total': 100, 'today_created': 5, 'today_completed': 3, 'overdue': 2}
            mock_eff.return_value = {'avg_duration_hours': 24.0, 'avg_response_hours': 12.0, 'completion_rate': 70.0}
            mock_status.return_value = [
                {'status': '进行中', 'count': 30, 'percentage': 30.0},
                {'status': '完成', 'count': 50, 'percentage': 50.0},
                {'status': '挂起', 'count': 20, 'percentage': 20.0}
            ]
            mock_importance.return_value = [
                {'importance': '高', 'count': 25, 'percentage': 25.0},
                {'importance': '中', 'count': 50, 'percentage': 50.0},
                {'importance': '低', 'count': 25, 'percentage': 25.0}
            ]
            mock_responder.return_value = [
                {'responder': '张三', 'total_tasks': 30, 'in_progress': 10, 'completed': 20, 'avg_duration': 24.0},
                {'responder': '李四', 'total_tasks': 25, 'in_progress': 5, 'completed': 20, 'avg_duration': 20.0}
            ]
            mock_module.return_value = [
                {'module': 'MAC认证', 'count': 20, 'percentage': 20.0},
                {'module': '802.1x认证', 'count': 15, 'percentage': 15.0}
            ]
            mock_dept.return_value = [
                {'department': '技术部', 'total': 50, 'completed': 35, 'in_progress': 15, 'percentage': 50.0}
            ]
            mock_trend.return_value = [
                {'date': '2026-06-18', 'created': 5, 'completed': 3},
                {'date': '2026-06-17', 'created': 8, 'completed': 6}
            ]
            
            filepath = os.path.join(self.temp_dir, 'full_report.xlsx')
            result = self.service.export_report(format='excel', filepath=filepath)
            
            self.assertTrue(os.path.exists(result))
            
            # 验证数据完整性
            try:
                from openpyxl import load_workbook
                wb = load_workbook(result)
                
                # 验证状态分布
                ws_status = wb['状态分布']
                self.assertEqual(ws_status.cell(row=4, column=1).value, '进行中')
                self.assertEqual(ws_status.cell(row=4, column=2).value, 30)
                
                # 验证责任人统计
                ws_responder = wb['责任人统计']
                self.assertEqual(ws_responder.cell(row=4, column=1).value, '张三')
                self.assertEqual(ws_responder.cell(row=4, column=2).value, 30)
                
                wb.close()
            except ImportError:
                self.skipTest("openpyxl未安装")
    
    def tearDown(self):
        """清理测试环境"""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)


# =============================================================================
# 测试运行器
# =============================================================================

if __name__ == '__main__':
    # 运行测试
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # 添加测试类
    suite.addTests(loader.loadTestsFromTestCase(TestWeChatParser))
    suite.addTests(loader.loadTestsFromTestCase(TestMappingScheduler))
    suite.addTests(loader.loadTestsFromTestCase(TestStatisticsService))
    suite.addTests(loader.loadTestsFromTestCase(TestStatisticsServiceDatabase))
    suite.addTests(loader.loadTestsFromTestCase(TestStatisticsExcelExport))
    
    # 运行
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    # 输出摘要
    print("\n" + "=" * 60)
    print("测试摘要")
    print("=" * 60)
    print(f"运行测试: {result.testsRun}")
    print(f"成功: {result.testsRun - len(result.failures) - len(result.errors)}")
    print(f"失败: {len(result.failures)}")
    print(f"错误: {len(result.errors)}")
