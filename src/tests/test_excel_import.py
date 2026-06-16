# -*- coding: utf-8 -*-
"""
Excel导入模块单元测试
"""

import os
import sys
import pytest
import tempfile
from datetime import datetime
from unittest.mock import Mock, patch, MagicMock
import pandas as pd
from openpyxl import Workbook

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from src.content.excel_import import (
    ExcelContact, ExcelHeaderMapper, ExcelImporter, ExcelExporter
)


class TestExcelHeaderMapper:
    """Excel表头映射器测试"""

    def test_find_header_mapping_standard(self):
        """测试标准表头识别"""
        headers = ['姓名', '电话', '邮箱', '公司', '部门', '职位', '备注']
        mapping = ExcelHeaderMapper.find_header_mapping(headers)

        assert '姓名' in mapping
        assert '电话' in mapping
        assert '邮箱' in mapping
        assert mapping['姓名'] == 0

    def test_find_header_mapping_english(self):
        """测试英文表头识别"""
        headers = ['name', 'phone', 'email', 'company', 'department', 'position', 'remark']
        mapping = ExcelHeaderMapper.find_header_mapping(headers)

        assert '姓名' in mapping
        assert '电话' in mapping
        assert '邮箱' in mapping

    def test_find_header_mapping_mixed(self):
        """测试混合表头识别"""
        headers = ['姓名', 'mobile', 'Email', '公司', 'dept', 'title', '']
        mapping = ExcelHeaderMapper.find_header_mapping(headers)

        assert '姓名' in mapping
        assert '电话' in mapping  # mobile
        assert '邮箱' in mapping  # Email
        assert '部门' in mapping   # dept

    def test_find_header_mapping_partial(self):
        """测试部分表头识别"""
        headers = ['姓名', '手机']  # 只有部分
        mapping = ExcelHeaderMapper.find_header_mapping(headers)

        assert '姓名' in mapping
        # phone应该找不到对应的中文映射
        # 但应该能找到phone相关的别名

    def test_find_header_mapping_empty(self):
        """测试空表头"""
        headers = []
        mapping = ExcelHeaderMapper.find_header_mapping(headers)
        assert mapping == {}


class TestExcelContact:
    """Excel联系人数据类测试"""

    def test_create_contact(self):
        """测试创建联系人"""
        contact = ExcelContact(
            姓名='张三',
            电话='13800138000',
            邮箱='zhangsan@example.com',
            公司='测试公司',
            部门='市场部',
            职位='经理',
            备注='VIP客户'
        )

        assert contact.姓名 == '张三'
        assert contact.电话 == '13800138000'
        assert contact.邮箱 == 'zhangsan@example.com'

    def test_to_dict(self):
        """测试转换为字典"""
        contact = ExcelContact(姓名='李四', 电话='13900139000')
        data = contact.to_dict()

        assert isinstance(data, dict)
        assert data['姓名'] == '李四'
        assert data['电话'] == '13900139000'

    def test_from_dict(self):
        """测试从字典创建"""
        data = {
            '姓名': '王五',
            '电话': '13700137000',
            '邮箱': '',
            '公司': '',
            '部门': '',
            '职位': '',
            '备注': '',
            'source': 'excel',
            'import_batch': '20240616'
        }
        contact = ExcelContact.from_dict(data)

        assert contact.姓名 == '王五'
        assert contact.source == 'excel'


class TestExcelImporter:
    """Excel导入器测试"""

    @pytest.fixture
    def mock_db(self):
        """创建模拟数据库"""
        db = Mock()
        db.get_all_contacts.return_value = []
        db.add_contact.return_value = True
        return db

    @pytest.fixture
    def temp_excel_file(self):
        """创建临时Excel文件"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "通讯录"

            # 写入表头
            headers = ['姓名', '电话', '邮箱', '公司', '部门', '职位', '备注']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 写入测试数据
            test_data = [
                ['张三', '13800138000', 'zhangsan@example.com', '公司A', '市场部', '经理', 'VIP'],
                ['李四', '13900139000', 'lisi@example.com', '公司B', '销售部', '主管', ''],
                ['王五', '13700137000', 'wangwu@example.com', '公司C', '技术部', '工程师', ''],
            ]

            for row_idx, row_data in enumerate(test_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(f.name)
            wb.close()

        yield f.name

        # 清理
        if os.path.exists(f.name):
            os.unlink(f.name)

    @pytest.fixture
    def temp_excel_empty(self):
        """创建空Excel文件"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "空表"
            wb.save(f.name)
            wb.close()

        yield f.name

        if os.path.exists(f.name):
            os.unlink(f.name)

    def test_import_from_file_success(self, mock_db, temp_excel_file):
        """测试成功导入"""
        importer = ExcelImporter(mock_db)
        result = importer.import_from_file(temp_excel_file)

        assert result['success'] == True
        assert result['total'] == 3
        assert result['added'] == 3
        assert result['duplicate'] == 0
        assert len(result['records']) == 3

    def test_import_from_file_empty(self, mock_db, temp_excel_empty):
        """测试导入空文件"""
        importer = ExcelImporter(mock_db)
        result = importer.import_from_file(temp_excel_empty)

        # 空文件应该能读取，但数据为空
        assert result['total'] == 0

    def test_import_duplicate_detection(self, mock_db, temp_excel_file):
        """测试重复检测"""
        # 模拟数据库已有数据
        mock_db.get_all_contacts.return_value = [
            {'name': '张三', 'phone': '13800138000'}
        ]

        importer = ExcelImporter(mock_db)
        result = importer.import_from_file(temp_excel_file)

        assert result['success'] == True
        assert result['total'] == 3
        assert result['added'] == 2  # 2个新的
        assert result['duplicate'] == 1  # 1个重复的

    def test_save_to_database(self, mock_db, temp_excel_file):
        """测试保存到数据库"""
        importer = ExcelImporter(mock_db)
        importer.import_from_file(temp_excel_file)

        saved = importer.save_to_database()

        assert saved == 3
        assert mock_db.add_contact.call_count == 3

    def test_save_to_excel(self, mock_db, temp_excel_file):
        """测试保存到Excel"""
        importer = ExcelImporter(mock_db)
        importer.import_from_file(temp_excel_file)

        output_path = tempfile.mktemp(suffix='.xlsx')
        saved_path = importer.save_to_excel(output_path)

        assert os.path.exists(saved_path)

        # 验证内容
        df = pd.read_excel(saved_path)
        assert len(df) == 3

        # 清理
        os.unlink(saved_path)

    def test_import_unsupported_format(self, mock_db):
        """测试不支持的格式"""
        importer = ExcelImporter(mock_db)
        result = importer.import_from_file('test.txt')

        assert result['success'] == False
        assert len(result['errors']) > 0


class TestExcelExporter:
    """Excel导出器测试"""

    @pytest.fixture
    def mock_db(self):
        """创建模拟数据库"""
        db = Mock()
        db.get_all_contacts.return_value = [
            {'name': '张三', 'phone': '13800138000', 'email': 'zhangsan@example.com',
             'company': '公司A', 'department': '市场部', 'position': '经理', 'remark': 'VIP'},
            {'name': '李四', 'phone': '13900139000', 'email': 'lisi@example.com',
             'company': '公司B', 'department': '销售部', 'position': '主管', 'remark': ''},
        ]
        db.search_contacts.return_value = [
            {'name': '张三', 'phone': '13800138000', 'email': 'zhangsan@example.com',
             'company': '公司A', 'department': '市场部', 'position': '经理', 'remark': 'VIP'},
        ]
        return db

    def test_export_all(self, mock_db):
        """测试导出全部"""
        exporter = ExcelExporter(mock_db)

        output_path = tempfile.mktemp(suffix='.xlsx')
        saved_path = exporter.export_all(output_path)

        assert os.path.exists(saved_path)

        # 验证内容
        df = pd.read_excel(saved_path)
        assert len(df) == 2

        # 清理
        os.unlink(saved_path)

    def test_export_filtered(self, mock_db):
        """测试导出筛选结果"""
        exporter = ExcelExporter(mock_db)

        output_path = tempfile.mktemp(suffix='.xlsx')
        saved_path = exporter.export_filtered(output_path, filters={'keyword': '张三'})

        assert os.path.exists(saved_path)

        # 验证内容
        df = pd.read_excel(saved_path)
        assert len(df) == 1
        assert df.iloc[0]['姓名'] == '张三'

        # 清理
        os.unlink(saved_path)

    def test_create_template(self, mock_db):
        """测试创建模板"""
        exporter = ExcelExporter(mock_db)

        output_path = tempfile.mktemp(suffix='.xlsx')
        saved_path = exporter.create_template(output_path)

        assert os.path.exists(saved_path)

        # 验证内容
        df = pd.read_excel(saved_path)
        # 应该有示例数据
        assert len(df) >= 1

        # 清理
        os.unlink(saved_path)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
