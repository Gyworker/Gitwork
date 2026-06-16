# -*- coding: utf-8 -*-
"""
OCR处理模块
支持图片文字识别和名片解析
版本：V2.1 (优化版)
"""

import os
import re
import json
import sys
import time
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, asdict, field
from pathlib import Path

# 图片处理
from PIL import Image, ImageEnhance, ImageFilter
import pytesseract


# =============================================================================
# 常量定义
# =============================================================================

# 置信度权重配置
CONFIDENCE_WEIGHTS = {
    'name': 0.3,
    'phone': 0.3,
    'email': 0.2,
    'company': 0.2,
}

# 电话号码正则
PHONE_PATTERN = re.compile(
    r'(?:电话|TEL|手机|Mobile|Phone)?[:：]?\s*'
    r'(\+?86)?\s*'
    r'(?:1[3-9]\d{9}|(?:010|021|022|023|024|025|027|028|029)\d{7,8})'
)

# 邮箱正则
EMAIL_PATTERN = re.compile(
    r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
)

# 常见职位关键词
POSITION_KEYWORDS = [
    '经理', '总监', '主管', '负责人', '主任', '工程师',
    'Consultant', 'Manager', 'Director', 'Chief', 'CEO', 'CTO', 'CFO',
    'President', 'Vice', 'Senior', 'Lead', 'Head'
]

# 常见部门关键词
DEPARTMENT_KEYWORDS = [
    '部', '部门', '科', '室',
    'Department', 'Dept', 'Division', 'Team'
]

# 公司关键词
COMPANY_KEYWORDS = ['公司', '有限公司', 'Co.', 'Ltd', 'Corp']

# 地址关键词
ADDRESS_KEYWORDS = ['地址', 'Address', '市', '区', '路', '街']

# 联系方式关键词
CONTACT_KEYWORDS = ['电话', '手机', '邮箱', 'email', 'phone', 'tel', '地址', 'address']


@dataclass
class OCRResult:
    """OCR识别结果"""
    raw_text: str = ""  # 原始识别文本
    name: str = ""      # 姓名
    phone: str = ""     # 电话
    email: str = ""     # 邮箱
    company: str = ""   # 公司
    department: str = "" # 部门
    position: str = ""  # 职位
    address: str = ""   # 地址
    confidence: float = 0.0  # 置信度
    source_image: str = ""    # 来源图片路径
    process_time: float = 0.0  # 处理时间

    def to_dict(self) -> Dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: Dict) -> 'OCRResult':
        return cls(**data)


class ImagePreprocessor:
    """图像预处理"""

    @staticmethod
    def preprocess(image_path: str) -> Image.Image:
        """图像预处理

        Args:
            image_path: 图片路径

        Returns:
            处理后的图像
        """
        img = Image.open(image_path)

        # 转为RGB（处理PNG透明通道等）
        if img.mode != 'RGB':
            img = img.convert('RGB')

        # 灰度转换
        img = img.convert('L')

        # 对比度增强
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(1.5)

        # 锐化
        img = img.filter(ImageFilter.SHARPEN)

        # 去噪声
        img = img.filter(ImageFilter.MedianFilter(size=3))

        return img

    @staticmethod
    def enhance_for_text(image: Image.Image) -> Image.Image:
        """增强文本识别效果

        Args:
            image: 输入图像

        Returns:
            增强后的图像
        """
        # 二值化
        img = image.point(lambda x: 0 if x < 128 else 255, '1')
        return img.convert('L')


class BusinessCardParser:
    """名片解析器 - 优化版，方法拆分提高可维护性"""

    def __init__(self):
        """初始化解析器"""
        self.phone_pattern = PHONE_PATTERN
        self.email_pattern = EMAIL_PATTERN

    def parse(self, text: str) -> Dict:
        """解析名片文本 - 主方法，调度各子方法

        Args:
            text: OCR识别的原始文本

        Returns:
            解析后的结构化数据
        """
        lines = self._split_lines(text)

        result = {
            'name': self._extract_name(lines),
            'phone': self._extract_phone(text),
            'email': self._extract_email(text),
            'company': self._extract_company(lines),
            'department': self._extract_department(lines),
            'position': self._extract_position(lines),
            'address': self._extract_address(lines),
        }

        # 如果没有识别到姓名，尝试从邮箱推断
        if not result['name'] and result['email']:
            result['name'] = self._extract_name_from_email(result['email'])

        return result

    @staticmethod
    def _split_lines(text: str) -> List[str]:
        """分割文本为行

        Args:
            text: 原始文本

        Returns:
            非空行列表
        """
        lines = text.split('\n')
        return [line.strip() for line in lines if line.strip()]

    @staticmethod
    def _extract_phone(text: str) -> str:
        """提取电话号码

        Args:
            text: 文本内容

        Returns:
            电话号码或空字符串
        """
        phone_match = PHONE_PATTERN.search(text)
        if phone_match:
            return BusinessCardParser._clean_phone(phone_match.group(0))
        return ""

    @staticmethod
    def _extract_email(text: str) -> str:
        """提取邮箱地址

        Args:
            text: 文本内容

        Returns:
            邮箱地址或空字符串
        """
        email_match = EMAIL_PATTERN.search(text)
        if email_match:
            return email_match.group(0).lower()
        return ""

    @staticmethod
    def _extract_name(lines: List[str]) -> str:
        """提取姓名

        Args:
            lines: 文本行列表

        Returns:
            姓名或空字符串
        """
        if not lines:
            return ""

        # 姓名通常是第一行的大字体文本
        first_line = lines[0]
        if len(first_line) <= 10 and not BusinessCardParser._is_contact_info(first_line):
            return first_line
        return ""

    @staticmethod
    def _extract_company(lines: List[str]) -> str:
        """提取公司名称

        Args:
            lines: 文本行列表

        Returns:
            公司名称或空字符串
        """
        for line in lines:
            if any(kw in line for kw in COMPANY_KEYWORDS):
                return line
        return ""

    @staticmethod
    def _extract_department(lines: List[str]) -> str:
        """提取部门信息

        Args:
            lines: 文本行列表

        Returns:
            部门信息或空字符串
        """
        for line in lines:
            # 部门关键词 + department关键字
            if any(kw in line for kw in DEPARTMENT_KEYWORDS):
                if 'department' in line.lower():
                    return line
        return ""

    @staticmethod
    def _extract_position(lines: List[str]) -> str:
        """提取职位信息

        Args:
            lines: 文本行列表

        Returns:
            职位信息或空字符串
        """
        for line in lines:
            if any(keyword in line for keyword in POSITION_KEYWORDS):
                return line
        return ""

    @staticmethod
    def _extract_address(lines: List[str]) -> str:
        """提取地址信息

        Args:
            lines: 文本行列表

        Returns:
            地址信息或空字符串
        """
        for line in lines:
            if any(kw in line for kw in ADDRESS_KEYWORDS):
                return line
        return ""

    @staticmethod
    def _clean_phone(phone_str: str) -> str:
        """清理电话号码"""
        # 去除非数字字符（保留+86）
        phone = re.sub(r'[^\d+]', '', phone_str)
        # 标准化格式
        if phone.startswith('86') and len(phone) > 10:
            phone = '+86' + phone[2:]
        return phone

    @staticmethod
    def _is_contact_info(line: str) -> bool:
        """判断是否为联系方式行"""
        return any(kw in line.lower() for kw in CONTACT_KEYWORDS)

    @staticmethod
    def _extract_name_from_email(email: str) -> str:
        """从邮箱提取可能的姓名"""
        name_part = email.split('@')[0]
        # 移除常见前缀
        name_part = re.sub(r'^(info|sales|support|admin|contact)', '', name_part)
        # 首字母大写
        return name_part.capitalize()

    def calculate_confidence(self, result: Dict) -> float:
        """计算识别置信度

        Args:
            result: 解析结果字典

        Returns:
            置信度值 (0-1)
        """
        confidence = 0.0

        for field, weight in CONFIDENCE_WEIGHTS.items():
            if result.get(field):
                confidence += weight

        return min(confidence, 1.0)


class OCRProcessor:
    """OCR处理器"""

    def __init__(self, tesseract_path: Optional[str] = None):
        """初始化OCR处理器

        Args:
            tesseract_path: Tesseract可执行文件路径
                          打包时会设置为相对路径
        """
        self._setup_tesseract(tesseract_path)
        self.preprocessor = ImagePreprocessor()
        self.parser = BusinessCardParser()
        self.batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")

    def _setup_tesseract(self, custom_path: Optional[str] = None):
        """设置Tesseract路径"""
        if custom_path:
            pytesseract.pytesseract.tesseract_cmd = custom_path
        else:
            # 打包后的路径
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
                tesseract_exe = os.path.join(base_path, 'tesseract', 'tesseract.exe')
                if os.path.exists(tesseract_exe):
                    pytesseract.pytesseract.tesseract_cmd = tesseract_exe
            else:
                # 开发环境使用系统安装的Tesseract
                pytesseract.pytesseract.tesseract_cmd = 'tesseract'

    def recognize_image(self, image_path: str,
                       language: str = 'eng+chi_sim') -> OCRResult:
        """识别图片文字

        Args:
            image_path: 图片路径
            language: 识别语言，默认为中英文

        Returns:
            OCR识别结果
        """
        start_time = time.time()

        result = OCRResult()
        result.source_image = os.path.basename(image_path)

        try:
            # 图像预处理
            processed_img = self.preprocessor.preprocess(image_path)

            # OCR识别
            raw_text = pytesseract.image_to_string(
                processed_img,
                lang=language,
                config='--psm 6'  # 假设统一文本块
            )
            result.raw_text = raw_text

            # 名片解析
            parsed = self.parser.parse(raw_text)
            result.name = parsed.get('name', '')
            result.phone = parsed.get('phone', '')
            result.email = parsed.get('email', '')
            result.company = parsed.get('company', '')
            result.department = parsed.get('department', '')
            result.position = parsed.get('position', '')
            result.address = parsed.get('address', '')

            # 计算置信度 - 使用parser的方法
            result.confidence = self.parser.calculate_confidence(parsed)

            result.process_time = time.time() - start_time

        except Exception as e:
            result.raw_text = f"OCR识别失败: {str(e)}"

        return result

    def recognize_batch(self, image_paths: List[str],
                       progress_callback: Optional[callable] = None) -> List[OCRResult]:
        """批量识别图片

        Args:
            image_paths: 图片路径列表
            progress_callback: 进度回调函数，参数为 (current, total)

        Returns:
            识别结果列表
        """
        results = []
        total = len(image_paths)

        for idx, path in enumerate(image_paths):
            result = self.recognize_image(path)
            results.append(result)

            # 调用进度回调
            if progress_callback:
                progress_callback(idx + 1, total)

        return results

    def recognize_and_save(self, image_path: str,
                          output_dir: Optional[str] = None) -> Tuple[str, OCRResult]:
        """识别图片并保存结果

        Args:
            image_path: 图片路径
            output_dir: 输出目录

        Returns:
            (Excel文件路径, 识别结果)
        """
        # 识别
        result = self.recognize_image(image_path)

        # 保存图片副本
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            # 保存处理后的图片
            processed_img = self.preprocessor.preprocess(image_path)
            img_name = os.path.basename(image_path)
            img_path = os.path.join(output_dir, f"processed_{img_name}")
            processed_img.save(img_path)

        # 导出为Excel
        if output_dir is None:
            output_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                'data', 'imports', 'ocr_images'
            )
        os.makedirs(output_dir, exist_ok=True)

        excel_path = os.path.join(
            output_dir,
            f"ocr_results_{self.batch_id}.xlsx"
        )

        self._export_to_excel([result], excel_path)

        return excel_path, result

    @staticmethod
    def _export_to_excel(results: List[OCRResult], output_path: str):
        """导出OCR结果到Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "OCR识别结果"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        headers = ['姓名', '电话', '邮箱', '公司', '部门', '职位', '地址', '置信度', '原图']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row_idx, result in enumerate(results, 2):
            values = [
                result.name, result.phone, result.email, result.company,
                result.department, result.position, result.address,
                f"{result.confidence:.1%}", result.source_image
            ]
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        wb.save(output_path)

    def get_supported_formats(self) -> List[str]:
        """获取支持的图片格式"""
        return ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif']

    def is_supported_format(self, file_path: str) -> bool:
        """检查是否支持该格式"""
        ext = os.path.splitext(file_path)[1].lower()
        return ext in self.get_supported_formats()
