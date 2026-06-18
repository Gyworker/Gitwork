# -*- coding: utf-8 -*-
"""
Excel导入模块
支持从Excel文件导入通讯录数据

版本：V4.1 (统一错误处理版)
"""

import os
import json
from datetime import datetime
from typing import List, Dict, Optional, Callable, Any
from dataclasses import dataclass, field
from enum import Enum

import pandas as pd
import openpyxl

# 导入项目模块
from src.database.contacts import ContactsDB
from src.utils.excel_utils import (
    create_excel_header,
    auto_adjust_column_width,
    write_data_rows,
    create_template,
    create_header_style,
    apply_style_to_cell,
    dict_to_row,
)
from src.utils.exceptions import AppException


# =============================================================================
# 异常定义
# =============================================================================

class ExcelImportError(AppException):
    """Excel导入异常基类"""
    pass


class FileFormatError(ExcelImportError):
    """文件格式错误"""
    pass


class HeaderMappingError(ExcelImportError):
    """表头映射错误"""
    pass


class DataParseError(ExcelImportError):
    """数据解析错误"""
    pass


# =============================================================================
# 错误处理枚举
# =============================================================================

class ErrorAction(Enum):
    """错误处理动作"""
    ABORT = "abort"      # 中止
    SKIP = "skip"        # 跳过
    PROCEED = "proceed"  # 继续


# =============================================================================
# 错误记录
# =============================================================================

@dataclass
class ImportError:
    """导入错误记录"""
    row_number: int           # 行号
    field_name: str = ""     # 字段名
    original_value: Any = ""  # 原始值
    error_type: str = ""      # 错误类型
    error_message: str = ""  # 错误信息
    
    def to_dict(self) -> Dict:
        return {
            'row': self.row_number,
            'field': self.field_name,
            'value': str(self.original_value),
            'type': self.error_type,
            'message': self.error_message,
        }


# =============================================================================
# 导入结果
# =============================================================================

@dataclass
class ImportResult:
    """导入结果"""
    success: bool = False
    total: int = 0
    added: int = 0
    duplicate: int = 0
    skipped: int = 0
    errors: List[ImportError] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    records: List[Dict] = field(default_factory=list)
    file_path: str = ""
    import_time: float = 0.0
    
    def add_error(self, row: int, field: str, value: Any, 
                  error_type: str, message: str):
        """添加错误"""
        self.errors.append(ImportError(
            row_number=row,
            field_name=field,
            original_value=value,
            error_type=error_type,
            error_message=message
        ))
    
    def add_warning(self, message: str):
        """添加警告"""
        self.warnings.append(message)
    
    def has_errors(self) -> bool:
        """是否有错误"""
        return len(self.errors) > 0
    
    def has_warnings(self) -> bool:
        """是否有警告"""
        return len(self.warnings) > 0
    
    def get_error_summary(self) -> str:
        """获取错误摘要"""
        if not self.errors:
            return ""
        
        summary = f"共发现 {len(self.errors)} 个错误:\n"
        error_types = {}
        for err in self.errors:
            key = err.error_type
            if key not in error_types:
                error_types[key] = 0
            error_types[key] += 1
        
        for err_type, count in error_types.items():
            summary += f"  - {err_type}: {count}个\n"
        
        return summary.strip()
    
    def to_dict(self) -> Dict:
        """转换为字典"""
        return {
            'success': self.success,
            'total': self.total,
            'added': self.added,
            'duplicate': self.duplicate,
            'skipped': self.skipped,
            'error_count': len(self.errors),
            'warning_count': len(self.warnings),
            'errors': [e.to_dict() for e in self.errors],
            'warnings': self.warnings,
        }


# =============================================================================
# 错误处理器
# =============================================================================

class ImportErrorHandler:
    """
    导入错误处理器
    统一管理导入过程中的错误处理
    """
    
    def __init__(self, 
                 on_error: Optional[Callable[[ImportError], ErrorAction]] = None,
                 max_errors: int = 100):
        """
        初始化错误处理器
        
        Args:
            on_error: 错误回调，返回处理动作
            max_errors: 最大错误数
        """
        self.on_error = on_error
        self.max_errors = max_errors
        self.errors: List[ImportError] = []
        self.warnings: List[str] = []
        self._should_stop = False
    
    def handle_error(self, row: int, field: str, value: Any,
                    error_type: str, message: str) -> ErrorAction:
        """
        处理错误
        
        Args:
            row: 行号
            field: 字段名
            value: 原始值
            error_type: 错误类型
            message: 错误信息
            
        Returns:
            处理动作
        """
        error = ImportError(
            row_number=row,
            field_name=field,
            original_value=value,
            error_type=error_type,
            error_message=message
        )
        
        self.errors.append(error)
        
        # 检查是否超过最大错误数
        if len(self.errors) >= self.max_errors:
            self._should_stop = True
            return ErrorAction.ABORT
        
        # 调用回调
        if self.on_error:
            return self.on_error(error)
        
        return ErrorAction.SKIP
    
    def add_warning(self, message: str):
        """添加警告"""
        self.warnings.append(message)
    
    def should_stop(self) -> bool:
        """是否应该停止"""
        return self._should_stop
    
    def get_result(self) -> ImportResult:
        """获取结果"""
        return ImportResult(
            success=not self._should_stop,
            errors=self.errors.copy(),
            warnings=self.warnings.copy(),
        )
    
    def reset(self):
        """重置"""
        self.errors.clear()
        self.warnings.clear()
        self._should_stop = False


# =============================================================================
# Excel联系人数据结构
# =============================================================================

@dataclass
class ExcelContact:
    """Excel联系人数据结构"""
    姓名: str = ""
    电话: str = ""
    邮箱: str = ""
    公司: str = ""
    部门: str = ""
    职位: str = ""
    备注: str = ""
    source: str = "excel"
    import_batch: str = ""
    
    def to_dict(self) -> Dict:
        """转换为字典"""
        return {
            '姓名': self.姓名,
            '电话': self.电话,
            '邮箱': self.邮箱,
            '公司': self.公司,
            '部门': self.部门,
            '职位': self.职位,
            '备注': self.备注,
            'source': self.source,
            'import_batch': self.import_batch,
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'ExcelContact':
        """从字典创建"""
        return cls(
            姓名=data.get('姓名', ''),
            电话=data.get('电话', ''),
            邮箱=data.get('邮箱', ''),
            公司=data.get('公司', ''),
            部门=data.get('部门', ''),
            职位=data.get('职位', ''),
            备注=data.get('备注', ''),
            source=data.get('source', 'excel'),
            import_batch=data.get('import_batch', ''),
        )
    
    def validate(self) -> List[str]:
        """验证数据"""
        errors = []
        if not self.姓名:
            errors.append("姓名为空")
        return errors


# =============================================================================
# Excel表头映射器
# =============================================================================

class ExcelHeaderMapper:
    """Excel表头映射器"""
    
    HEADER_MAPPINGS = {
        '姓名': ['姓名', 'name', '名称', '名字', 'contact'],
        '电话': ['电话', 'phone', '手机', 'mobile', 'tel', 'telephone', '号码'],
        '邮箱': ['邮箱', 'email', '邮件', 'e-mail', 'mail'],
        '公司': ['公司', 'company', '企业', 'organization', 'org'],
        '部门': ['部门', 'department', 'dept', '科室'],
        '职位': ['职位', 'position', 'title', '职务', 'job'],
        '备注': ['备注', 'remark', 'note', 'notes', '说明'],
    }
    
    @classmethod
    def find_header_mapping(cls, headers: List[str]) -> Dict[str, int]:
        """智能查找表头映射"""
        mapping = {}
        normalized_headers = [h.strip().lower() for h in headers]
        
        for field, aliases in cls.HEADER_MAPPINGS.items():
            for idx, header in enumerate(normalized_headers):
                if header in [a.lower() for a in aliases]:
                    mapping[field] = idx
                    break
        
        return mapping
    
    @classmethod
    def validate_mapping(cls, mapping: Dict[str, int]) -> tuple:
        """验证映射"""
        if '姓名' not in mapping:
            return False, "无法识别姓名列"
        return True, ""


# =============================================================================
# Excel导入器
# =============================================================================

class ExcelImporter:
    """Excel导入器 - 统一错误处理版"""
    
    def __init__(self, db: ContactsDB,
                 error_handler: Optional[ImportErrorHandler] = None):
        self.db = db
        self.batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.last_import_path = ""
        self._import_records: List[ExcelContact] = []
        self._error_handler = error_handler or ImportErrorHandler()
    
    def import_from_file(self, file_path: str,
                        sheet_name: Optional[str] = None,
                        header_row: int = 1,
                        progress_callback: Optional[Callable[[int, int], None]] = None) -> ImportResult:
        """从Excel文件导入通讯录"""
        import time
        start_time = time.time()
        
        result = ImportResult()
        result.file_path = file_path
        
        try:
            # 验证文件格式
            if not self._validate_file(file_path, result):
                return result
            
            # 读取Excel文件
            df = self._read_excel(file_path, sheet_name, header_row, result)
            if result.errors:
                return result
            
            # 获取表头映射
            headers = df.columns.tolist()
            header_map = ExcelHeaderMapper.find_header_mapping(headers)
            
            # 验证映射
            is_valid, error_msg = ExcelHeaderMapper.validate_mapping(header_map)
            if not is_valid:
                result.add_error(0, "header", "", "HeaderMapping", error_msg)
                return result
            
            # 重置错误处理器
            self._error_handler.reset()
            
            # 处理数据
            total_rows = len(df)
            for idx, row in df.iterrows():
                self._process_row(idx, row, header_map, result)
                
                if self._error_handler.should_stop():
                    break
                
                if progress_callback:
                    progress_callback(idx + 1, total_rows)
            
            # 检测重复
            unique_records = self._detect_duplicates(result.records)
            result.added = len(unique_records)
            result.duplicate = result.total - len(unique_records)
            
            # 保存导入记录
            self._import_records = [ExcelContact.from_dict(r) for r in unique_records]
            self.last_import_path = file_path
            result.success = True
            
        except Exception as e:
            result.add_error(0, "", "", "Unexpected", f"导入失败: {str(e)}")
        
        result.import_time = time.time() - start_time
        return result
    
    def _validate_file(self, file_path: str, result: ImportResult) -> bool:
        """验证文件"""
        if not os.path.exists(file_path):
            result.add_error(0, "file", file_path, "FileNotFound", "文件不存在")
            return False
        
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in ['.xlsx', '.xls']:
            result.add_error(0, "file", file_path, "FileFormat", 
                           f"不支持的文件格式: {ext}")
            return False
        
        return True
    
    def _read_excel(self, file_path: str, sheet_name: Optional[str],
                   header_row: int, result: ImportResult) -> pd.DataFrame:
        """读取Excel文件"""
        try:
            engine = 'openpyxl' if file_path.endswith('.xlsx') else 'xlrd'
            df = pd.read_excel(file_path, sheet_name=sheet_name,
                              header=header_row - 1, engine=engine)
            return df
        except Exception as e:
            result.add_error(0, "file", file_path, "ReadError", 
                           f"文件读取失败: {str(e)}")
            return pd.DataFrame()
    
    def _process_row(self, idx: int, row: pd.Series, 
                   header_map: Dict[str, int], result: ImportResult):
        """处理单行数据"""
        row_num = idx + 2
        
        try:
            contact = self._parse_row(row, header_map, result)
            errors = contact.validate()
            if errors:
                for error in errors:
                    result.add_error(row_num, "姓名", contact.姓名, "Validation", error)
                return
            
            if contact.姓名:
                result.records.append(contact.to_dict())
                result.total += 1
                
        except Exception as e:
            result.add_error(row_num, "", "", "ParseError", f"解析错误: {str(e)}")
    
    def _parse_row(self, row: pd.Series, header_map: Dict[str, int],
                  result: ImportResult) -> ExcelContact:
        """解析单行数据"""
        contact = ExcelContact()
        contact.import_batch = self.batch_id
        
        for field, col_idx in header_map.items():
            if col_idx < len(row):
                value = row.iloc[col_idx]
                try:
                    str_value = str(value).strip() if pd.notna(value) else ""
                    if str_value and str_value != 'nan':
                        setattr(contact, field, str_value)
                except Exception as e:
                    result.add_error(row.number, field, value, "Conversion", str(e))
        
        return contact
    
    def _detect_duplicates(self, records: List[Dict]) -> List[Dict]:
        """检测重复记录"""
        existing_contacts = self.db.get_all_contacts()
        existing_keys = set(
            f"{c.get('name', '')}|{c.get('phone', '')}"
            for c in existing_contacts
        )
        
        unique_records = []
        for record in records:
            key = f"{record.get('姓名', '')}|{record.get('电话', '')}"
            if key not in existing_keys:
                unique_records.append(record)
        
        return unique_records
    
    def save_to_database(self) -> int:
        """保存到数据库"""
        saved = 0
        for record in self._import_records:
            self.db.add_contact(
                name=record.姓名,
                phone=record.电话,
                email=record.邮箱,
                company=record.公司,
                department=record.部门,
                position=record.职位,
                remark=record.备注,
                source=record.source,
                import_batch=record.import_batch,
                raw_data=json.dumps(record.to_dict(), ensure_ascii=False)
            )
            saved += 1
        return saved
    
    def save_to_excel(self, output_path: Optional[str] = None) -> str:
        """保存导入记录到Excel文件"""
        if not self._import_records:
            raise ValueError("没有导入记录")
        
        if output_path is None:
            imports_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                'data', 'imports'
            )
            os.makedirs(imports_dir, exist_ok=True)
            output_path = os.path.join(
                imports_dir,
                f"contacts_{self.batch_id}.xlsx"
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "导入通讯录"
        
        headers = ['姓名', '电话', '邮箱', '公司', '部门', '职位', '备注', '导入批次']
        create_excel_header(ws, headers)
        
        for row_idx, record in enumerate(self._import_records, 2):
            values = [
                record.姓名, record.电话, record.邮箱, record.公司,
                record.部门, record.职位, record.备注, record.import_batch
            ]
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        auto_adjust_column_width(ws)
        wb.save(output_path)
        return output_path
    
    def load_from_excel(self, file_path: str) -> List[ExcelContact]:
        """从Excel文件读取已编辑的导入记录"""
        if file_path.endswith('.xlsx'):
            df = pd.read_excel(file_path, engine='openpyxl')
        else:
            df = pd.read_excel(file_path, engine='xlrd')
        
        headers = df.columns.tolist()
        header_map = ExcelHeaderMapper.find_header_mapping(headers)
        
        records = []
        for idx, row in df.iterrows():
            contact = self._parse_row(row, header_map, ImportResult())
            if contact.姓名:
                records.append(contact)
        
        self._import_records = records
        return records


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self, db: ContactsDB):
        self.db = db
    
    def export_all(self, output_path: str) -> str:
        """导出全部通讯录"""
        contacts = self.db.get_all_contacts()
        return self._create_excel(contacts, output_path)
    
    def export_filtered(self, output_path: str,
                       filters: Optional[Dict] = None) -> str:
        """导出筛选后的通讯录"""
        contacts = self.db.search_contacts(
            keyword=filters.get('keyword', '') if filters else '',
            company=filters.get('company', '') if filters else ''
        )
        return self._create_excel(contacts, output_path)
    
    def _create_excel(self, contacts: List[Dict], output_path: str) -> str:
        """创建Excel文件"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "通讯录"
        
        headers = ['姓名', '电话', '邮箱', '公司', '部门', '职位', '备注']
        create_excel_header(ws, headers)
        
        for row_idx, contact in enumerate(contacts, 2):
            values = [
                contact.get('name', ''),
                contact.get('phone', ''),
                contact.get('email', ''),
                contact.get('company', ''),
                contact.get('department', ''),
                contact.get('position', ''),
                contact.get('remark', ''),
            ]
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        auto_adjust_column_width(ws)
        wb.save(output_path)
        return output_path
    
    def create_template(self, output_path: str) -> str:
        """创建导入模板"""
        headers = ['姓名*', '电话', '邮箱', '公司', '部门', '职位', '备注']
        example_data = [
            ['张三', '13800138000', 'zhangsan@example.com', '示例公司', '市场部', '经理', 'VIP客户'],
            ['李四', '13900139000', 'lisi@example.com', '示例公司', '销售部', '主管', ''],
        ]
        
        return create_template(output_path, headers, example_data, "70AD47")
