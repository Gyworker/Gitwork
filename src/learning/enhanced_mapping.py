# -*- coding: utf-8 -*-
"""
增强映射学习模块
支持从Excel文件和历史数据学习映射关系
版本：V2.1 (优化版)
"""

import os
import re
import json
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, asdict
from collections import defaultdict
import pandas as pd
from difflib import SequenceMatcher

# 导入项目模块
from src.database.contacts import ContactsDB


@dataclass
class MappingRule:
    """映射规则"""
    id: int = 0
    keyword: str = ""
    responsible: str = ""
    confidence: float = 1.0
    source: str = "manual"  # 'manual', 'excel', 'history'
    usage_count: int = 0
    last_used: Optional[str] = None
    created_at: str = ""
    updated_at: str = ""

    def to_dict(self) -> Dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: Dict) -> 'MappingRule':
        return cls(**data)


# =============================================================================
# 停用词定义 - V2.1完善版
# =============================================================================

# 中文停用词
CHINESE_STOP_WORDS = {
    # 常见助词
    '的', '了', '在', '是', '我', '有', '和', '就', '不', '人', '都', '一', '一个',
    '上', '也', '很', '到', '说', '要', '去', '你', '会', '着', '没有', '看', '好',
    '自己', '这', '那', '里', '为', '而且', '但是', '所以', '如果', '因为', '还是',
    # 常见动词
    '做', '进行', '使用', '需要', '可以', '能够', '应该', '必须', '需要', '希望',
    # 常见副词
    '很', '非常', '特别', '比较', '最', '更', '还', '再', '又', '也', '都', '只',
    # 常见介词
    '从', '到', '把', '被', '让', '给', '向', '对', '关于', '对于',
    # 常见连词
    '和', '或', '但', '而', '则', '因此', '所以', '于是', '既然',
    # 常见量词
    '个', '些', '种', '类', '点', '次', '件', '条', '把', '位',
    # 其他常见词
    '什么', '怎么', '如何', '怎样', '哪', '哪个', '哪些', '谁', '多少', '几',
    '这个', '那个', '这些', '那些', '这样', '那样', '如此',
}

# 英文停用词
ENGLISH_STOP_WORDS = {
    'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for',
    'of', 'with', 'by', 'from', 'as', 'is', 'was', 'are', 'were', 'been',
    'be', 'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could',
    'should', 'may', 'might', 'must', 'can', 'this', 'that', 'these', 'those',
    'i', 'you', 'he', 'she', 'it', 'we', 'they', 'what', 'which', 'who',
    'when', 'where', 'why', 'how', 'all', 'each', 'every', 'both', 'few',
    'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only',
    'own', 'same', 'so', 'than', 'too', 'very', 'just',
}

# 合并所有停用词
ALL_STOP_WORDS = CHINESE_STOP_WORDS | ENGLISH_STOP_WORDS


class MappingLearner:
    """映射学习器"""

    def __init__(self, db: ContactsDB):
        self.db = db
        self.rules: Dict[str, MappingRule] = {}
        self.stop_words = ALL_STOP_WORDS  # 使用完善的停用词列表
        self._load_rules()

    def _load_rules(self):
        """从数据库加载映射规则"""
        try:
            rules_data = self.db.execute_query(
                "SELECT * FROM mapping_rules ORDER BY usage_count DESC"
            )
            for row in rules_data:
                rule = MappingRule(
                    id=row['id'],
                    keyword=row['keyword'],
                    responsible=row['responsible'],
                    confidence=row.get('confidence', 1.0),
                    source=row.get('source', 'manual'),
                    usage_count=row.get('usage_count', 0),
                    last_used=row.get('last_used'),
                    created_at=row.get('created_at', ''),
                    updated_at=row.get('updated_at', '')
                )
                self.rules[rule.keyword.lower()] = rule
        except Exception:
            # 表可能不存在，忽略
            pass

    def _ensure_table_exists(self):
        """确保映射规则表存在"""
        try:
            self.db.execute_query("""
                CREATE TABLE IF NOT EXISTS mapping_rules (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    keyword VARCHAR(200) NOT NULL,
                    responsible VARCHAR(100) NOT NULL,
                    confidence FLOAT DEFAULT 1.0,
                    source VARCHAR(50) DEFAULT 'manual',
                    usage_count INTEGER DEFAULT 0,
                    last_used VARCHAR(50),
                    created_at VARCHAR(50),
                    updated_at VARCHAR(50)
                )
            """)
        except Exception:
            pass

    def learn_from_excel(self, file_path: str,
                        keyword_col: str = '关键词',
                        responsible_col: str = '责任人') -> int:
        """从Excel文件学习映射

        Args:
            file_path: Excel文件路径
            keyword_col: 关键词列名
            responsible_col: 责任人列名

        Returns:
            学习到的规则数量
        """
        self._ensure_table_exists()

        # 读取Excel
        if file_path.endswith('.xlsx'):
            df = pd.read_excel(file_path, engine='openpyxl')
        else:
            df = pd.read_excel(file_path, engine='xlrd')

        # 查找列
        headers = df.columns.tolist()
        keyword_idx = self._find_column(headers, keyword_col, ['关键词', 'keyword', '模块', 'task'])
        responsible_idx = self._find_column(headers, responsible_col, ['责任人', 'responsible', '负责人', 'owner'])

        if keyword_idx is None or responsible_idx is None:
            raise ValueError("无法识别关键词或责任人列")

        # 学习规则
        count = 0
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for _, row in df.iterrows():
            keyword = str(row.iloc[keyword_idx]).strip()
            responsible = str(row.iloc[responsible_idx]).strip()

            if keyword and keyword != 'nan' and responsible and responsible != 'nan':
                self._add_rule(
                    keyword=keyword,
                    responsible=responsible,
                    source='excel'
                )
                count += 1

        return count

    def learn_from_text(self, text: str, responsible: str) -> MappingRule:
        """从文本学习映射

        Args:
            text: 包含关键词的文本
            responsible: 责任人

        Returns:
            创建的映射规则
        """
        self._ensure_table_exists()

        # 提取关键词
        keywords = self._extract_keywords(text)
        if keywords:
            keyword = keywords[0]
        else:
            keyword = text[:50]  # 使用前50个字符

        return self._add_rule(keyword, responsible, source='manual')

    def learn_from_history(self, tasks: List[Dict]) -> int:
        """从历史任务数据学习映射

        Args:
            tasks: 历史任务列表，每个任务包含 module, responsible 等字段

        Returns:
            学习到的规则数量
        """
        self._ensure_table_exists()

        # 统计模块-责任人的出现频率
        module_responsible_count = defaultdict(lambda: defaultdict(int))

        for task in tasks:
            module = task.get('module', '')
            responsible = task.get('responsible', '')
            if module and responsible:
                module_responsible_count[module][responsible] += 1

        # 学习频率最高的映射
        count = 0
        for module, responsible_dict in module_responsible_count.items():
            if responsible_dict:
                # 选择出现频率最高的
                best_responsible = max(responsible_dict.items(), key=lambda x: x[1])
                self._add_rule(
                    keyword=module,
                    responsible=best_responsible[0],
                    confidence=min(best_responsible[1] / 10, 1.0),  # 根据频率调整置信度
                    source='history'
                )
                count += 1

        return count

    def _find_column(self, headers: List[str], target: str, alternatives: List[str]) -> Optional[int]:
        """查找匹配的列索引"""
        target_lower = target.lower()
        for idx, header in enumerate(headers):
            header_lower = header.strip().lower()
            if header_lower == target_lower or header_lower in alternatives:
                return idx
        return None

    def _extract_keywords(self, text: str) -> List[str]:
        """从文本提取关键词 - 使用完善的停用词列表

        Args:
            text: 文本内容

        Returns:
            关键词列表
        """
        # 提取中文和英文词
        chinese_words = re.findall(r'[\u4e00-\u9fa5]+', text)
        english_words = re.findall(r'[a-zA-Z]+', text)
        numbers = re.findall(r'\d+', text)

        # 过滤停用词
        chinese_keywords = [
            w for w in chinese_words
            if len(w) >= 2 and w not in self.stop_words
        ]
        english_keywords = [
            w.lower() for w in english_words
            if len(w) >= 2 and w.lower() not in self.stop_words
        ]

        # 数字序列（保留作为关键词）
        number_keywords = [n for n in numbers if len(n) >= 2]

        # 合并结果
        keywords = chinese_keywords + english_keywords + number_keywords

        # 去重
        seen = set()
        unique_keywords = []
        for kw in keywords:
            kw_lower = kw.lower() if isinstance(kw, str) else str(kw)
            if kw_lower not in seen:
                seen.add(kw_lower)
                unique_keywords.append(kw)

        return unique_keywords

    def _add_rule(self, keyword: str, responsible: str,
                  confidence: float = 1.0, source: str = 'manual') -> MappingRule:
        """添加映射规则"""
        keyword_lower = keyword.lower()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if keyword_lower in self.rules:
            # 更新现有规则
            rule = self.rules[keyword_lower]
            rule.responsible = responsible
            rule.confidence = max(rule.confidence, confidence)
            rule.source = source
            rule.updated_at = now
        else:
            # 创建新规则
            rule = MappingRule(
                keyword=keyword,
                responsible=responsible,
                confidence=confidence,
                source=source,
                created_at=now,
                updated_at=now
            )
            self.rules[keyword_lower] = rule

        # 保存到数据库
        self._save_rule(rule)

        return rule

    def _save_rule(self, rule: MappingRule):
        """保存规则到数据库"""
        try:
            self.db.execute_query(
                """INSERT OR REPLACE INTO mapping_rules
                   (keyword, responsible, confidence, source, usage_count, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (rule.keyword, rule.responsible, rule.confidence, rule.source,
                 rule.usage_count, rule.created_at, rule.updated_at)
            )
        except Exception:
            pass

    def recommend(self, task_text: str) -> Tuple[Optional[str], float]:
        """推荐责任人

        Args:
            task_text: 任务描述文本

        Returns:
            (推荐的责任人, 置信度)
        """
        # 提取关键词
        keywords = self._extract_keywords(task_text)
        if not keywords:
            return None, 0.0

        # 匹配规则
        best_match = self._find_best_match(keywords)

        # 更新使用统计
        if best_match[0]:
            self._update_usage(best_match[0])

        return best_match

    def _find_best_match(self, keywords: List[str]) -> Tuple[Optional[str], float]:
        """查找最佳匹配

        Args:
            keywords: 关键词列表

        Returns:
            (责任人, 置信度)
        """
        best_match = None
        best_score = 0.0

        for keyword in keywords:
            keyword_lower = keyword.lower()

            # 精确匹配
            exact_result = self._exact_match(keyword_lower)
            if exact_result and exact_result[1] > best_score:
                best_score = exact_result[1]
                best_match = exact_result[0]

            # 模糊匹配
            fuzzy_result = self._fuzzy_match(keyword_lower)
            if fuzzy_result and fuzzy_result[1] > best_score:
                best_score = fuzzy_result[1]
                best_match = fuzzy_result[0]

        return best_match, best_score

    def _exact_match(self, keyword: str) -> Optional[Tuple[str, float]]:
        """精确匹配

        Args:
            keyword: 关键词

        Returns:
            (责任人, 置信度) 或 None
        """
        if keyword in self.rules:
            rule = self.rules[keyword]
            # 精确匹配权重为1.0
            return rule.responsible, rule.confidence * 1.0
        return None

    def _fuzzy_match(self, keyword: str, threshold: float = 0.8) -> Optional[Tuple[str, float]]:
        """模糊匹配

        Args:
            keyword: 关键词
            threshold: 相似度阈值

        Returns:
            (责任人, 置信度) 或 None
        """
        best_score = 0.0
        best_responsible = None

        for rule_key, rule in self.rules.items():
            similarity = SequenceMatcher(None, keyword, rule_key).ratio()
            if similarity >= threshold:
                # 模糊匹配权重为0.8
                score = rule.confidence * similarity * 0.8
                if score > best_score:
                    best_score = score
                    best_responsible = rule.responsible

        if best_responsible:
            return best_responsible, best_score
        return None

    def _update_usage(self, responsible: str):
        """更新使用统计"""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for rule in self.rules.values():
            if rule.responsible == responsible:
                rule.usage_count += 1
                rule.last_used = now
                self._save_rule(rule)

    def get_all_rules(self) -> List[MappingRule]:
        """获取所有映射规则"""
        return sorted(self.rules.values(), key=lambda x: x.usage_count, reverse=True)

    def export_to_excel(self, output_path: str) -> str:
        """导出映射规则到Excel

        Args:
            output_path: 输出文件路径

        Returns:
            保存的文件路径
        """
        from openpyxl import Workbook
        from src.utils.excel_utils import (
            create_excel_header,
            auto_adjust_column_width,
            create_header_style,
            apply_style_to_cell,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "映射规则"

        # 使用公共工具
        headers = ['关键词', '责任人', '置信度', '来源', '使用次数', '最后使用', '创建时间']
        style = create_header_style()

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            apply_style_to_cell(cell, style)

        # 写入数据
        rules = self.get_all_rules()
        for row_idx, rule in enumerate(rules, 2):
            values = [
                rule.keyword, rule.responsible, f"{rule.confidence:.1%}",
                rule.source, rule.usage_count, rule.last_used or '-', rule.created_at
            ]
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动调整列宽
        auto_adjust_column_width(ws)

        wb.save(output_path)
        return output_path

    def delete_rule(self, keyword: str) -> bool:
        """删除映射规则

        Args:
            keyword: 要删除的关键词

        Returns:
            是否删除成功
        """
        keyword_lower = keyword.lower()
        if keyword_lower in self.rules:
            del self.rules[keyword_lower]
            try:
                self.db.execute_query(
                    "DELETE FROM mapping_rules WHERE LOWER(keyword) = ?",
                    (keyword_lower,)
                )
                return True
            except Exception:
                return False
        return False

    def add_stop_words(self, words: List[str]):
        """添加自定义停用词

        Args:
            words: 停用词列表
        """
        self.stop_words.update(words)

    def get_stop_words(self) -> set:
        """获取当前停用词集合

        Returns:
            停用词集合
        """
        return self.stop_words.copy()
