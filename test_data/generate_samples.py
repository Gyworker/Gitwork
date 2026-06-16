# -*- coding: utf-8 -*-
"""
生成测试样本数据
运行此脚本生成测试用的Excel和图片样本
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 测试数据目录
TEST_DATA_DIR = Path(__file__).parent
SAMPLES_DIR = TEST_DATA_DIR / "samples"


def create_sample_excel():
    """创建示例Excel通讯录"""
    wb = Workbook()
    ws = wb.active
    ws.title = "通讯录"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    headers = ['姓名', '电话', '邮箱', '公司', '部门', '职位', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入测试数据
    test_data = [
        ['张三', '13800138000', 'zhangsan@example.com', '科技有限公司', '市场部', '经理', 'VIP客户'],
        ['李四', '13900139000', 'lisi@example.com', '实业集团', '销售部', '主管', ''],
        ['王五', '13700137000', 'wangwu@example.com', '互联网公司', '技术部', '工程师', '技术骨干'],
        ['赵六', '13600136000', 'zhaoliu@example.com', '贸易公司', '采购部', '专员', ''],
        ['钱七', '13500135000', 'qianqi@example.com', '金融集团', '财务部', '总监', ''],
    ]

    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 30)

    # 保存
    output_path = SAMPLES_DIR / "sample_contacts.xlsx"
    wb.save(output_path)
    print(f"✓ 创建示例通讯录: {output_path}")
    return output_path


def create_english_headers_excel():
    """创建英文表头的Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    headers = ['name', 'phone', 'email', 'company', 'department', 'position', 'remark']
    test_data = [
        ['John Doe', '+1-234-567-8900', 'john@example.com', 'Tech Corp', 'Marketing', 'Director', 'Key client'],
        ['Jane Smith', '+1-234-567-8901', 'jane@example.com', 'Sales Inc', 'Sales', 'Manager', ''],
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)

    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    output_path = SAMPLES_DIR / "sample_contacts_english.xlsx"
    wb.save(output_path)
    print(f"✓ 创建英文表头通讯录: {output_path}")
    return output_path


def create_mixed_headers_excel():
    """创建混合表头的Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed Headers"

    # 混合中英文表头
    headers = ['姓名', 'mobile', 'Email', '公司', 'dept', 'title', '备注']
    test_data = [
        ['孙八', '15000150000', 'sunba@example.com', '混合公司', '运营部', '经理', ''],
        ['周九', '15100151000', 'zhoujiu@example.com', '测试企业', 'HR', '主管', '招聘负责人'],
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)

    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    output_path = SAMPLES_DIR / "sample_contacts_mixed.xlsx"
    wb.save(output_path)
    print(f"✓ 创建混合表头通讯录: {output_path}")
    return output_path


def create_mapping_rules_excel():
    """创建映射规则Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "映射规则"

    headers = ['关键词', '责任人', '置信度', '来源']
    test_data = [
        ['市场推广', '张三', 0.95, 'excel'],
        ['技术开发', '李四', 0.90, 'excel'],
        ['产品设计', '王五', 0.85, 'excel'],
        ['财务审计', '赵六', 0.88, 'excel'],
        ['人力资源', '钱七', 0.82, 'excel'],
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)

    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    output_path = SAMPLES_DIR / "sample_mapping_rules.xlsx"
    wb.save(output_path)
    print(f"✓ 创建映射规则Excel: {output_path}")
    return output_path


def create_template_excel():
    """创建导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    # 绿色表头（模板风格）
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    headers = ['姓名*', '电话', '邮箱', '公司', '部门', '职位', '备注']
    example_data = [
        ['张三', '13800138000', 'zhangsan@example.com', '示例公司', '市场部', '经理', 'VIP客户'],
        ['李四', '13900139000', 'lisi@example.com', '示例公司', '销售部', '主管', ''],
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    output_path = SAMPLES_DIR / "import_template.xlsx"
    wb.save(output_path)
    print(f"✓ 创建导入模板: {output_path}")
    return output_path


def create_test_images_info():
    """创建测试图片信息文件（OCR测试需要手动准备图片）"""
    info = """
# OCR测试图片准备说明

为了完整测试OCR功能，请准备以下测试图片：

## 图片1: 清晰中文名片 (test_card_1.jpg)
- 内容：姓名、电话、邮箱、公司、职位
- 格式：JPG/PNG
- 分辨率：建议 300 DPI 以上
- 预期识别准确率：> 90%

## 图片2: 英文名片 (test_card_2.jpg)
- 内容：Name、Phone、Email、Company
- 格式：JPG/PNG
- 预期识别准确率：> 85%

## 图片3: 低质量图片 (test_card_blurry.jpg)
- 用途：测试降噪处理
- 模糊/有噪点
- 预期识别准确率：< 70%

## 图片4: 多名片合集 (test_multiple.jpg)
- 多张名片在同一图片
- 测试批量识别功能

## 放置位置
将测试图片放置在: test_data/images/ 目录
"""
    output_path = SAMPLES_DIR / "ocr_test_images_README.txt"
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(info)
    print(f"✓ 创建OCR测试说明: {output_path}")
    return output_path


def main():
    """主函数"""
    # 确保目录存在
    SAMPLES_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 50)
    print("  生成测试样本数据")
    print("=" * 50)
    print()

    # 创建Excel样本
    create_sample_excel()
    create_english_headers_excel()
    create_mixed_headers_excel()
    create_mapping_rules_excel()
    create_template_excel()

    # 创建OCR测试说明
    create_test_images_info()

    print()
    print("=" * 50)
    print("  样本数据生成完成！")
    print("=" * 50)
    print()
    print("样本文件位置: test_data/samples/")
    print()
    print("下一步:")
    print("1. 运行单元测试: python run_tests.py unit")
    print("2. 运行系统测试: python run_tests.py system")
    print("3. 运行集成测试: python run_tests.py integration")
    print("4. 完整测试: python run_tests.py")


if __name__ == "__main__":
    main()
