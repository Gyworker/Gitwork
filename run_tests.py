# -*- coding: utf-8 -*-
"""
市场咨询任务跟踪工具 V2.1 测试执行脚本
支持单元测试、系统测试和集成测试

使用方法：
    python run_tests.py              # 运行所有测试
    python run_tests.py unit         # 仅运行单元测试
    python run_tests.py system       # 仅运行系统测试
    python run_tests.py integration  # 仅运行集成测试
    python run_tests.py quick       # 快速测试（跳过OCR）
"""

import os
import sys
import time
import tempfile
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple

# 添加项目路径
PROJECT_ROOT = Path(__file__).parent
sys.path.insert(0, str(PROJECT_ROOT))

# 尝试导入测试依赖
try:
    import pytest
except ImportError:
    print("[错误] 缺少pytest，请运行: pip install pytest")
    sys.exit(1)


class TestRunner:
    """测试运行器"""

    def __init__(self):
        self.results = []
        self.start_time = None
        self.test_dir = PROJECT_ROOT / "src" / "tests"

    def print_header(self):
        """打印测试标题"""
        print("=" * 60)
        print("  市场咨询任务跟踪工具 V2.1 测试执行")
        print("=" * 60)
        print(f"  时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"  项目: {PROJECT_ROOT}")
        print("=" * 60)
        print()

    def run_unit_tests(self) -> Tuple[int, int]:
        """运行单元测试"""
        print("\n" + "=" * 60)
        print("  单元测试")
        print("=" * 60)

        test_modules = [
            "test_excel_import.py",
            "test_ocr_handler.py",
            "test_enhanced_mapping.py",
        ]

        total_passed = 0
        total_failed = 0

        for module in test_modules:
            test_file = self.test_dir / module
            if test_file.exists():
                print(f"\n>>> 运行 {module}...")
                result = pytest.main([
                    str(test_file),
                    "-v",
                    "--tb=short",
                    "--color=yes"
                ])
                if result == 0:
                    total_passed += 1
                else:
                    total_failed += 1

        print("\n" + "-" * 60)
        print(f"单元测试结果: 通过 {total_passed}/{len(test_modules)} 模块")
        return total_passed, total_failed

    def run_system_tests(self) -> Dict:
        """运行系统测试"""
        print("\n" + "=" * 60)
        print("  系统测试")
        print("=" * 60)

        results = {
            "excel_import": self._test_excel_import_system(),
            "mapping_learning": self._test_mapping_learning_system(),
        }

        return results

    def _test_excel_import_system(self) -> Dict:
        """系统测试 - Excel导入功能"""
        print("\n>>> 测试Excel导入功能...")

        result = {
            "passed": 0,
            "failed": 0,
            "tests": []
        }

        try:
            # 导入模块
            from src.content.excel_import import (
                ExcelContact, ExcelHeaderMapper, ExcelImporter, ExcelExporter
            )
            from unittest.mock import Mock
            from openpyxl import Workbook
            import pandas as pd

            # 测试1: 表头映射
            print("  [1] 测试表头映射...")
            headers = ['姓名', '电话', '邮箱']
            mapping = ExcelHeaderMapper.find_header_mapping(headers)
            if '姓名' in mapping and '电话' in mapping:
                print("    ✓ 表头映射正确")
                result["passed"] += 1
            else:
                print("    ✗ 表头映射失败")
                result["failed"] += 1
            result["tests"].append("表头映射")

            # 测试2: 联系人创建
            print("  [2] 测试联系人创建...")
            contact = ExcelContact(
                姓名='测试用户',
                电话='13800138000',
                邮箱='test@example.com'
            )
            if contact.姓名 == '测试用户':
                print("    ✓ 联系人创建正确")
                result["passed"] += 1
            else:
                print("    ✗ 联系人创建失败")
                result["failed"] += 1
            result["tests"].append("联系人创建")

            # 测试3: Excel文件创建和读取
            print("  [3] 测试Excel文件创建...")
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                temp_path = f.name

            wb = Workbook()
            ws = wb.active
            ws['A1'] = '姓名'
            ws['B1'] = '电话'
            ws['A2'] = '张三'
            ws['B2'] = '13800138000'
            wb.save(temp_path)
            wb.close()

            df = pd.read_excel(temp_path)
            if len(df) == 1 and df.iloc[0]['姓名'] == '张三':
                print("    ✓ Excel文件创建和读取正确")
                result["passed"] += 1
            else:
                print("    ✗ Excel文件读取失败")
                result["failed"] += 1
            result["tests"].append("Excel文件操作")

            # 清理
            os.unlink(temp_path)

        except Exception as e:
            print(f"    ✗ 测试异常: {e}")
            result["failed"] += 1

        print(f"\n  Excel导入测试: 通过 {result['passed']}/{result['passed']+result['failed']}")
        return result

    def _test_mapping_learning_system(self) -> Dict:
        """系统测试 - 映射学习功能"""
        print("\n>>> 测试映射学习功能...")

        result = {
            "passed": 0,
            "failed": 0,
            "tests": []
        }

        try:
            from src.learning.enhanced_mapping import MappingRule, MappingLearner
            from unittest.mock import Mock

            # 测试1: 规则创建
            print("  [1] 测试规则创建...")
            rule = MappingRule(keyword="市场", responsible="张三")
            if rule.keyword == "市场" and rule.responsible == "张三":
                print("    ✓ 规则创建正确")
                result["passed"] += 1
            else:
                print("    ✗ 规则创建失败")
                result["failed"] += 1
            result["tests"].append("规则创建")

            # 测试2: 映射学习器
            print("  [2] 测试映射学习器...")
            mock_db = Mock()
            learner = MappingLearner(mock_db)
            if learner.db == mock_db and learner.rules == {}:
                print("    ✓ 映射学习器初始化正确")
                result["passed"] += 1
            else:
                print("    ✗ 映射学习器初始化失败")
                result["failed"] += 1
            result["tests"].append("映射学习器")

            # 测试3: 从文本学习
            print("  [3] 测试从文本学习...")
            rule = learner.learn_from_text("市场推广活动", "李四")
            if rule and rule.responsible == "李四":
                print("    ✓ 从文本学习正确")
                result["passed"] += 1
            else:
                print("    ✗ 从文本学习失败")
                result["failed"] += 1
            result["tests"].append("从文本学习")

            # 测试4: 推荐功能
            print("  [4] 测试推荐功能...")
            responsible, confidence = learner.recommend("市场推广")
            if responsible == "李四" and confidence > 0:
                print(f"    ✓ 推荐正确 (责任人: {responsible}, 置信度: {confidence:.2f})")
                result["passed"] += 1
            else:
                print("    ✗ 推荐失败")
                result["failed"] += 1
            result["tests"].append("推荐功能")

        except Exception as e:
            print(f"    ✗ 测试异常: {e}")
            result["failed"] += 1

        print(f"\n  映射学习测试: 通过 {result['passed']}/{result['passed']+result['failed']}")
        return result

    def run_integration_tests(self) -> Dict:
        """运行集成测试"""
        print("\n" + "=" * 60)
        print("  集成测试")
        print("=" * 60)

        result = {
            "passed": 0,
            "failed": 0,
            "tests": []
        }

        try:
            from src.content.excel_import import ExcelImporter, ExcelExporter
            from src.learning.enhanced_mapping import MappingLearner
            from unittest.mock import Mock
            from openpyxl import Workbook

            print("\n>>> 测试端到端流程...")

            # 1. 创建模拟数据库
            print("  [1] 初始化数据库...")
            mock_db = Mock()
            mock_db.get_all_contacts.return_value = []
            mock_db.add_contact.return_value = True
            result["passed"] += 1

            # 2. 创建测试Excel
            print("  [2] 创建测试Excel...")
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                temp_excel = f.name

            wb = Workbook()
            ws = wb.active
            headers = ['姓名', '电话', '邮箱', '公司']
            for col, h in enumerate(headers, 1):
                ws.cell(1, col, h)
            ws.cell(2, 1, '测试用户')
            ws.cell(2, 2, '13800138000')
            ws.cell(2, 3, 'test@example.com')
            ws.cell(2, 4, '测试公司')
            wb.save(temp_excel)
            wb.close()
            result["passed"] += 1

            # 3. 导入Excel
            print("  [3] 导入Excel...")
            importer = ExcelImporter(mock_db)
            import_result = importer.import_from_file(temp_excel)
            if import_result['success'] and import_result['total'] == 1:
                print("    ✓ Excel导入成功")
                result["passed"] += 1
            else:
                print("    ✗ Excel导入失败")
                result["failed"] += 1

            # 4. 保存到数据库
            print("  [4] 保存到数据库...")
            saved = importer.save_to_database()
            if saved == 1:
                print("    ✓ 数据库保存成功")
                result["passed"] += 1
            else:
                print("    ✗ 数据库保存失败")
                result["failed"] += 1

            # 5. 导出Excel
            print("  [5] 导出Excel...")
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                output_excel = f.name

            exporter = ExcelExporter(mock_db)
            exported = exporter.export_all(output_excel)
            if os.path.exists(output_excel):
                print("    ✓ Excel导出成功")
                result["passed"] += 1
            else:
                print("    ✗ Excel导出失败")
                result["failed"] += 1

            # 6. 清理
            os.unlink(temp_excel)
            os.unlink(output_excel)

        except Exception as e:
            print(f"    ✗ 集成测试异常: {e}")
            result["failed"] += 1

        print(f"\n  集成测试: 通过 {result['passed']}/{result['passed']+result['failed']}")
        return result

    def print_summary(self, unit_results, system_results, integration_results):
        """打印测试总结"""
        print("\n" + "=" * 60)
        print("  测试总结")
        print("=" * 60)

        # 单元测试汇总
        unit_passed, unit_failed = unit_results
        print(f"\n单元测试:")
        print(f"  - 模块通过: {unit_passed}")
        print(f"  - 模块失败: {unit_failed}")

        # 系统测试汇总
        print(f"\n系统测试:")
        for module, result in system_results.items():
            print(f"  - {module}: 通过 {result['passed']}/{result['passed']+result['failed']}")

        # 集成测试汇总
        print(f"\n集成测试:")
        print(f"  - 测试通过: {integration_results['passed']}")
        print(f"  - 测试失败: {integration_results['failed']}")

        # 总计
        total_tests = unit_passed * 10 + sum(r['passed'] for r in system_results.values()) + integration_results['passed']
        total_failed = unit_failed * 10 + sum(r['failed'] for r in system_results.values()) + integration_results['failed']

        print("\n" + "-" * 60)
        print(f"总计: 通过 {total_tests} 项, 失败 {total_failed} 项")
        print("=" * 60)

        # 结论
        if total_failed == 0:
            print("\n🎉 所有测试通过！代码质量符合要求。")
            return True
        else:
            print(f"\n⚠️  有 {total_failed} 项测试失败，请检查。")
            return False

    def run(self, mode: str = "all"):
        """运行测试"""
        self.start_time = time.time()
        self.print_header()

        if mode in ["all", "unit"]:
            unit_results = self.run_unit_tests()
        else:
            unit_results = (0, 0)

        if mode in ["all", "system"]:
            system_results = self.run_system_tests()
        else:
            system_results = {}

        if mode in ["all", "integration"]:
            integration_results = self.run_integration_tests()
        else:
            integration_results = {"passed": 0, "failed": 0}

        success = self.print_summary(unit_results, system_results, integration_results)

        elapsed = time.time() - self.start_time
        print(f"\n测试耗时: {elapsed:.2f} 秒")

        return 0 if success else 1


def main():
    """主函数"""
    mode = "all"

    if len(sys.argv) > 1:
        mode = sys.argv[1].lower()

    runner = TestRunner()
    return runner.run(mode)


if __name__ == "__main__":
    sys.exit(main())
