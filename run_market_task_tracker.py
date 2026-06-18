# 市场咨询任务跟踪工具 V4.6
# Market Task Tracker V4.6
# 可直接运行的完整应用程序

import sys
import os
import sqlite3
import csv
import json
import tempfile
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from pathlib import Path

# ==================== 依赖检查 ====================
def check_dependencies():
    """检查依赖并提供友好的错误提示"""
    missing = []
    
    try:
        from PyQt5.QtWidgets import QApplication
    except ImportError:
        missing.append("PyQt5")
    
    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")
    
    try:
        import pandas
    except ImportError:
        missing.append("pandas")
    
    if missing:
        print("=" * 60)
        print("⚠️  缺少必要的依赖库")
        print("=" * 60)
        print(f"请运行以下命令安装缺失的依赖:\n")
        print(f"pip install {' '.join(missing)}")
        print("=" * 60)
        sys.exit(1)

check_dependencies()

# ==================== PyQt5 导入 ====================
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QFormLayout, QGridLayout, QSplitter, QTabWidget, QStackedWidget,
    QLabel, QLineEdit, QTextEdit, QComboBox, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QDateTimeEdit, QCalendarWidget,
    QGroupBox, QFrame, QScrollArea, QDialog, QMessageBox, QFileDialog,
    QInputDialog, QListWidget, QListWidgetItem, QProgressBar, QSpinBox,
    QCheckBox, QRadioButton, QButtonGroup, QToolBar, QStatusBar,
    QMenuBar, QMenu, QAction, QShortcut, QKeySequence, QFontDialog,
    QColorDialog, QPrintDialog, QPrinter, QPageSetupDialog, QWizard,
    QFileSystemModel, QTreeView, QDirModel, QDockWidget, QTabBar,
    QSplashScreen, QStyleFactory, QDesktopWidget, QApplication
)
from PyQt5.QtCore import (
    Qt, QTimer, QDate, QTime, QDateTime, QSize, QPoint, QRect,
    QObject, pyqtSignal, pyqtSlot, QThread, QMutex, QSemaphore,
    QProcess, QSettings, QTranslator, QLocale, QLibraryInfo,
    QItemSelectionModel, QItemSelection, QSortFilterProxyModel,
    QAbstractTableModel, QAbstractListModel, QAbstractItemModel,
    QModelIndex, QPersistentModelIndex, QEvent, QSizePolicy,
    QStackedLayout, QGridLayout, QBoxLayout
)
from PyQt5.QtGui import (
    QIcon, QFont, QPalette, QColor, QBrush, QPen, QPainter,
    QPicture, QPixmap, QImage, QMovie, QClipboard, QCursor,
    QTextCursor, QTextCharFormat, QTextBlockFormat, QTextFormat,
    QKeyEvent, QMouseEvent, QWheelEvent, QFocusEvent, QCloseEvent,
    QShowEvent, QHideEvent, QResizeEvent, QMoveEvent, QDragEnterEvent,
    QDropEvent, QDragMoveEvent, QDragLeaveEvent, QContextMenuEvent,
    QInputMethodEvent, QInputMethodQueryEvent, QIcon
)

# ==================== 数据库模块 ====================
class DatabaseManager:
    """数据库管理器"""
    
    def __init__(self, db_path: str = None):
        if db_path is None:
            db_path = os.path.join(tempfile.gettempdir(), "market_task_tracker.db")
        self.db_path = db_path
        self.conn = None
        self._init_database()
    
    def _init_database(self):
        """初始化数据库"""
        self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._create_tables()
    
    def _create_tables(self):
        """创建数据表"""
        cursor = self.conn.cursor()
        
        # 任务表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS tasks (
                task_id TEXT PRIMARY KEY,
                task_name TEXT NOT NULL,
                consultant TEXT,
                responder TEXT,
                department TEXT,
                key_module TEXT,
                product_model TEXT,
                status TEXT DEFAULT '待处理',
                importance TEXT DEFAULT '中',
                description TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                completed_at TEXT,
                due_date TEXT,
                resolution TEXT
            )
        ''')
        
        # 跟踪记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS track_records (
                record_id TEXT PRIMARY KEY,
                task_id TEXT NOT NULL,
                content TEXT NOT NULL,
                operator TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (task_id) REFERENCES tasks (task_id)
            )
        ''')
        
        # 推荐库表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS recommendations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                key_module TEXT,
                employee_id TEXT,
                phone TEXT,
                email TEXT,
                department TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 通讯录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                employee_id TEXT,
                phone TEXT,
                email TEXT,
                department TEXT,
                title TEXT,
                remark TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 提醒表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS reminders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id TEXT NOT NULL,
                remind_time TEXT NOT NULL,
                message TEXT,
                is_reminded INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (task_id) REFERENCES tasks (task_id)
            )
        ''')
        
        self.conn.commit()
        
        # 插入示例数据
        self._insert_sample_data()
    
    def _insert_sample_data(self):
        """插入示例数据"""
        cursor = self.conn.cursor()
        
        # 检查是否已有数据
        cursor.execute("SELECT COUNT(*) FROM tasks")
        if cursor.fetchone()[0] > 0:
            return
        
        # 插入示例任务
        sample_tasks = [
            ("TASK001", "MAC认证故障排查", "李经理", "张工", "技术部", "MAC认证", "S5750E-48GT4XS-HP", "处理中", "高", "客户反映MAC认证失败"),
            ("TASK002", "802.1x认证配置咨询", "王总监", "李工", "技术部", "802.1x认证", "S5750E-24GT4XS-HP", "已完成", "中", "需要配置802.1x认证方案"),
            ("TASK003", "Portal认证问题", "赵经理", "王工", "运维部", "Portal认证", "Agile Controller", "待处理", "高", "Portal认证页面无法跳转"),
            ("TASK004", "交换机端口安全", "钱总", "张工", "技术部", "端口安全", "S5750E-48GT4XS-HP", "已完成", "中", "端口安全策略配置"),
            ("TASK005", "AAA服务器配置", "孙经理", "李工", "技术部", "AAA认证", "CAMS服务器", "处理中", "高", "AAA服务器无法连接"),
        ]
        
        for task in sample_tasks:
            cursor.execute('''
                INSERT OR IGNORE INTO tasks 
                (task_id, task_name, consultant, responder, department, 
                 key_module, product_model, status, importance, description)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', task)
        
        # 插入示例推荐库
        sample_recommendations = [
            ("张工", "MAC认证", "EMP001", "13800001001", "zhang@company.com", "技术部"),
            ("李工", "802.1x认证", "EMP002", "13800001002", "li@company.com", "技术部"),
            ("王工", "Portal认证", "EMP003", "13800001003", "wang@company.com", "运维部"),
            ("赵工", "端口安全", "EMP004", "13800001004", "zhao@company.com", "技术部"),
            ("刘工", "AAA认证", "EMP005", "13800001005", "liu@company.com", "技术部"),
        ]
        
        for rec in sample_recommendations:
            cursor.execute('''
                INSERT OR IGNORE INTO recommendations 
                (name, key_module, employee_id, phone, email, department)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', rec)
        
        # 插入示例通讯录
        sample_contacts = [
            ("李经理", "MGR001", "13900001001", "limanager@company.com", "市场部", "经理"),
            ("王总监", "DIR001", "13900001002", "wang@company.com", "市场部", "总监"),
            ("张工", "EMP001", "13800001001", "zhang@company.com", "技术部", "工程师"),
            ("李工", "EMP002", "13800001002", "li@company.com", "技术部", "工程师"),
            ("王工", "EMP003", "13800001003", "wang@company.com", "运维部", "工程师"),
        ]
        
        for contact in sample_contacts:
            cursor.execute('''
                INSERT OR IGNORE INTO contacts 
                (name, employee_id, phone, email, department, title)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', contact)
        
        self.conn.commit()
    
    def execute(self, sql: str, params: tuple = ()) -> sqlite3.Cursor:
        """执行SQL"""
        cursor = self.conn.cursor()
        cursor.execute(sql, params)
        self.conn.commit()
        return cursor
    
    def fetchall(self, sql: str, params: tuple = ()) -> List[sqlite3.Row]:
        """查询所有"""
        cursor = self.conn.cursor()
        cursor.execute(sql, params)
        return cursor.fetchall()
    
    def fetchone(self, sql: str, params: tuple = ()) -> Optional[sqlite3.Row]:
        """查询一条"""
        cursor = self.conn.cursor()
        cursor.execute(sql, params)
        return cursor.fetchone()
    
    def close(self):
        """关闭连接"""
        if self.conn:
            self.conn.close()

# 全局数据库实例
db = DatabaseManager()

# ==================== 任务服务 ====================
class TaskService:
    """任务服务"""
    
    @staticmethod
    def generate_task_id() -> str:
        """生成任务ID"""
        cursor = db.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM tasks")
        count = cursor.fetchone()[0]
        return f"TASK{str(count + 1).zfill(3)}"
    
    @staticmethod
    def get_all_tasks() -> List[Dict]:
        """获取所有任务"""
        rows = db.fetchall("SELECT * FROM tasks ORDER BY created_at DESC")
        return [dict(row) for row in rows]
    
    @staticmethod
    def create_task(task_data: Dict) -> str:
        """创建任务"""
        task_id = TaskService.generate_task_id()
        db.execute('''
            INSERT INTO tasks 
            (task_id, task_name, consultant, responder, department, 
             key_module, product_model, status, importance, description, due_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            task_id,
            task_data.get('task_name', ''),
            task_data.get('consultant', ''),
            task_data.get('responder', ''),
            task_data.get('department', ''),
            task_data.get('key_module', ''),
            task_data.get('product_model', ''),
            task_data.get('status', '待处理'),
            task_data.get('importance', '中'),
            task_data.get('description', ''),
            task_data.get('due_date', '')
        ))
        return task_id
    
    @staticmethod
    def update_task(task_id: str, task_data: Dict) -> bool:
        """更新任务"""
        db.execute('''
            UPDATE tasks SET
                task_name = ?,
                consultant = ?,
                responder = ?,
                department = ?,
                key_module = ?,
                product_model = ?,
                status = ?,
                importance = ?,
                description = ?,
                due_date = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE task_id = ?
        ''', (
            task_data.get('task_name', ''),
            task_data.get('consultant', ''),
            task_data.get('responder', ''),
            task_data.get('department', ''),
            task_data.get('key_module', ''),
            task_data.get('product_model', ''),
            task_data.get('status', '待处理'),
            task_data.get('importance', '中'),
            task_data.get('description', ''),
            task_data.get('due_date', ''),
            task_id
        ))
        return True
    
    @staticmethod
    def delete_task(task_id: str) -> bool:
        """删除任务"""
        db.execute("DELETE FROM tasks WHERE task_id = ?", (task_id,))
        return True
    
    @staticmethod
    def search_tasks(keyword: str) -> List[Dict]:
        """搜索任务"""
        rows = db.fetchall('''
            SELECT * FROM tasks 
            WHERE task_name LIKE ? OR consultant LIKE ? OR responder LIKE ? 
               OR key_module LIKE ? OR description LIKE ?
            ORDER BY created_at DESC
        ''', (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))
        return [dict(row) for row in rows]

# ==================== 跟踪记录服务 ====================
class TrackService:
    """跟踪记录服务"""
    
    @staticmethod
    def get_records_by_task(task_id: str) -> List[Dict]:
        """获取任务跟踪记录"""
        rows = db.fetchall(
            "SELECT * FROM track_records WHERE task_id = ? ORDER BY created_at DESC",
            (task_id,)
        )
        return [dict(row) for row in rows]
    
    @staticmethod
    def add_record(task_id: str, content: str, operator: str = "系统") -> str:
        """添加跟踪记录"""
        import uuid
        record_id = f"REC{uuid.uuid4().hex[:8].upper()}"
        db.execute('''
            INSERT INTO track_records (record_id, task_id, content, operator)
            VALUES (?, ?, ?, ?)
        ''', (record_id, task_id, content, operator))
        return record_id
    
    @staticmethod
    def delete_record(record_id: str) -> bool:
        """删除跟踪记录"""
        db.execute("DELETE FROM track_records WHERE record_id = ?", (record_id,))
        return True

# ==================== 推荐服务 ====================
class RecommendationService:
    """推荐服务"""
    
    @staticmethod
    def get_all_recommendations() -> List[Dict]:
        """获取所有推荐"""
        rows = db.fetchall("SELECT * FROM recommendations ORDER BY name")
        return [dict(row) for row in rows]
    
    @staticmethod
    def search_by_module(key_module: str) -> List[Dict]:
        """按关键模块搜索推荐"""
        rows = db.fetchall(
            "SELECT * FROM recommendations WHERE key_module LIKE ? ORDER BY name",
            (f'%{key_module}%',)
        )
        return [dict(row) for row in rows]
    
    @staticmethod
    def add_recommendation(data: Dict) -> int:
        """添加推荐"""
        cursor = db.execute('''
            INSERT INTO recommendations (name, key_module, employee_id, phone, email, department)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            data.get('name', ''),
            data.get('key_module', ''),
            data.get('employee_id', ''),
            data.get('phone', ''),
            data.get('email', ''),
            data.get('department', '')
        ))
        return cursor.lastrowid
    
    @staticmethod
    def delete_recommendation(id: int) -> bool:
        """删除推荐"""
        db.execute("DELETE FROM recommendations WHERE id = ?", (id,))
        return True

# ==================== 通讯录服务 ====================
class ContactsService:
    """通讯录服务"""
    
    @staticmethod
    def get_all_contacts() -> List[Dict]:
        """获取所有联系人"""
        rows = db.fetchall("SELECT * FROM contacts ORDER BY name")
        return [dict(row) for row in rows]
    
    @staticmethod
    def search_contacts(keyword: str) -> List[Dict]:
        """搜索联系人"""
        rows = db.fetchall('''
            SELECT * FROM contacts 
            WHERE name LIKE ? OR employee_id LIKE ? OR phone LIKE ? OR department LIKE ?
            ORDER BY name
        ''', (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))
        return [dict(row) for row in rows]
    
    @staticmethod
    def add_contact(data: Dict) -> int:
        """添加联系人"""
        cursor = db.execute('''
            INSERT INTO contacts (name, employee_id, phone, email, department, title, remark)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('name', ''),
            data.get('employee_id', ''),
            data.get('phone', ''),
            data.get('email', ''),
            data.get('department', ''),
            data.get('title', ''),
            data.get('remark', '')
        ))
        return cursor.lastrowid
    
    @staticmethod
    def update_contact(id: int, data: Dict) -> bool:
        """更新联系人"""
        db.execute('''
            UPDATE contacts SET
                name = ?,
                employee_id = ?,
                phone = ?,
                email = ?,
                department = ?,
                title = ?,
                remark = ?
            WHERE id = ?
        ''', (
            data.get('name', ''),
            data.get('employee_id', ''),
            data.get('phone', ''),
            data.get('email', ''),
            data.get('department', ''),
            data.get('title', ''),
            data.get('remark', ''),
            id
        ))
        return True
    
    @staticmethod
    def delete_contact(id: int) -> bool:
        """删除联系人"""
        db.execute("DELETE FROM contacts WHERE id = ?", (id,))
        return True

# ==================== 统计服务 ====================
class StatisticsService:
    """统计服务"""
    
    @staticmethod
    def get_summary() -> Dict:
        """获取统计摘要"""
        total = db.fetchone("SELECT COUNT(*) as count FROM tasks")
        today = datetime.now().strftime('%Y-%m-%d')
        today_created = db.fetchone(
            "SELECT COUNT(*) as count FROM tasks WHERE created_at LIKE ?",
            (f'{today}%',)
        )
        completed = db.fetchone(
            "SELECT COUNT(*) as count FROM tasks WHERE status = '已完成'"
        )
        in_progress = db.fetchone(
            "SELECT COUNT(*) as count FROM tasks WHERE status = '处理中'"
        )
        overdue = db.fetchone(
            "SELECT COUNT(*) as count FROM tasks WHERE status NOT IN ('已完成', '已取消')"
        )
        
        return {
            'total': total[0] if total else 0,
            'today_created': today_created[0] if today_created else 0,
            'completed': completed[0] if completed else 0,
            'in_progress': in_progress[0] if in_progress else 0,
            'pending': overdue[0] if overdue else 0
        }
    
    @staticmethod
    def get_status_distribution() -> List[Dict]:
        """获取状态分布"""
        rows = db.fetchall('''
            SELECT status, COUNT(*) as count 
            FROM tasks 
            GROUP BY status 
            ORDER BY count DESC
        ''')
        total = db.fetchone("SELECT COUNT(*) FROM tasks")[0]
        return [
            {
                'status': row[0],
                'count': row[1],
                'percentage': round(row[1] / total * 100, 1) if total > 0 else 0
            }
            for row in rows
        ]
    
    @staticmethod
    def get_department_distribution() -> List[Dict]:
        """获取部门分布"""
        rows = db.fetchall('''
            SELECT department, COUNT(*) as count 
            FROM tasks 
            WHERE department IS NOT NULL AND department != ''
            GROUP BY department 
            ORDER BY count DESC
        ''')
        return [{'department': row[0], 'count': row[1]} for row in rows]
    
    @staticmethod
    def get_module_distribution() -> List[Dict]:
        """获取模块分布"""
        rows = db.fetchall('''
            SELECT key_module, COUNT(*) as count 
            FROM tasks 
            WHERE key_module IS NOT NULL AND key_module != ''
            GROUP BY key_module 
            ORDER BY count DESC
            LIMIT 10
        ''')
        return [{'module': row[0], 'count': row[1]} for row in rows]

# ==================== 导出服务 ====================
class ExportService:
    """导出服务"""
    
    @staticmethod
    def generate_filename(prefix: str, ext: str) -> str:
        """生成带时间戳的文件名"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{prefix}_{timestamp}.{ext}"
    
    @staticmethod
    def export_tasks_to_excel(tasks: List[Dict], filepath: str) -> bool:
        """导出任务到Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "任务列表"
            
            # 表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # 写入表头
            headers = ["任务ID", "任务名称", "咨询者", "答复人", "部门", 
                      "关键模块", "产品型号", "状态", "重要程度", "创建时间"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # 写入数据
            for row_idx, task in enumerate(tasks, 2):
                ws.cell(row=row_idx, column=1, value=task.get('task_id', '')).border = thin_border
                ws.cell(row=row_idx, column=2, value=task.get('task_name', '')).border = thin_border
                ws.cell(row=row_idx, column=3, value=task.get('consultant', '')).border = thin_border
                ws.cell(row=row_idx, column=4, value=task.get('responder', '')).border = thin_border
                ws.cell(row=row_idx, column=5, value=task.get('department', '')).border = thin_border
                ws.cell(row=row_idx, column=6, value=task.get('key_module', '')).border = thin_border
                ws.cell(row=row_idx, column=7, value=task.get('product_model', '')).border = thin_border
                ws.cell(row=row_idx, column=8, value=task.get('status', '')).border = thin_border
                ws.cell(row=row_idx, column=9, value=task.get('importance', '')).border = thin_border
                ws.cell(row=row_idx, column=10, value=task.get('created_at', '')).border = thin_border
            
            # 设置列宽
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 20
            
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"导出Excel失败: {e}")
            return False
    
    @staticmethod
    def export_tasks_to_csv(tasks: List[Dict], filepath: str) -> bool:
        """导出任务到CSV"""
        try:
            with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(["任务ID", "任务名称", "咨询者", "答复人", "部门", 
                               "关键模块", "产品型号", "状态", "重要程度", "创建时间"])
                for task in tasks:
                    writer.writerow([
                        task.get('task_id', ''),
                        task.get('task_name', ''),
                        task.get('consultant', ''),
                        task.get('responder', ''),
                        task.get('department', ''),
                        task.get('key_module', ''),
                        task.get('product_model', ''),
                        task.get('status', ''),
                        task.get('importance', ''),
                        task.get('created_at', '')
                    ])
            return True
        except Exception as e:
            print(f"导出CSV失败: {e}")
            return False
    
    @staticmethod
    def export_statistics_to_excel(filepath: str) -> bool:
        """导出统计报表到Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            
            # 样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            title_font = Font(bold=True, size=14)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Sheet1: 统计概览
            ws1 = wb.active
            ws1.title = "统计概览"
            
            summary = StatisticsService.get_summary()
            
            ws1.cell(row=1, column=1, value="市场咨询任务跟踪工具 - 统计报告").font = title_font
            ws1.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            ws1.cell(row=4, column=1, value="【任务概览】").font = Font(bold=True, size=12)
            overview_data = [
                ("总任务数", summary['total']),
                ("今日新增", summary['today_created']),
                ("进行中", summary['in_progress']),
                ("已完成", summary['completed']),
                ("待处理", summary['pending'])
            ]
            for idx, (label, value) in enumerate(overview_data, 5):
                ws1.cell(row=idx, column=1, value=label).border = thin_border
                ws1.cell(row=idx, column=2, value=value).border = thin_border
            
            ws1.column_dimensions['A'].width = 15
            ws1.column_dimensions['B'].width = 15
            
            # Sheet2: 状态分布
            ws2 = wb.create_sheet("状态分布")
            ws2.cell(row=1, column=1, value="任务状态分布").font = title_font
            
            headers = ["状态", "数量", "占比"]
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            
            status_data = StatisticsService.get_status_distribution()
            for row_idx, item in enumerate(status_data, 4):
                ws2.cell(row=row_idx, column=1, value=item['status']).border = thin_border
                ws2.cell(row=row_idx, column=2, value=item['count']).border = thin_border
                ws2.cell(row=row_idx, column=3, value=f"{item['percentage']}%").border = thin_border
            
            # Sheet3: 部门分布
            ws3 = wb.create_sheet("部门分布")
            ws3.cell(row=1, column=1, value="部门任务分布").font = title_font
            
            headers = ["部门", "任务数"]
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            
            dept_data = StatisticsService.get_department_distribution()
            for row_idx, item in enumerate(dept_data, 4):
                ws3.cell(row=row_idx, column=1, value=item['department']).border = thin_border
                ws3.cell(row=row_idx, column=2, value=item['count']).border = thin_border
            
            # Sheet4: 模块分布
            ws4 = wb.create_sheet("模块分布")
            ws4.cell(row=1, column=1, value="关键模块分布").font = title_font
            
            headers = ["模块", "任务数"]
            for col, header in enumerate(headers, 1):
                cell = ws4.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            
            module_data = StatisticsService.get_module_distribution()
            for row_idx, item in enumerate(module_data, 4):
                ws4.cell(row=row_idx, column=1, value=item['module']).border = thin_border
                ws4.cell(row=row_idx, column=2, value=item['count']).border = thin_border
            
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"导出统计报表失败: {e}")
            return False

# ==================== 主窗口 ====================
class MainWindow(QMainWindow):
    """主窗口"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("市场咨询任务跟踪工具 V4.6")
        self.setGeometry(100, 100, 1400, 800)
        
        # 初始化UI
        self._init_ui()
        self._init_menu()
        self._init_toolbar()
        self._init_statusbar()
        
        # 加载数据
        self._load_tasks()
        
        # 定时刷新
        self.timer = QTimer()
        self.timer.timeout.connect(self._refresh_data)
        self.timer.start(30000)  # 每30秒刷新
    
    def _init_ui(self):
        """初始化UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QHBoxLayout(central_widget)
        
        # 左侧导航
        self.nav_list = QListWidget()
        self.nav_list.setMaximumWidth(180)
        self.nav_list.currentRowChanged.connect(self._on_nav_changed)
        
        nav_items = [
            "📋 任务信息",
            "📊 任务跟踪",
            "🔮 智能推荐",
            "🔔 提醒管理",
            "📁 数据导入",
            "📈 统计分析",
            "📚 推荐库",
            "📒 通讯录"
        ]
        for item_text in nav_items:
            self.nav_list.addItem(item_text)
        
        # 右侧工作区
        self.work_area = QStackedWidget()
        
        # 创建各功能页面
        self.task_page = self._create_task_page()
        self.track_page = self._create_track_page()
        self.recommend_page = self._create_recommend_page()
        self.reminder_page = self._create_reminder_page()
        self.import_page = self._create_import_page()
        self.stats_page = self._create_stats_page()
        self.library_page = self._create_library_page()
        self.contacts_page = self._create_contacts_page()
        
        self.work_area.addWidget(self.task_page)
        self.work_area.addWidget(self.track_page)
        self.work_area.addWidget(self.recommend_page)
        self.work_area.addWidget(self.reminder_page)
        self.work_area.addWidget(self.import_page)
        self.work_area.addWidget(self.stats_page)
        self.work_area.addWidget(self.library_page)
        self.work_area.addWidget(self.contacts_page)
        
        # 布局
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(self.nav_list)
        splitter.addWidget(self.work_area)
        splitter.setStretchFactor(1, 1)
        
        main_layout.addWidget(splitter)
        
        # 默认选择任务页面
        self.nav_list.setCurrentRow(0)
    
    def _create_task_page(self) -> QWidget:
        """创建任务页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 标题栏
        title_layout = QHBoxLayout()
        title_label = QLabel("📋 任务信息管理")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        self.new_task_btn = QPushButton("➕ 新建任务")
        self.new_task_btn.clicked.connect(self._on_new_task)
        title_layout.addWidget(self.new_task_btn)
        
        self.refresh_task_btn = QPushButton("🔄 刷新")
        self.refresh_task_btn.clicked.connect(self._load_tasks)
        title_layout.addWidget(self.refresh_task_btn)
        
        layout.addLayout(title_layout)
        
        # 搜索栏
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 搜索:"))
        self.task_search = QLineEdit()
        self.task_search.setPlaceholderText("输入关键词搜索任务...")
        self.task_search.textChanged.connect(self._on_task_search)
        search_layout.addWidget(self.task_search)
        layout.addLayout(search_layout)
        
        # 任务表格
        self.task_table = QTableWidget()
        self.task_table.setColumnCount(10)
        self.task_table.setHorizontalHeaderLabels([
            "任务ID", "任务名称", "咨询者", "答复人", "部门",
            "关键模块", "状态", "重要程度", "创建时间", "操作"
        ])
        self.task_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.task_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.task_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.task_table.itemDoubleClicked.connect(self._on_task_double_click)
        layout.addWidget(self.task_table)
        
        return widget
    
    def _create_track_page(self) -> QWidget:
        """创建跟踪页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 标题
        title_label = QLabel("📊 任务跟踪")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(title_label)
        
        # 选择任务
        select_layout = QHBoxLayout()
        select_layout.addWidget(QLabel("选择任务:"))
        self.track_task_combo = QComboBox()
        self.track_task_combo.currentIndexChanged.connect(self._on_track_task_changed)
        select_layout.addWidget(self.track_task_combo)
        layout.addLayout(select_layout)
        
        # 跟踪记录
        record_group = QGroupBox("跟踪记录")
        record_layout = QVBoxLayout()
        
        self.track_table = QTableWidget()
        self.track_table.setColumnCount(4)
        self.track_table.setHorizontalHeaderLabels(["记录ID", "操作人", "内容", "时间"])
        self.track_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.track_table.setEditTriggers(QTableWidget.NoEditTriggers)
        record_layout.addWidget(self.track_table)
        
        record_group.setLayout(record_layout)
        layout.addWidget(record_group)
        
        # 添加记录
        add_group = QGroupBox("添加跟踪记录")
        add_layout = QVBoxLayout()
        
        self.track_content = QTextEdit()
        self.track_content.setMaximumHeight(80)
        add_layout.addWidget(QLabel("记录内容:"))
        add_layout.addWidget(self.track_content)
        
        btn_layout = QHBoxLayout()
        self.add_record_btn = QPushButton("➕ 添加记录")
        self.add_record_btn.clicked.connect(self._on_add_record)
        btn_layout.addWidget(self.add_record_btn)
        btn_layout.addStretch()
        
        add_layout.addLayout(btn_layout)
        add_group.setLayout(add_layout)
        layout.addWidget(add_group)
        
        return widget
    
    def _create_recommend_page(self) -> QWidget:
        """创建推荐页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        title_label = QLabel("🔮 智能推荐")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(title_label)
        
        # 搜索
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("关键模块:"))
        self.recommend_search = QLineEdit()
        self.recommend_search.setPlaceholderText("输入关键模块搜索专家...")
        search_layout.addWidget(self.recommend_search)
        self.recommend_search_btn = QPushButton("🔍 搜索")
        self.recommend_search_btn.clicked.connect(self._on_recommend_search)
        search_layout.addWidget(self.recommend_search_btn)
        layout.addLayout(search_layout)
        
        # 推荐列表
        self.recommend_table = QTableWidget()
        self.recommend_table.setColumnCount(6)
        self.recommend_table.setHorizontalHeaderLabels([
            "姓名", "关键模块", "工号", "电话", "邮箱", "部门"
        ])
        self.recommend_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.recommend_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.recommend_table)
        
        # 加载推荐
        self._load_recommendations()
        
        return widget
    
    def _create_reminder_page(self) -> QWidget:
        """创建提醒页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        title_label = QLabel("🔔 提醒管理")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(title_label)
        
        info_label = QLabel("提醒功能开发中... 设置任务到期提醒")
        info_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(info_label)
        
        return widget
    
    def _create_import_page(self) -> QWidget:
        """创建导入页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        title_label = QLabel("📁 数据导入")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(title_label)
        
        # 导入选项
        options_group = QGroupBox("导入选项")
        options_layout = QVBoxLayout()
        
        self.import_excel_btn = QPushButton("📊 导入Excel文件")
        self.import_excel_btn.clicked.connect(self._on_import_excel)
        options_layout.addWidget(self.import_excel_btn)
        
        self.import_csv_btn = QPushButton("📄 导入CSV文件")
        self.import_csv_btn.clicked.connect(self._on_import_csv)
        options_layout.addWidget(self.import_csv_btn)
        
        self.import_txt_btn = QPushButton("📝 从文本导入")
        self.import_txt_btn.clicked.connect(self._on_import_txt)
        options_layout.addWidget(self.import_txt_btn)
        
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        
        layout.addStretch()
        
        return widget
    
    def _create_stats_page(self) -> QWidget:
        """创建统计页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 标题栏
        title_layout = QHBoxLayout()
        title_label = QLabel("📈 数据统计分析")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        self.export_stats_btn = QPushButton("📊 导出报表")
        self.export_stats_btn.clicked.connect(self._on_export_statistics)
        title_layout.addWidget(self.export_stats_btn)
        
        self.refresh_stats_btn = QPushButton("🔄 刷新")
        self.refresh_stats_btn.clicked.connect(self._load_statistics)
        title_layout.addWidget(self.refresh_stats_btn)
        
        layout.addLayout(title_layout)
        
        # 统计卡片
        cards_layout = QGridLayout()
        
        self.total_card = self._create_stat_card("总任务数", "0", "#4472C4")
        self.today_card = self._create_stat_card("今日新增", "0", "#2E7D32")
        self.progress_card = self._create_stat_card("进行中", "0", "#ED6C02")
        self.complete_card = self._create_stat_card("已完成", "0", "#00897B")
        
        cards_layout.addWidget(self.total_card, 0, 0)
        cards_layout.addWidget(self.today_card, 0, 1)
        cards_layout.addWidget(self.progress_card, 0, 2)
        cards_layout.addWidget(self.complete_card, 0, 3)
        
        layout.addLayout(cards_layout)
        
        # 状态分布
        status_group = QGroupBox("状态分布")
        status_layout = QVBoxLayout()
        self.status_table = QTableWidget()
        self.status_table.setColumnCount(3)
        self.status_table.setHorizontalHeaderLabels(["状态", "数量", "占比"])
        self.status_table.setEditTriggers(QTableWidget.NoEditTriggers)
        status_layout.addWidget(self.status_table)
        status_group.setLayout(status_layout)
        layout.addWidget(status_group)
        
        # 部门分布
        dept_group = QGroupBox("部门分布")
        dept_layout = QVBoxLayout()
        self.dept_table = QTableWidget()
        self.dept_table.setColumnCount(2)
        self.dept_table.setHorizontalHeaderLabels(["部门", "任务数"])
        self.dept_table.setEditTriggers(QTableWidget.NoEditTriggers)
        dept_layout.addWidget(self.dept_table)
        dept_group.setLayout(dept_layout)
        layout.addWidget(dept_group)
        
        # 加载统计
        self._load_statistics()
        
        return widget
    
    def _create_stat_card(self, title: str, value: str, color: str) -> QWidget:
        """创建统计卡片"""
        card = QFrame()
        card.setFrameStyle(QFrame.StyledPanel | QFrame.Raised)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 8px;
                padding: 15px;
            }}
        """)
        
        layout = QVBoxLayout(card)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("color: white; font-size: 12px;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        value_label = QLabel(value)
        value_label.setObjectName("stat_value")
        value_label.setStyleSheet("color: white; font-size: 28px; font-weight: bold;")
        value_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(value_label)
        
        return card
    
    def _create_library_page(self) -> QWidget:
        """创建推荐库页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        title_layout = QHBoxLayout()
        title_label = QLabel("📚 推荐库管理")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        self.add_recommend_btn = QPushButton("➕ 添加专家")
        self.add_recommend_btn.clicked.connect(self._on_add_recommendation)
        title_layout.addWidget(self.add_recommend_btn)
        
        self.export_recommend_btn = QPushButton("📤 导出")
        self.export_recommend_btn.clicked.connect(self._on_export_recommendations)
        title_layout.addWidget(self.export_recommend_btn)
        
        layout.addLayout(title_layout)
        
        # 推荐库表格
        self.library_table = QTableWidget()
        self.library_table.setColumnCount(7)
        self.library_table.setHorizontalHeaderLabels([
            "ID", "姓名", "关键模块", "工号", "电话", "邮箱", "部门"
        ])
        self.library_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.library_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.library_table)
        
        # 加载推荐库
        self._load_library()
        
        return widget
    
    def _create_contacts_page(self) -> QWidget:
        """创建通讯录页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        title_layout = QHBoxLayout()
        title_label = QLabel("📒 通讯录")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        self.add_contact_btn = QPushButton("➕ 添加联系人")
        self.add_contact_btn.clicked.connect(self._on_add_contact)
        title_layout.addWidget(self.add_contact_btn)
        
        self.export_contacts_btn = QPushButton("📤 导出")
        self.export_contacts_btn.clicked.connect(self._on_export_contacts)
        title_layout.addWidget(self.export_contacts_btn)
        
        layout.addLayout(title_layout)
        
        # 搜索
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 搜索:"))
        self.contacts_search = QLineEdit()
        self.contacts_search.setPlaceholderText("输入姓名、工号或部门搜索...")
        self.contacts_search.textChanged.connect(self._on_contacts_search)
        search_layout.addWidget(self.contacts_search)
        layout.addLayout(search_layout)
        
        # 通讯录表格
        self.contacts_table = QTableWidget()
        self.contacts_table.setColumnCount(7)
        self.contacts_table.setHorizontalHeaderLabels([
            "ID", "姓名", "工号", "电话", "邮箱", "部门", "职位"
        ])
        self.contacts_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.contacts_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.contacts_table)
        
        # 加载通讯录
        self._load_contacts()
        
        return widget
    
    def _init_menu(self):
        """初始化菜单"""
        menubar = self.menuBar()
        
        # 文件菜单
        file_menu = menubar.addMenu("文件")
        
        new_action = QAction("新建任务", self)
        new_action.setShortcut("Ctrl+N")
        new_action.triggered.connect(self._on_new_task)
        file_menu.addAction(new_action)
        
        file_menu.addSeparator()
        
        import_action = QAction("导入...", self)
        import_action.setShortcut("Ctrl+I")
        import_action.triggered.connect(self._on_import_excel)
        file_menu.addAction(import_action)
        
        export_action = QAction("导出...", self)
        export_action.setShortcut("Ctrl+E")
        export_action.triggered.connect(self._on_export_tasks)
        file_menu.addAction(export_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("退出", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # 工具菜单
        tools_menu = menubar.addMenu("工具")
        
        stats_action = QAction("统计报表", self)
        stats_action.triggered.connect(self._on_export_statistics)
        tools_menu.addAction(stats_action)
        
        # 帮助菜单
        help_menu = menubar.addMenu("帮助")
        
        about_action = QAction("关于", self)
        about_action.triggered.connect(self._on_about)
        help_menu.addAction(about_action)
    
    def _init_toolbar(self):
        """初始化工具栏"""
        toolbar = QToolBar()
        self.addToolBar(toolbar)
        
        # 新建任务
        new_task_action = QAction("📋 新建任务", self)
        new_task_action.triggered.connect(self._on_new_task)
        toolbar.addAction(new_task_action)
        
        toolbar.addSeparator()
        
        # 导入
        import_action = QAction("📥 导入", self)
        import_action.triggered.connect(self._on_import_excel)
        toolbar.addAction(import_action)
        
        # 导出
        export_action = QAction("📤 导出", self)
        export_action.triggered.connect(self._on_export_tasks)
        toolbar.addAction(export_action)
        
        # 统计报表
        stats_action = QAction("📊 统计报表", self)
        stats_action.triggered.connect(self._on_export_statistics)
        toolbar.addAction(stats_action)
        
        toolbar.addSeparator()
        
        # 刷新
        refresh_action = QAction("🔄 刷新", self)
        refresh_action.triggered.connect(self._refresh_data)
        toolbar.addAction(refresh_action)
    
    def _init_statusbar(self):
        """初始化状态栏"""
        self.statusbar = QStatusBar()
        self.setStatusBar(self.statusbar)
        self.statusbar.showMessage("市场咨询任务跟踪工具 V4.6 就绪")
    
    # ==================== 数据加载方法 ====================
    
    def _load_tasks(self):
        """加载任务列表"""
        tasks = TaskService.get_all_tasks()
        self.task_table.setRowCount(len(tasks))
        
        for row, task in enumerate(tasks):
            self.task_table.setItem(row, 0, QTableWidgetItem(task.get('task_id', '')))
            self.task_table.setItem(row, 1, QTableWidgetItem(task.get('task_name', '')))
            self.task_table.setItem(row, 2, QTableWidgetItem(task.get('consultant', '')))
            self.task_table.setItem(row, 3, QTableWidgetItem(task.get('responder', '')))
            self.task_table.setItem(row, 4, QTableWidgetItem(task.get('department', '')))
            self.task_table.setItem(row, 5, QTableWidgetItem(task.get('key_module', '')))
            self.task_table.setItem(row, 6, QTableWidgetItem(task.get('status', '')))
            
            importance = task.get('importance', '中')
            importance_item = QTableWidgetItem(importance)
            if importance == '高':
                importance_item.setBackground(QBrush(QColor(255, 200, 200)))
            elif importance == '低':
                importance_item.setBackground(QBrush(QColor(200, 255, 200)))
            self.task_table.setItem(row, 7, importance_item)
            
            created_at = task.get('created_at', '')
            if created_at:
                created_at = created_at[:19] if len(created_at) > 19 else created_at
            self.task_table.setItem(row, 8, QTableWidgetItem(created_at))
            
            # 操作按钮
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("编辑")
            edit_btn.setMaximumWidth(50)
            edit_btn.clicked.connect(lambda _, t=task: self._on_edit_task(t))
            btn_layout.addWidget(edit_btn)
            
            delete_btn = QPushButton("删除")
            delete_btn.setMaximumWidth(50)
            delete_btn.clicked.connect(lambda _, t=task: self._on_delete_task(t))
            btn_layout.addWidget(delete_btn)
            
            self.task_table.setCellWidget(row, 9, btn_widget)
        
        # 更新跟踪页面的任务下拉框
        self.track_task_combo.clear()
        for task in tasks:
            self.track_task_combo.addItem(
                f"{task.get('task_id')} - {task.get('task_name')}",
                task.get('task_id')
            )
        
        self.statusbar.showMessage(f"已加载 {len(tasks)} 个任务")
    
    def _load_recommendations(self):
        """加载推荐列表"""
        recommendations = RecommendationService.get_all_recommendations()
        self.recommend_table.setRowCount(len(recommendations))
        
        for row, rec in enumerate(recommendations):
            self.recommend_table.setItem(row, 0, QTableWidgetItem(rec.get('name', '')))
            self.recommend_table.setItem(row, 1, QTableWidgetItem(rec.get('key_module', '')))
            self.recommend_table.setItem(row, 2, QTableWidgetItem(rec.get('employee_id', '')))
            self.recommend_table.setItem(row, 3, QTableWidgetItem(rec.get('phone', '')))
            self.recommend_table.setItem(row, 4, QTableWidgetItem(rec.get('email', '')))
            self.recommend_table.setItem(row, 5, QTableWidgetItem(rec.get('department', '')))
    
    def _load_library(self):
        """加载推荐库"""
        recommendations = RecommendationService.get_all_recommendations()
        self.library_table.setRowCount(len(recommendations))
        
        for row, rec in enumerate(recommendations):
            self.library_table.setItem(row, 0, QTableWidgetItem(str(rec.get('id', ''))))
            self.library_table.setItem(row, 1, QTableWidgetItem(rec.get('name', '')))
            self.library_table.setItem(row, 2, QTableWidgetItem(rec.get('key_module', '')))
            self.library_table.setItem(row, 3, QTableWidgetItem(rec.get('employee_id', '')))
            self.library_table.setItem(row, 4, QTableWidgetItem(rec.get('phone', '')))
            self.library_table.setItem(row, 5, QTableWidgetItem(rec.get('email', '')))
            self.library_table.setItem(row, 6, QTableWidgetItem(rec.get('department', '')))
    
    def _load_contacts(self):
        """加载通讯录"""
        contacts = ContactsService.get_all_contacts()
        self.contacts_table.setRowCount(len(contacts))
        
        for row, contact in enumerate(contacts):
            self.contacts_table.setItem(row, 0, QTableWidgetItem(str(contact.get('id', ''))))
            self.contacts_table.setItem(row, 1, QTableWidgetItem(contact.get('name', '')))
            self.contacts_table.setItem(row, 2, QTableWidgetItem(contact.get('employee_id', '')))
            self.contacts_table.setItem(row, 3, QTableWidgetItem(contact.get('phone', '')))
            self.contacts_table.setItem(row, 4, QTableWidgetItem(contact.get('email', '')))
            self.contacts_table.setItem(row, 5, QTableWidgetItem(contact.get('department', '')))
            self.contacts_table.setItem(row, 6, QTableWidgetItem(contact.get('title', '')))
    
    def _load_statistics(self):
        """加载统计数据"""
        summary = StatisticsService.get_summary()
        
        # 更新卡片
        cards = self.total_card.findChildren(QLabel, "stat_value")
        if cards:
            cards[0].setText(str(summary['total']))
        
        cards = self.today_card.findChildren(QLabel, "stat_value")
        if cards:
            cards[0].setText(str(summary['today_created']))
        
        cards = self.progress_card.findChildren(QLabel, "stat_value")
        if cards:
            cards[0].setText(str(summary['in_progress']))
        
        cards = self.complete_card.findChildren(QLabel, "stat_value")
        if cards:
            cards[0].setText(str(summary['completed']))
        
        # 状态分布
        status_data = StatisticsService.get_status_distribution()
        self.status_table.setRowCount(len(status_data))
        for row, item in enumerate(status_data):
            self.status_table.setItem(row, 0, QTableWidgetItem(item['status']))
            self.status_table.setItem(row, 1, QTableWidgetItem(str(item['count'])))
            self.status_table.setItem(row, 2, QTableWidgetItem(f"{item['percentage']}%"))
        
        # 部门分布
        dept_data = StatisticsService.get_department_distribution()
        self.dept_table.setRowCount(len(dept_data))
        for row, item in enumerate(dept_data):
            self.dept_table.setItem(row, 0, QTableWidgetItem(item['department']))
            self.dept_table.setItem(row, 1, QTableWidgetItem(str(item['count'])))
    
    def _refresh_data(self):
        """刷新所有数据"""
        self._load_tasks()
        self._load_statistics()
        self.statusbar.showMessage(f"数据已刷新 - {datetime.now().strftime('%H:%M:%S')}")
    
    # ==================== 事件处理方法 ====================
    
    def _on_nav_changed(self, index: int):
        """导航切换"""
        self.work_area.setCurrentIndex(index)
        
        if index == 2:  # 智能推荐
            self._load_recommendations()
        elif index == 5:  # 统计分析
            self._load_statistics()
        elif index == 6:  # 推荐库
            self._load_library()
        elif index == 7:  # 通讯录
            self._load_contacts()
    
    def _on_new_task(self):
        """新建任务"""
        dialog = TaskDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            task_data = dialog.get_task_data()
            task_id = TaskService.create_task(task_data)
            self._load_tasks()
            self.statusbar.showMessage(f"任务创建成功: {task_id}")
    
    def _on_edit_task(self, task: Dict):
        """编辑任务"""
        dialog = TaskDialog(self, task)
        if dialog.exec_() == QDialog.Accepted:
            task_data = dialog.get_task_data()
            TaskService.update_task(task['task_id'], task_data)
            self._load_tasks()
            self.statusbar.showMessage(f"任务更新成功: {task['task_id']}")
    
    def _on_delete_task(self, task: Dict):
        """删除任务"""
        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除任务 {task.get('task_id')} 吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            TaskService.delete_task(task['task_id'])
            self._load_tasks()
            self.statusbar.showMessage(f"任务已删除: {task['task_id']}")
    
    def _on_task_double_click(self, item: QTableWidgetItem):
        """双击任务"""
        row = item.row()
        task_id = self.task_table.item(row, 0).text()
        tasks = TaskService.get_all_tasks()
        for task in tasks:
            if task.get('task_id') == task_id:
                self._on_edit_task(task)
                break
    
    def _on_task_search(self, keyword: str):
        """搜索任务"""
        if keyword:
            tasks = TaskService.search_tasks(keyword)
        else:
            tasks = TaskService.get_all_tasks()
        
        self.task_table.setRowCount(len(tasks))
        for row, task in enumerate(tasks):
            self.task_table.setItem(row, 0, QTableWidgetItem(task.get('task_id', '')))
            self.task_table.setItem(row, 1, QTableWidgetItem(task.get('task_name', '')))
            self.task_table.setItem(row, 2, QTableWidgetItem(task.get('consultant', '')))
            self.task_table.setItem(row, 3, QTableWidgetItem(task.get('responder', '')))
            self.task_table.setItem(row, 4, QTableWidgetItem(task.get('department', '')))
            self.task_table.setItem(row, 5, QTableWidgetItem(task.get('key_module', '')))
            self.task_table.setItem(row, 6, QTableWidgetItem(task.get('status', '')))
            self.task_table.setItem(row, 7, QTableWidgetItem(task.get('importance', '')))
            self.task_table.setItem(row, 8, QTableWidgetItem(task.get('created_at', '')[:19] if task.get('created_at') else ''))
    
    def _on_track_task_changed(self, index: int):
        """跟踪任务切换"""
        if index < 0:
            return
        
        task_id = self.track_task_combo.currentData()
        records = TrackService.get_records_by_task(task_id)
        
        self.track_table.setRowCount(len(records))
        for row, record in enumerate(records):
            self.track_table.setItem(row, 0, QTableWidgetItem(record.get('record_id', '')))
            self.track_table.setItem(row, 1, QTableWidgetItem(record.get('operator', '')))
            self.track_table.setItem(row, 2, QTableWidgetItem(record.get('content', '')))
            self.track_table.setItem(row, 3, QTableWidgetItem(record.get('created_at', '')[:19] if record.get('created_at') else ''))
    
    def _on_add_record(self):
        """添加跟踪记录"""
        task_id = self.track_task_combo.currentData()
        if not task_id:
            QMessageBox.warning(self, "警告", "请先选择一个任务")
            return
        
        content = self.track_content.toPlainText().strip()
        if not content:
            QMessageBox.warning(self, "警告", "请输入记录内容")
            return
        
        record_id = TrackService.add_record(task_id, content)
        self.track_content.clear()
        self._on_track_task_changed(self.track_task_combo.currentIndex())
        self.statusbar.showMessage(f"跟踪记录已添加: {record_id}")
    
    def _on_recommend_search(self):
        """搜索推荐"""
        keyword = self.recommend_search.text().strip()
        if keyword:
            recommendations = RecommendationService.search_by_module(keyword)
        else:
            recommendations = RecommendationService.get_all_recommendations()
        
        self.recommend_table.setRowCount(len(recommendations))
        for row, rec in enumerate(recommendations):
            self.recommend_table.setItem(row, 0, QTableWidgetItem(rec.get('name', '')))
            self.recommend_table.setItem(row, 1, QTableWidgetItem(rec.get('key_module', '')))
            self.recommend_table.setItem(row, 2, QTableWidgetItem(rec.get('employee_id', '')))
            self.recommend_table.setItem(row, 3, QTableWidgetItem(rec.get('phone', '')))
            self.recommend_table.setItem(row, 4, QTableWidgetItem(rec.get('email', '')))
            self.recommend_table.setItem(row, 5, QTableWidgetItem(rec.get('department', '')))
    
    def _on_import_excel(self):
        """导入Excel"""
        filepath, _ = QFileDialog.getOpenFileName(
            self, "导入Excel", "", "Excel文件 (*.xlsx *.xls)"
        )
        if filepath:
            QMessageBox.information(self, "提示", f"Excel导入功能开发中...\n文件: {filepath}")
    
    def _on_import_csv(self):
        """导入CSV"""
        filepath, _ = QFileDialog.getOpenFileName(
            self, "导入CSV", "", "CSV文件 (*.csv)"
        )
        if filepath:
            QMessageBox.information(self, "提示", f"CSV导入功能开发中...\n文件: {filepath}")
    
    def _on_import_txt(self):
        """导入文本"""
        text, ok = QInputDialog.getMultiLineText(
            self, "导入文本", "请输入任务信息:"
        )
        if ok and text.strip():
            QMessageBox.information(self, "提示", "文本导入功能开发中...")
    
    def _on_export_tasks(self):
        """导出任务"""
        format_items = ["Excel (.xlsx)", "CSV (.csv)"]
        format_choice, ok = QInputDialog.getItem(
            self, "导出任务", "选择导出格式:", format_items, 0, False
        )
        if not ok:
            return
        
        if format_choice == "Excel (.xlsx)":
            filepath, _ = QFileDialog.getSaveFileName(
                self, "保存任务", ExportService.generate_filename("任务列表", "xlsx"),
                "Excel文件 (*.xlsx)"
            )
            if filepath:
                tasks = TaskService.get_all_tasks()
                if ExportService.export_tasks_to_excel(tasks, filepath):
                    QMessageBox.information(self, "成功", f"任务列表已导出到:\n{filepath}")
                else:
                    QMessageBox.critical(self, "失败", "导出失败")
        else:
            filepath, _ = QFileDialog.getSaveFileName(
                self, "保存任务", ExportService.generate_filename("任务列表", "csv"),
                "CSV文件 (*.csv)"
            )
            if filepath:
                tasks = TaskService.get_all_tasks()
                if ExportService.export_tasks_to_csv(tasks, filepath):
                    QMessageBox.information(self, "成功", f"任务列表已导出到:\n{filepath}")
                else:
                    QMessageBox.critical(self, "失败", "导出失败")
    
    def _on_export_statistics(self):
        """导出统计报表"""
        format_items = ["Excel (.xlsx)", "JSON (.json)", "CSV (.csv)"]
        format_choice, ok = QInputDialog.getItem(
            self, "导出统计报表", "选择导出格式:", format_items, 0, False
        )
        if not ok:
            return
        
        if format_choice == "Excel (.xlsx)":
            filepath, _ = QFileDialog.getSaveFileName(
                self, "保存统计报表", ExportService.generate_filename("统计报告", "xlsx"),
                "Excel文件 (*.xlsx)"
            )
            if filepath:
                if ExportService.export_statistics_to_excel(filepath):
                    QMessageBox.information(self, "成功", f"统计报表已导出到:\n{filepath}")
                else:
                    QMessageBox.critical(self, "失败", "导出失败")
        elif format_choice == "JSON (.json)":
            filepath, _ = QFileDialog.getSaveFileName(
                self, "保存统计报表", ExportService.generate_filename("统计报告", "json"),
                "JSON文件 (*.json)"
            )
            if filepath:
                summary = StatisticsService.get_summary()
                status_data = StatisticsService.get_status_distribution()
                data = {'summary': summary, 'status_distribution': status_data}
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                QMessageBox.information(self, "成功", f"统计报表已导出到:\n{filepath}")
        else:
            filepath, _ = QFileDialog.getSaveFileName(
                self, "保存统计报表", ExportService.generate_filename("统计报告", "csv"),
                "CSV文件 (*.csv)"
            )
            if filepath:
                summary = StatisticsService.get_summary()
                with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(["指标", "数值"])
                    writer.writerow(["总任务数", summary['total']])
                    writer.writerow(["今日新增", summary['today_created']])
                    writer.writerow(["进行中", summary['in_progress']])
                    writer.writerow(["已完成", summary['completed']])
                    writer.writerow(["待处理", summary['pending']])
                QMessageBox.information(self, "成功", f"统计报表已导出到:\n{filepath}")
    
    def _on_add_recommendation(self):
        """添加推荐"""
        dialog = RecommendationDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            RecommendationService.add_recommendation(data)
            self._load_library()
            self.statusbar.showMessage("专家已添加")
    
    def _on_export_recommendations(self):
        """导出推荐库"""
        filepath, _ = QFileDialog.getSaveFileName(
            self, "导出推荐库", ExportService.generate_filename("推荐库", "xlsx"),
            "Excel文件 (*.xlsx)"
        )
        if filepath:
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                
                headers = ["姓名", "关键模块", "工号", "电话", "邮箱", "部门"]
                ws.append(headers)
                
                recommendations = RecommendationService.get_all_recommendations()
                for rec in recommendations:
                    ws.append([
                        rec.get('name', ''),
                        rec.get('key_module', ''),
                        rec.get('employee_id', ''),
                        rec.get('phone', ''),
                        rec.get('email', ''),
                        rec.get('department', '')
                    ])
                
                wb.save(filepath)
                QMessageBox.information(self, "成功", f"推荐库已导出到:\n{filepath}")
            except Exception as e:
                QMessageBox.critical(self, "失败", f"导出失败: {e}")
    
    def _on_add_contact(self):
        """添加联系人"""
        dialog = ContactDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            ContactsService.add_contact(data)
            self._load_contacts()
            self.statusbar.showMessage("联系人已添加")
    
    def _on_export_contacts(self):
        """导出通讯录"""
        filepath, _ = QFileDialog.getSaveFileName(
            self, "导出通讯录", ExportService.generate_filename("通讯录", "xlsx"),
            "Excel文件 (*.xlsx)"
        )
        if filepath:
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                
                headers = ["姓名", "工号", "电话", "邮箱", "部门", "职位", "备注"]
                ws.append(headers)
                
                contacts = ContactsService.get_all_contacts()
                for contact in contacts:
                    ws.append([
                        contact.get('name', ''),
                        contact.get('employee_id', ''),
                        contact.get('phone', ''),
                        contact.get('email', ''),
                        contact.get('department', ''),
                        contact.get('title', ''),
                        contact.get('remark', '')
                    ])
                
                wb.save(filepath)
                QMessageBox.information(self, "成功", f"通讯录已导出到:\n{filepath}")
            except Exception as e:
                QMessageBox.critical(self, "失败", f"导出失败: {e}")
    
    def _on_contacts_search(self, keyword: str):
        """搜索通讯录"""
        if keyword:
            contacts = ContactsService.search_contacts(keyword)
        else:
            contacts = ContactsService.get_all_contacts()
        
        self.contacts_table.setRowCount(len(contacts))
        for row, contact in enumerate(contacts):
            self.contacts_table.setItem(row, 0, QTableWidgetItem(str(contact.get('id', ''))))
            self.contacts_table.setItem(row, 1, QTableWidgetItem(contact.get('name', '')))
            self.contacts_table.setItem(row, 2, QTableWidgetItem(contact.get('employee_id', '')))
            self.contacts_table.setItem(row, 3, QTableWidgetItem(contact.get('phone', '')))
            self.contacts_table.setItem(row, 4, QTableWidgetItem(contact.get('email', '')))
            self.contacts_table.setItem(row, 5, QTableWidgetItem(contact.get('department', '')))
            self.contacts_table.setItem(row, 6, QTableWidgetItem(contact.get('title', '')))
    
    def _on_about(self):
        """关于"""
        QMessageBox.about(
            self, "关于",
            "市场咨询任务跟踪工具 V4.6\n\n"
            "一个高效的市场咨询任务管理和跟踪系统。\n\n"
            "功能特性:\n"
            "• 任务信息管理\n"
            "• 任务跟踪\n"
            "• 智能推荐\n"
            "• 统计分析\n"
            "• 数据导入导出\n\n"
            "© 2026 市场咨询团队"
        )

# ==================== 任务对话框 ====================
class TaskDialog(QDialog):
    """任务对话框"""
    
    def __init__(self, parent=None, task: Dict = None):
        super().__init__(parent)
        self.task = task
        self.setWindowTitle("编辑任务" if task else "新建任务")
        self.setMinimumWidth(500)
        self._init_ui()
        
        if task:
            self._load_task(task)
    
    def _init_ui(self):
        """初始化UI"""
        layout = QFormLayout(self)
        
        # 任务名称
        self.task_name = QLineEdit()
        self.task_name.setPlaceholderText("请输入任务名称")
        layout.addRow("任务名称 *:", self.task_name)
        
        # 咨询者
        self.consultant = QLineEdit()
        self.consultant.setPlaceholderText("请输入咨询者姓名")
        layout.addRow("咨询者:", self.consultant)
        
        # 答复人
        self.responder = QLineEdit()
        self.responder.setPlaceholderText("请输入答复人姓名")
        layout.addRow("答复人:", self.responder)
        
        # 部门
        self.department = QComboBox()
        self.department.addItems(["", "技术部", "运维部", "市场部", "客服部", "其他"])
        layout.addRow("部门:", self.department)
        
        # 关键模块
        self.key_module = QComboBox()
        self.key_module.addItems([
            "", "MAC认证", "802.1x认证", "Portal认证", "端口安全", 
            "AAA认证", "交换机配置", "网络故障", "其他"
        ])
        self.key_module.setEditable(True)
        layout.addRow("关键模块:", self.key_module)
        
        # 产品型号
        self.product_model = QLineEdit()
        self.product_model.setPlaceholderText("请输入产品型号")
        layout.addRow("产品型号:", self.product_model)
        
        # 状态
        self.status = QComboBox()
        self.status.addItems(["待处理", "处理中", "已完成", "已取消"])
        layout.addRow("状态:", self.status)
        
        # 重要程度
        self.importance = QComboBox()
        self.importance.addItems(["低", "中", "高"])
        layout.addRow("重要程度:", self.importance)
        
        # 截止日期
        self.due_date = QDateTimeEdit()
        self.due_date.setCalendarPopup(True)
        self.due_date.setDateTime(QDateTime.currentDateTime().addDays(7))
        layout.addRow("截止日期:", self.due_date)
        
        # 描述
        self.description = QTextEdit()
        self.description.setPlaceholderText("请输入任务描述...")
        self.description.setMaximumHeight(100)
        layout.addRow("描述:", self.description)
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addRow(btn_layout)
    
    def _load_task(self, task: Dict):
        """加载任务数据"""
        self.task_name.setText(task.get('task_name', ''))
        self.consultant.setText(task.get('consultant', ''))
        self.responder.setText(task.get('responder', ''))
        
        dept = task.get('department', '')
        index = self.department.findText(dept)
        if index >= 0:
            self.department.setCurrentIndex(index)
        
        module = task.get('key_module', '')
        index = self.key_module.findText(module)
        if index >= 0:
            self.key_module.setCurrentIndex(index)
        
        self.product_model.setText(task.get('product_model', ''))
        
        status = task.get('status', '待处理')
        index = self.status.findText(status)
        if index >= 0:
            self.status.setCurrentIndex(index)
        
        importance = task.get('importance', '中')
        index = self.importance.findText(importance)
        if index >= 0:
            self.importance.setCurrentIndex(index)
        
        self.description.setText(task.get('description', ''))
    
    def get_task_data(self) -> Dict:
        """获取任务数据"""
        return {
            'task_name': self.task_name.text().strip(),
            'consultant': self.consultant.text().strip(),
            'responder': self.responder.text().strip(),
            'department': self.department.currentText(),
            'key_module': self.key_module.currentText(),
            'product_model': self.product_model.text().strip(),
            'status': self.status.currentText(),
            'importance': self.importance.currentText(),
            'description': self.description.toPlainText().strip(),
            'due_date': self.due_date.dateTime().toString("yyyy-MM-dd HH:mm")
        }

# ==================== 推荐对话框 ====================
class RecommendationDialog(QDialog):
    """推荐对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("添加专家")
        self.setMinimumWidth(400)
        self._init_ui()
    
    def _init_ui(self):
        """初始化UI"""
        layout = QFormLayout(self)
        
        self.name = QLineEdit()
        self.name.setPlaceholderText("请输入姓名")
        layout.addRow("姓名 *:", self.name)
        
        self.key_module = QComboBox()
        self.key_module.addItems([
            "MAC认证", "802.1x认证", "Portal认证", "端口安全", 
            "AAA认证", "交换机配置", "网络故障", "其他"
        ])
        self.key_module.setEditable(True)
        layout.addRow("关键模块:", self.key_module)
        
        self.employee_id = QLineEdit()
        self.employee_id.setPlaceholderText("请输入工号")
        layout.addRow("工号:", self.employee_id)
        
        self.phone = QLineEdit()
        self.phone.setPlaceholderText("请输入电话")
        layout.addRow("电话:", self.phone)
        
        self.email = QLineEdit()
        self.email.setPlaceholderText("请输入邮箱")
        layout.addRow("邮箱:", self.email)
        
        self.department = QComboBox()
        self.department.addItems(["技术部", "运维部", "市场部", "客服部", "其他"])
        layout.addRow("部门:", self.department)
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addRow(btn_layout)
    
    def get_data(self) -> Dict:
        """获取数据"""
        return {
            'name': self.name.text().strip(),
            'key_module': self.key_module.currentText(),
            'employee_id': self.employee_id.text().strip(),
            'phone': self.phone.text().strip(),
            'email': self.email.text().strip(),
            'department': self.department.currentText()
        }

# ==================== 联系人对话框 ====================
class ContactDialog(QDialog):
    """联系人对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("添加联系人")
        self.setMinimumWidth(400)
        self._init_ui()
    
    def _init_ui(self):
        """初始化UI"""
        layout = QFormLayout(self)
        
        self.name = QLineEdit()
        self.name.setPlaceholderText("请输入姓名")
        layout.addRow("姓名 *:", self.name)
        
        self.employee_id = QLineEdit()
        self.employee_id.setPlaceholderText("请输入工号")
        layout.addRow("工号:", self.employee_id)
        
        self.phone = QLineEdit()
        self.phone.setPlaceholderText("请输入电话")
        layout.addRow("电话:", self.phone)
        
        self.email = QLineEdit()
        self.email.setPlaceholderText("请输入邮箱")
        layout.addRow("邮箱:", self.email)
        
        self.department = QComboBox()
        self.department.addItems(["技术部", "运维部", "市场部", "客服部", "其他"])
        layout.addRow("部门:", self.department)
        
        self.title = QLineEdit()
        self.title.setPlaceholderText("请输入职位")
        layout.addRow("职位:", self.title)
        
        self.remark = QTextEdit()
        self.remark.setMaximumHeight(60)
        self.remark.setPlaceholderText("备注信息...")
        layout.addRow("备注:", self.remark)
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addRow(btn_layout)
    
    def get_data(self) -> Dict:
        """获取数据"""
        return {
            'name': self.name.text().strip(),
            'employee_id': self.employee_id.text().strip(),
            'phone': self.phone.text().strip(),
            'email': self.email.text().strip(),
            'department': self.department.currentText(),
            'title': self.title.text().strip(),
            'remark': self.remark.toPlainText().strip()
        }

# ==================== 主程序入口 ====================
def main():
    """主程序入口"""
    app = QApplication(sys.argv)
    app.setApplicationName("市场咨询任务跟踪工具")
    app.setApplicationVersion("4.6")
    app.setStyle("Fusion")
    
    # 设置中文
    font = QFont()
    font.setFamily("Microsoft YaHei")
    font.setPointSize(10)
    app.setFont(font)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
